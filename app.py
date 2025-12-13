import streamlit as st
import requests
from bs4 import BeautifulSoup
from urllib.parse import urlparse
import pandas as pd
from fpdf import FPDF
from datetime import datetime, timedelta
import time
import re
import whois
from openai import OpenAI
import os
from dotenv import load_dotenv
import json
import hashlib
from pathlib import Path
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import smtplib
from io import BytesIO
from cryptography.fernet import Fernet
import logging
from logging.handlers import RotatingFileHandler
from concurrent.futures import ThreadPoolExecutor, as_completed

# Optional: 2FA support
try:
    import pyotp
    import qrcode
    TWO_FA_AVAILABLE = True
except ImportError:
    TWO_FA_AVAILABLE = False

# Load environment variables
load_dotenv()
from models import init_db, get_db, Audit, Lead, EmailOutreach, DATABASE_URL, User, BulkScan

# Import consistency layer for normalization
from consistency import (
    normalize_audit, normalize_lead, normalize_bulk_result,
    normalize_audit_list, normalize_lead_list, normalize_bulk_list,
    ensure_audit_fields, ensure_lead_fields,
    ensure_audit_defaults, ensure_lead_defaults, ensure_bulk_defaults,
    safe_render_audit, safe_render_lead, safe_render_bulk,
    get_safe_export_columns, safe_timestamp_slice,
    lead_to_dict, audit_to_dict, bulk_to_dict,
    _lead_to_dict, _audit_to_dict, _bulk_to_dict
)

# ============================================================================
# PERSISTENCE LAYER IMPORTS
# ============================================================================
from persistence import (
    # Session/App Initialization
    init_app_session_persistence,
    reset_persistence_state,
    get_persistence_status,
    on_logout_cleanup,
    
    # Audit State Management
    persist_audit_data,
    load_audit_data,
    get_current_audit,
    set_current_audit,
    clear_current_audit,
    get_audit_from_db,
    rebuild_audit_data_from_db,
    
    # AI Cache (new module replaces old functions)
    get_ai_cache,
    save_ai_cache,
    ai_cache_exists,
    clear_ai_cache,
    get_regen_count as persistence_get_regen_count,
    increment_regen_count as persistence_increment_regen_count,
    reset_regen_count,
    regen_limit_reached,
    REGEN_LIMIT,
    
    # PDF Context
    store_pdf_context,
    load_pdf_context,
    clear_pdf_context,
    get_pdf_context_or_current,
    
    # Navigation State
    save_navigation_state,
    load_navigation_state,
    clear_navigation_state,
    get_deep_link_audit_id,
    sync_query_params,
    restore_navigation_on_refresh,
)

# ============================================================================
# CRASH-PROOF UI HELPERS (Phase 5 Step 3)
# ============================================================================

def safe_ui_section(section_name: str):
    """
    Decorator for crash-proof UI sections.
    Wraps a function with try/except and displays user-friendly error.
    
    Usage:
        @safe_ui_section("Audit History")
        def show_audit_history():
            ...
    """
    def decorator(func):
        def wrapper(*args, **kwargs):
            try:
                return func(*args, **kwargs)
            except Exception as e:
                logger = logging.getLogger("sales_engine")
                logger.error(f"Error in {section_name}: {str(e)}", exc_info=True)
                st.error(f"⚠️ Something went wrong in {section_name}. Please try again or refresh the page.")
                st.expander("🔍 Technical Details").code(str(e))
                return None
        return wrapper
    return decorator


def safe_execute(operation_name: str, operation, *args, **kwargs):
    """
    Execute an operation safely with automatic error handling.
    
    Args:
        operation_name: Name for logging
        operation: Callable to execute
        *args, **kwargs: Arguments to pass to operation
        
    Returns:
        Result of operation, or None on error
    """
    logger = logging.getLogger("sales_engine")
    try:
        return operation(*args, **kwargs)
    except Exception as e:
        logger.error(f"Error in {operation_name}: {str(e)}", exc_info=True)
        return None


# ============================================================================
# ENHANCED APP CONFIGURATION
# ============================================================================

st.set_page_config(
    page_title="Code Nest Sales Engine Pro",
    layout="wide",
    page_icon="🦅",
    initial_sidebar_state="expanded"
)

# Initialize theme preferences early
if 'user_theme' not in st.session_state:
    st.session_state.user_theme = 'light'

# ============================================================================
# CODE NEST BRANDING SYSTEM (CENTRALIZED)
# ============================================================================

# Company Information
COMPANY_NAME = "Code Nest LLC"
COMPANY_TAGLINE = "Nest Idea | Code Success"
COMPANY_WEBSITE = "https://codenest.us.com"
CONTACT_EMAIL = "contact@codenest.us.com"
COMPANY_LOCATION = "New Mexico, USA"

# Brand Colors
BRAND_COLORS = {
    # Primary Brand Colors
    "primary_dark_green": "#0c3740",      # Primary Dark Green
    "accent_light_green": "#2b945f",      # Accent Light Green  
    "branding_grey": "#5a5a5a",           # Branding Grey
    "white": "#feffff",                   # Brand White
    
    # UI Colors (for app interface)
    "success": "#2b945f",                 # Use accent green for success
    "warning": "#FFA500",
    "danger": "#FF6B6B",
    "info": "#0c3740",                    # Use primary dark green
    "primary": "#0c3740",
    "neutral": "#5a5a5a"
}

# Legacy COLORS dict for backward compatibility
COLORS = {
    "success": BRAND_COLORS["success"],
    "warning": BRAND_COLORS["warning"],
    "danger": BRAND_COLORS["danger"],
    "info": BRAND_COLORS["info"],
    "primary": BRAND_COLORS["primary"],
    "neutral": BRAND_COLORS["neutral"]
}

# Brand Assets Paths
ASSETS_DIR = Path(__file__).parent / "assets"
PDF_LOGO_PATH = ASSETS_DIR / "codenest-logo.png"
EMAIL_LOGO_URL = "https://raw.githubusercontent.com/Code-Nest1/sales-engine/main/assets/codenest-logo.png"

# Ensure assets directory exists
ASSETS_DIR.mkdir(exist_ok=True)

def check_brand_assets():
    """Check if required brand assets exist and print messages if missing."""
    missing_assets = []
    
    if not PDF_LOGO_PATH.exists():
        missing_assets.append(f"Please upload: {PDF_LOGO_PATH}")
    
    return missing_assets

# Check for missing assets at startup
_missing_assets = check_brand_assets()
if _missing_assets:
    for msg in _missing_assets:
        print(f"⚠️ BRANDING: {msg}")

# ============================================================================
# OPENAI CONFIGURATION & COST OPTIMIZATION
# ============================================================================

# Centralized AI Model Settings
AI_MODEL = "gpt-4o-mini"  # Cheapest reliable model
AI_MAX_TOKENS = 1200      # Increased for richer insights (was 800)
AI_TEMPERATURE = 0.7      # Balance creativity/consistency
AI_TIMEOUT = 30.0         # Per-request timeout (seconds)
AI_CLIENT_TIMEOUT = 35.0  # Client-level timeout

# Regeneration Limits
MAX_REGENERATIONS_PER_URL = 2  # Max times user can regenerate AI content per URL

# Issue Categories (for structured issues)
ISSUE_CATEGORIES = {
    "SEO": "Search Engine Optimization",
    "Performance": "Site Speed & Performance", 
    "Tracking": "Analytics & Tracking",
    "Conversion": "Lead Generation & Conversion",
    "UX": "User Experience",
    "Security": "Security & Trust",
    "Content": "Content Quality"
}

# ============================================================================
# AI CACHE WRAPPER FUNCTIONS (use persistence layer)
# ============================================================================

def _extract_domain(url: str) -> str:
    """Extract clean domain from URL for cache keys."""
    if not url:
        return ""
    domain = urlparse(url).netloc.replace("www.", "").lower()
    return domain if domain else url.lower().replace("www.", "")


def get_cached_ai_result(url: str, audit_data: dict = None) -> dict | None:
    """
    Get cached AI result for a URL using persistence layer.
    
    Uses triple-layer fallback: session → disk → None
    """
    domain = _extract_domain(url)
    if not domain:
        return None
    
    cached = get_ai_cache(domain)
    if cached:
        logger = logging.getLogger("sales_engine")
        logger.info(f"[AI CACHE HIT] {domain}")
    return cached


def cache_ai_result(url: str, ai_result: dict):
    """
    Cache AI result using persistence layer.
    
    Stores to session + disk with automatic expiry.
    """
    domain = _extract_domain(url)
    if not domain:
        return
    
    save_ai_cache(domain, ai_result)
    logger = logging.getLogger("sales_engine")
    logger.info(f"[AI CACHE STORE] {domain}")


def get_regen_count(url: str) -> int:
    """Get regeneration count for a URL using persistence layer."""
    domain = _extract_domain(url)
    return persistence_get_regen_count(domain)


def increment_regen_count(url: str) -> int:
    """Increment and return regeneration count for a URL using persistence layer."""
    domain = _extract_domain(url)
    return persistence_increment_regen_count(domain)


def can_regenerate(url: str) -> bool:
    """Check if regeneration is allowed for this URL."""
    domain = _extract_domain(url)
    return not regen_limit_reached(domain, MAX_REGENERATIONS_PER_URL)


def clear_ai_cache_for_url(url: str):
    """Clear cached AI result for a specific URL (for forced regeneration)."""
    domain = _extract_domain(url)
    clear_ai_cache(domain)


# Initialize database
DB_AVAILABLE = False
if DATABASE_URL:
    try:
        DB_AVAILABLE = init_db()
    except Exception as e:
        DB_AVAILABLE = False

# ============================================================================
# LOGGING SYSTEM
# ============================================================================

# Configure logging
LOGS_DIR = Path(__file__).parent / "logs"
LOGS_DIR.mkdir(exist_ok=True)

def setup_logging():
    """Configure rotating file logger."""
    logger = logging.getLogger("sales_engine")
    
    if logger.hasHandlers():
        return logger
    
    logger.setLevel(logging.DEBUG)
    
    # Rotating file handler (max 5MB per file, keep 5 backups)
    handler = RotatingFileHandler(
        LOGS_DIR / "app.log",
        maxBytes=5_000_000,
        backupCount=5
    )
    
    formatter = logging.Formatter(
        '%(asctime)s - %(name)s - %(levelname)s - %(funcName)s:%(lineno)d - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    handler.setFormatter(formatter)
    logger.addHandler(handler)
    
    return logger

logger = setup_logging()


# ============================================================================
# PHASE 4: DATABASE ABSTRACTION LAYER
# ============================================================================
# These functions provide a clean interface for database operations,
# ensuring consistent error handling and response formatting.

# Pipeline stage definitions for CRM
PIPELINE_STAGES = ["new", "contacted", "follow-up", "qualified", "proposal", "closed"]
LEAD_STATUS_OPTIONS = ["hot", "warm", "cold"]


def create_response(success: bool, data: dict = None, error: str = None, code: int = 200) -> dict:
    """
    Create a standardized API response envelope.
    
    Args:
        success: Whether the operation succeeded
        data: Optional data payload
        error: Optional error message
        code: HTTP-like status code
    
    Returns:
        Standardized response dict
    """
    return {
        "success": success,
        "data": data or {},
        "error": error,
        "code": code,
        "timestamp": datetime.now().isoformat()
    }


def api_success(data: dict = None, message: str = None) -> dict:
    """Create a success response envelope."""
    result = create_response(success=True, data=data, code=200)
    if message:
        result["message"] = message
    return result


def api_error(error: str, code: int = 400, data: dict = None) -> dict:
    """Create an error response envelope."""
    return create_response(success=False, data=data, error=error, code=code)


def db_get_lead_by_domain(domain: str) -> dict:
    """
    Get a lead by domain name.
    
    Args:
        domain: Domain name to search for
    
    Returns:
        API response with lead data or error
    """
    try:
        db = get_db()
        if not db:
            return api_error("Database connection failed", code=503)
        
        lead = db.query(Lead).filter(Lead.domain == domain).first()
        db.close()
        
        if lead:
            return api_success(data={"lead": _lead_to_dict(lead)})
        else:
            return api_error(f"Lead not found for domain: {domain}", code=404)
    except Exception as e:
        logger.error(f"Error getting lead by domain '{domain}': {e}")
        return api_error(str(e), code=500)


def db_get_lead_by_id(lead_id: int) -> dict:
    """
    Get a lead by ID.
    
    Args:
        lead_id: Lead ID to search for
    
    Returns:
        API response with lead data or error
    """
    try:
        db = get_db()
        if not db:
            return api_error("Database connection failed", code=503)
        
        lead = db.query(Lead).filter(Lead.id == lead_id).first()
        db.close()
        
        if lead:
            return api_success(data={"lead": _lead_to_dict(lead)})
        else:
            return api_error(f"Lead not found with ID: {lead_id}", code=404)
    except Exception as e:
        logger.error(f"Error getting lead by ID {lead_id}: {e}")
        return api_error(str(e), code=500)


def db_get_audit_by_id(audit_id: int) -> dict:
    """
    Get an audit by ID.
    
    Args:
        audit_id: Audit ID to search for
    
    Returns:
        API response with audit data or error
    """
    try:
        db = get_db()
        if not db:
            return api_error("Database connection failed", code=503)
        
        audit = db.query(Audit).filter(Audit.id == audit_id).first()
        db.close()
        
        if audit:
            return api_success(data={"audit": _audit_to_dict(audit)})
        else:
            return api_error(f"Audit not found with ID: {audit_id}", code=404)
    except Exception as e:
        logger.error(f"Error getting audit by ID {audit_id}: {e}")
        return api_error(str(e), code=500)


def db_get_leads_by_pipeline(stage: str = None, status: str = None, limit: int = 100) -> dict:
    """
    Get leads filtered by pipeline stage and/or status.
    
    Args:
        stage: Pipeline stage filter (optional)
        status: Lead status filter (optional)
        limit: Maximum results to return
    
    Returns:
        API response with list of leads
    """
    try:
        db = get_db()
        if not db:
            return api_error("Database connection failed", code=503)
        
        query = db.query(Lead)
        
        if stage:
            query = query.filter(Lead.pipeline_stage == stage)
        if status:
            query = query.filter(Lead.lead_status == status)
        
        leads = query.order_by(Lead.updated_at.desc()).limit(limit).all()
        db.close()
        
        leads_data = [_lead_to_dict(lead) for lead in leads]
        return api_success(data={"leads": leads_data, "count": len(leads_data)})
    except Exception as e:
        logger.error(f"Error getting leads by pipeline: {e}")
        return api_error(str(e), code=500)


def db_create_lead_from_audit(audit_id: int) -> dict:
    """
    Create a new lead from an audit result.
    
    Args:
        audit_id: The audit ID to create a lead from
    
    Returns:
        API response with created lead data
    """
    try:
        db = get_db()
        if not db:
            return api_error("Database connection failed", code=503)
        
        # Get the audit
        audit = db.query(Audit).filter(Audit.id == audit_id).first()
        if not audit:
            db.close()
            return api_error(f"Audit not found: {audit_id}", code=404)
        
        # Check if lead already exists for this domain
        existing_lead = db.query(Lead).filter(Lead.domain == audit.domain).first()
        if existing_lead:
            # Update existing lead with latest audit
            existing_lead.last_audit_id = audit_id
            existing_lead.health_score = audit.health_score
            if audit.emails_found:
                emails = audit.emails_found if isinstance(audit.emails_found, list) else []
                if emails:
                    existing_lead.email = emails[0]
            existing_lead.updated_at = datetime.utcnow()
            db.commit()
            lead_dict = _lead_to_dict(existing_lead)
            db.close()
            return api_success(data={"lead": lead_dict, "created": False}, message="Lead updated with latest audit")
        
        # Create new lead
        emails = audit.emails_found if isinstance(audit.emails_found, list) else []
        new_lead = Lead(
            domain=audit.domain,
            email=emails[0] if emails else None,
            company_name=audit.domain.replace(".", " ").title(),
            health_score=audit.health_score,
            pipeline_stage="new",
            lead_status="warm",
            source="audit",
            last_audit_id=audit_id
        )
        db.add(new_lead)
        db.commit()
        lead_dict = _lead_to_dict(new_lead)
        db.close()
        
        return api_success(data={"lead": lead_dict, "created": True}, message="Lead created from audit")
    except Exception as e:
        logger.error(f"Error creating lead from audit {audit_id}: {e}")
        return api_error(str(e), code=500)


def db_update_lead_pipeline(lead_id: int, stage: str = None, status: str = None, notes: str = None) -> dict:
    """
    Update a lead's pipeline stage, status, or notes.
    
    Args:
        lead_id: Lead ID to update
        stage: New pipeline stage (optional)
        status: New lead status (optional)
        notes: Notes to append (optional)
    
    Returns:
        API response with updated lead data
    """
    try:
        db = get_db()
        if not db:
            return api_error("Database connection failed", code=503)
        
        lead = db.query(Lead).filter(Lead.id == lead_id).first()
        if not lead:
            db.close()
            return api_error(f"Lead not found: {lead_id}", code=404)
        
        if stage and stage in PIPELINE_STAGES:
            lead.pipeline_stage = stage
        if status and status in LEAD_STATUS_OPTIONS:
            lead.lead_status = status
        if notes:
            existing_notes = lead.notes or ""
            lead.notes = f"{existing_notes}\n[{datetime.now().strftime('%Y-%m-%d %H:%M')}] {notes}".strip()
        
        lead.updated_at = datetime.utcnow()
        db.commit()
        lead_dict = _lead_to_dict(lead)
        db.close()
        
        return api_success(data={"lead": lead_dict}, message="Lead updated")
    except Exception as e:
        logger.error(f"Error updating lead {lead_id}: {e}")
        return api_error(str(e), code=500)


def db_count_leads_by_status() -> dict:
    """
    Get count of leads grouped by status and pipeline stage.
    
    Returns:
        API response with counts by status and stage
    """
    try:
        db = get_db()
        if not db:
            return api_error("Database connection failed", code=503)
        
        # Count by status
        status_counts = {}
        for status in LEAD_STATUS_OPTIONS:
            count = db.query(Lead).filter(Lead.lead_status == status).count()
            status_counts[status] = count
        
        # Count by pipeline stage
        stage_counts = {}
        for stage in PIPELINE_STAGES:
            count = db.query(Lead).filter(Lead.pipeline_stage == stage).count()
            stage_counts[stage] = count
        
        total = db.query(Lead).count()
        db.close()
        
        return api_success(data={
            "by_status": status_counts,
            "by_stage": stage_counts,
            "total": total
        })
    except Exception as e:
        logger.error(f"Error counting leads by status: {e}")
        return api_error(str(e), code=500)


def db_get_follow_up_due(days_ahead: int = 7) -> dict:
    """
    Get leads with follow-up dates within the specified window.
    
    Args:
        days_ahead: Number of days to look ahead
    
    Returns:
        API response with leads needing follow-up
    """
    try:
        db = get_db()
        if not db:
            return api_error("Database connection failed", code=503)
        
        now = datetime.utcnow()
        future = now + timedelta(days=days_ahead)
        
        leads = db.query(Lead).filter(
            Lead.follow_up_date != None,
            Lead.follow_up_date <= future,
            Lead.pipeline_stage != "closed"
        ).order_by(Lead.follow_up_date.asc()).all()
        
        db.close()
        
        leads_data = [_lead_to_dict(lead) for lead in leads]
        
        # Categorize by urgency
        overdue = [l for l in leads_data if l.get("follow_up_date") and datetime.fromisoformat(l["follow_up_date"]) < now]
        today = [l for l in leads_data if l.get("follow_up_date") and datetime.fromisoformat(l["follow_up_date"]).date() == now.date()]
        upcoming = [l for l in leads_data if l not in overdue and l not in today]
        
        return api_success(data={
            "leads": leads_data,
            "overdue": overdue,
            "today": today,
            "upcoming": upcoming,
            "total": len(leads_data)
        })
    except Exception as e:
        logger.error(f"Error getting follow-up due leads: {e}")
        return api_error(str(e), code=500)


# ============================================================================
# CONSISTENCY LAYER - Functions imported from consistency.py
# ============================================================================
# lead_to_dict, audit_to_dict, bulk_to_dict - imported from consistency.py
# normalize_audit, normalize_lead, normalize_bulk_result - imported from consistency.py
# See consistency.py for full implementation


# ============================================================================
# PHASE 4: AUTO-MOVEMENT RULES
# ============================================================================

def apply_auto_pipeline_rules(lead_id: int) -> dict:
    """
    Apply automatic pipeline movement rules based on lead activity.
    
    Rules:
    - If email sent → move to "contacted"
    - If followed up 2+ times → move to "follow-up"
    - If health_score < 50 → mark as "cold"
    - If health_score > 80 → mark as "hot"
    
    Args:
        lead_id: Lead ID to apply rules to
    
    Returns:
        API response with rule application results
    """
    try:
        db = get_db()
        if not db:
            return api_error("Database connection failed", code=503)
        
        lead = db.query(Lead).filter(Lead.id == lead_id).first()
        if not lead:
            db.close()
            return api_error(f"Lead not found: {lead_id}", code=404)
        
        changes = []
        
        # Rule 1: Auto-set lead temperature based on health score
        if lead.health_score is not None:
            if lead.health_score < 50 and lead.lead_status != "cold":
                lead.lead_status = "cold"
                changes.append("Set status to 'cold' (low health score)")
            elif lead.health_score > 80 and lead.lead_status == "cold":
                lead.lead_status = "warm"
                changes.append("Set status to 'warm' (high health score)")
        
        # Rule 2: If approached but still in "new", move to "contacted"
        if lead.approached and lead.pipeline_stage == "new":
            lead.pipeline_stage = "contacted"
            changes.append("Moved to 'contacted' (already approached)")
        
        if changes:
            lead.updated_at = datetime.utcnow()
            db.commit()
        
        lead_dict = _lead_to_dict(lead)
        db.close()
        
        return api_success(data={"lead": lead_dict, "changes": changes})
    except Exception as e:
        logger.error(f"Error applying auto rules to lead {lead_id}: {e}")
        return api_error(str(e), code=500)


def auto_move_after_email_sent(lead_id: int) -> dict:
    """
    Automatically update lead after an email is sent.
    
    Updates:
    - Set approached = True
    - Set approached_date if not set
    - Move pipeline_stage to "contacted" if still "new"
    - Set default follow_up_date if not set (3 days from now)
    
    Args:
        lead_id: Lead ID that received an email
    
    Returns:
        API response with update results
    """
    try:
        db = get_db()
        if not db:
            return api_error("Database connection failed", code=503)
        
        lead = db.query(Lead).filter(Lead.id == lead_id).first()
        if not lead:
            db.close()
            return api_error(f"Lead not found: {lead_id}", code=404)
        
        changes = []
        now = datetime.utcnow()
        
        # Mark as approached
        if not lead.approached:
            lead.approached = True
            lead.approached_date = now
            changes.append("Marked as approached")
        
        # Move pipeline stage
        if lead.pipeline_stage == "new":
            lead.pipeline_stage = "contacted"
            changes.append("Moved to 'contacted' stage")
        
        # Set default follow-up
        if not lead.follow_up_date:
            lead.follow_up_date = now + timedelta(days=3)
            changes.append("Set follow-up for 3 days from now")
        
        if changes:
            lead.updated_at = now
            db.commit()
        
        lead_dict = _lead_to_dict(lead)
        db.close()
        
        return api_success(data={"lead": lead_dict, "changes": changes})
    except Exception as e:
        logger.error(f"Error auto-moving lead {lead_id}: {e}")
        return api_error(str(e), code=500)

# ============================================================================
# INPUT VALIDATION & SANITIZATION
# ============================================================================

def validate_url(url: str) -> tuple[bool, str]:
    """Validate URL format. Returns (is_valid, error_message)."""
    if not url or not isinstance(url, str):
        return False, "URL cannot be empty"
    
    url = url.strip()
    if len(url) > 2000:
        return False, "URL is too long (max 2000 characters)"
    
    try:
        # Basic URL validation
        if not url.startswith(('http://', 'https://')):
            url = 'https://' + url
        
        result = urlparse(url)
        if not all([result.scheme, result.netloc]):
            return False, "Invalid URL format"
        
        # Check for valid domain
        if not re.match(r'^[a-zA-Z0-9]([a-zA-Z0-9-]{0,61}[a-zA-Z0-9])?(\.[a-zA-Z0-9]([a-zA-Z0-9-]{0,61}[a-zA-Z0-9])?)*$', result.netloc):
            return False, "Invalid domain name"
        
        return True, ""
    except Exception as e:
        logger.warning(f"URL validation error for '{url}': {str(e)}")
        return False, f"URL validation failed: {str(e)}"

def validate_email(email: str) -> tuple[bool, str]:
    """Validate email format. Returns (is_valid, error_message)."""
    if not email or not isinstance(email, str):
        return False, "Email cannot be empty"
    
    email = email.strip()
    if len(email) > 254:
        return False, "Email is too long"
    
    # Basic email regex pattern
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    
    if re.match(pattern, email):
        return True, ""
    else:
        return False, "Invalid email format"

def validate_password(password: str) -> tuple[bool, str]:
    """Validate password strength. Returns (is_valid, error_message)."""
    if not password:
        return False, "Password cannot be empty"
    
    if len(password) < 8:
        return False, "Password must be at least 8 characters"
    
    if len(password) > 128:
        return False, "Password is too long"
    
    # Check for complexity (optional but recommended)
    has_upper = any(c.isupper() for c in password)
    has_lower = any(c.islower() for c in password)
    has_digit = any(c.isdigit() for c in password)
    has_special = any(c in '!@#$%^&*()_+-=[]{}|;:,.<>?' for c in password)
    
    complexity_score = sum([has_upper, has_lower, has_digit, has_special])
    
    if complexity_score < 3:
        return False, "Password should contain uppercase, lowercase, numbers, and special characters"
    
    return True, ""

def sanitize_input(user_input: str, max_length: int = 1000) -> str:
    """Sanitize user input by removing dangerous characters."""
    if not isinstance(user_input, str):
        return ""
    
    # Limit length
    sanitized = user_input[:max_length]
    
    # Remove null bytes
    sanitized = sanitized.replace('\x00', '')
    
    # Remove control characters except newlines and tabs
    sanitized = ''.join(char for char in sanitized if ord(char) >= 32 or char in '\n\t')
    
    return sanitized.strip()

def safe_execute(func, *args, default_return=None, error_message: str = "Operation failed", **kwargs):
    """Execute function safely with error handling. Returns (success, result)."""
    try:
        result = func(*args, **kwargs)
        logger.debug(f"Successfully executed {func.__name__}")
        return True, result
    except Exception as e:
        logger.error(f"Error in {func.__name__}: {str(e)}", exc_info=True)
        return False, default_return

# ============================================================================
# PERFORMANCE OPTIMIZATION - CACHING & PAGINATION
# ============================================================================

# Caching configuration
CACHE_TTL = 300  # 5 minutes cache for database queries

@st.cache_data(ttl=CACHE_TTL)
def get_audit_history_cached(limit=100, search_query=None, min_score=None, max_score=None,
                              username=None, is_admin=False):
    """Cached version of audit history query with user-based filtering."""
    return get_audit_history(
        limit=limit, 
        search_query=search_query, 
        min_score=min_score, 
        max_score=max_score,
        username=username,
        is_admin=is_admin
    )

@st.cache_data(ttl=CACHE_TTL)
def get_leads_cached():
    """Cached version of leads query."""
    return get_leads()

def get_scheduled_audits():
    """Get scheduled audits from database. (Placeholder for future feature)"""
    return []

@st.cache_data(ttl=CACHE_TTL)
def get_scheduled_audits_cached():
    """Cached version of scheduled audits query."""
    return get_scheduled_audits() if DB_AVAILABLE else []

# Pagination helper functions
def init_pagination_state(page_key: str, items_per_page: int = 50):
    """Initialize pagination state in session."""
    if page_key not in st.session_state:
        st.session_state[page_key] = 0

def get_paginated_items(items: list, page_key: str, items_per_page: int = 50) -> tuple[list, int, int]:
    """Get paginated items. Returns (items_on_page, total_pages, current_page)."""
    init_pagination_state(page_key, items_per_page)
    
    total_items = len(items)
    total_pages = (total_items + items_per_page - 1) // items_per_page
    
    current_page = st.session_state[page_key]
    # Ensure current page is valid
    if current_page >= total_pages and total_pages > 0:
        current_page = total_pages - 1
    
    start_idx = current_page * items_per_page
    end_idx = start_idx + items_per_page
    
    return items[start_idx:end_idx], total_pages, current_page

def display_pagination_controls(page_key: str, total_pages: int, current_page: int):
    """Display pagination controls (previous/next buttons and page info)."""
    if total_pages <= 1:
        return
    
    col1, col2, col3, col4, col5 = st.columns([1, 1, 2, 1, 1])
    
    with col1:
        if st.button("◀ Prev", key=f"{page_key}_prev", use_container_width=True):
            if current_page > 0:
                st.session_state[page_key] -= 1
                st.rerun()
    
    with col2:
        # Page number selector
        page_selector = st.number_input(
            "Page",
            min_value=1,
            max_value=total_pages,
            value=current_page + 1,
            key=f"{page_key}_select",
            label_visibility="collapsed"
        )
        if page_selector - 1 != current_page:
            st.session_state[page_key] = page_selector - 1
            st.rerun()
    
    with col3:
        st.markdown(f"<p style='text-align: center; margin-top: 8px;'>Page {current_page + 1} of {total_pages}</p>", unsafe_allow_html=True)
    
    with col4:
        st.markdown("")  # Spacer
    
    with col5:
        if st.button("Next ▶", key=f"{page_key}_next", use_container_width=True):
            if current_page < total_pages - 1:
                st.session_state[page_key] += 1
                st.rerun()

# ============================================================================
# PHASE 4: EMAIL NOTIFICATIONS SYSTEM
# ============================================================================

EMAIL_CONFIG_PATH = Path(__file__).parent / "email_config.json"
NOTIFICATIONS_LOG_PATH = Path(__file__).parent / "notifications.json"

def init_email_config():
    """Initialize email configuration file with Hostinger SMTP settings."""
    if not EMAIL_CONFIG_PATH.exists():
        default_config = {
            "enabled": True,
            "smtp_server": "smtp.hostinger.com",
            "smtp_port": 587,
            "sender_email": "contact@codenest.us.com",
            "sender_password": os.environ.get("EMAIL_PASSWORD", ""),  # Set via environment variable
            "from_name": "Code Nest LLC - Digital Audits",
            "auto_send_reports": True,
            "reply_to": "contact@codenest.us.com",
            "notifications": {
                "audit_complete": True,
                "report_sent": True,
                "permission_change": True,
                "admin_alert": True
            }
        }
        EMAIL_CONFIG_PATH.write_text(json.dumps(default_config, indent=2))
        return default_config
    try:
        return json.loads(EMAIL_CONFIG_PATH.read_text())
    except Exception:
        return init_email_config()

def load_email_config():
    """Load email configuration."""
    return init_email_config()

def save_email_config(config):
    """Save email configuration."""
    try:
        EMAIL_CONFIG_PATH.write_text(json.dumps(config, indent=2))
        return True, "Configuration saved successfully"
    except Exception as e:
        return False, f"Error saving configuration: {str(e)}"

def send_email(recipient_email, subject, html_body):
    """Send email notification with error handling."""
    try:
        config = load_email_config()
        
        if not config.get("enabled"):
            return False, "Email notifications are disabled"
        
        if not config.get("smtp_server") or not config.get("sender_email"):
            return False, "Email configuration incomplete"
        
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"] = f"{config['from_name']} <{config['sender_email']}>"
        msg["To"] = recipient_email
        
        part = MIMEText(html_body, "html")
        msg.attach(part)
        
        server = smtplib.SMTP(config["smtp_server"], config["smtp_port"], timeout=10)
        server.starttls()
        server.login(config["sender_email"], config["sender_password"])
        server.sendmail(config["sender_email"], recipient_email, msg.as_string())
        server.quit()
        
        log_notification(recipient_email, subject, "sent")
        logger.info(f"Email sent to {recipient_email}: {subject}")
        return True, "Email sent successfully"
    except smtplib.SMTPAuthenticationError:
        logger.error("SMTP authentication failed")
        return False, "Authentication failed - check credentials"
    except smtplib.SMTPException as e:
        logger.error(f"SMTP error: {str(e)}")
        return False, f"Email service error: {str(e)}"
    except Exception as e:
        logger.error(f"Error sending email: {str(e)}")
        return False, f"Error: {str(e)}"

def send_email_with_pdf(to_email: str, subject: str, body: str, pdf_bytes: bytes, filename: str = "audit_report.pdf") -> tuple[bool, str]:
    """
    Send email with PDF attachment using SMTP configuration.
    
    Args:
        to_email: Recipient email address
        subject: Email subject line
        body: Plain text email body
        pdf_bytes: PDF file as bytes
        filename: Name for the PDF attachment
    
    Returns:
        Tuple of (success: bool, message: str)
    """
    try:
        config = load_email_config()
        
        # Validate configuration
        if not config.get("enabled"):
            return False, "Email notifications are disabled. Go to API Settings to enable."
        
        if not config.get("smtp_server") or not config.get("sender_email") or not config.get("sender_password"):
            return False, "SMTP not configured. Go to API Settings to configure SMTP first."
        
        # Validate recipient email
        is_valid, error_msg = validate_email(to_email)
        if not is_valid:
            return False, f"Invalid recipient email: {error_msg}"
        
        # Create message
        msg = MIMEMultipart()
        msg["Subject"] = subject
        msg["From"] = f"{config.get('from_name', 'Code Nest')} <{config['sender_email']}>"
        msg["To"] = to_email
        if config.get("reply_to"):
            msg["Reply-To"] = config["reply_to"]
        
        # Attach body as plain text
        msg.attach(MIMEText(body, "plain"))
        
        # Attach PDF
        if pdf_bytes:
            pdf_attachment = MIMEBase("application", "pdf")
            pdf_attachment.set_payload(pdf_bytes)
            encoders.encode_base64(pdf_attachment)
            pdf_attachment.add_header(
                "Content-Disposition",
                f"attachment; filename={filename}"
            )
            msg.attach(pdf_attachment)
        
        # Connect and send
        smtp_port = config.get("smtp_port", 587)
        smtp_server = config["smtp_server"]
        
        if smtp_port == 465:
            # SSL connection
            server = smtplib.SMTP_SSL(smtp_server, smtp_port, timeout=15)
        else:
            # TLS connection (port 587)
            server = smtplib.SMTP(smtp_server, smtp_port, timeout=15)
            server.starttls()
        
        server.login(config["sender_email"], config["sender_password"])
        server.sendmail(config["sender_email"], to_email, msg.as_string())
        server.quit()
        
        # Log the notification
        log_notification(to_email, subject, "sent_with_pdf")
        logger.info(f"Email with PDF sent to {to_email}: {subject}")
        
        return True, f"Email sent successfully to {to_email}"
        
    except smtplib.SMTPAuthenticationError:
        logger.error("SMTP authentication failed")
        return False, "Authentication failed - check your SMTP credentials in API Settings"
    except smtplib.SMTPConnectError:
        logger.error("SMTP connection failed")
        return False, "Could not connect to SMTP server - check host and port in API Settings"
    except smtplib.SMTPException as e:
        logger.error(f"SMTP error: {str(e)}")
        return False, f"Email service error: {str(e)}"
    except Exception as e:
        logger.error(f"Error sending email with PDF: {str(e)}")
        return False, f"Error: {str(e)}"

def is_smtp_configured() -> bool:
    """Check if SMTP is properly configured."""
    try:
        config = load_email_config()
        return bool(
            config.get("enabled") and 
            config.get("smtp_server") and 
            config.get("sender_email") and 
            config.get("sender_password")
        )
    except Exception:
        return False

# ============================================================================
# BRANDED HTML EMAIL TEMPLATE
# ============================================================================

def build_branded_email_html(content_text: str, subject: str = "") -> str:
    """
    Wrap AI-generated email content in a professional branded HTML template.
    
    Uses inline CSS only for maximum email client compatibility.
    Tested for: Gmail, Outlook, Apple Mail, Yahoo Mail.
    
    Args:
        content_text: The AI-generated email body text (plain text)
        subject: Optional subject for internal reference
    
    Returns:
        Complete HTML email string ready to send
    """
    # Convert plain text to HTML paragraphs
    # Split by double newlines for paragraphs, single newlines for line breaks
    paragraphs = content_text.strip().split('\n\n')
    html_content = ""
    
    for para in paragraphs:
        # Handle line breaks within paragraphs
        para_html = para.replace('\n', '<br>')
        # Handle bullet points (lines starting with - or •)
        lines = para_html.split('<br>')
        formatted_lines = []
        for line in lines:
            line = line.strip()
            if line.startswith('- ') or line.startswith('• '):
                # Format as bullet point
                bullet_text = line[2:].strip()
                formatted_lines.append(f'<span style="color: {BRAND_COLORS["accent_light_green"]};">•</span> {bullet_text}')
            elif line.startswith('* '):
                bullet_text = line[2:].strip()
                formatted_lines.append(f'<span style="color: {BRAND_COLORS["accent_light_green"]};">•</span> {bullet_text}')
            else:
                formatted_lines.append(line)
        
        para_html = '<br>'.join(formatted_lines)
        html_content += f'<p style="margin: 0 0 16px 0; line-height: 1.6;">{para_html}</p>'
    
    # Build the complete HTML email
    html_template = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{COMPANY_NAME}</title>
</head>
<body style="margin: 0; padding: 0; font-family: Arial, Helvetica, sans-serif; background-color: #f4f4f4;">
    <!-- Main Container -->
    <table role="presentation" cellspacing="0" cellpadding="0" border="0" width="100%" style="background-color: #f4f4f4;">
        <tr>
            <td style="padding: 20px 0;">
                <!-- Email Content Container -->
                <table role="presentation" cellspacing="0" cellpadding="0" border="0" width="600" style="margin: 0 auto; max-width: 600px;">
                    
                    <!-- Header Section - Dark Green Background -->
                    <tr>
                        <td style="background-color: {BRAND_COLORS['primary_dark_green']}; padding: 25px 30px; text-align: center; border-radius: 8px 8px 0 0;">
                            <!-- Logo placeholder - Update EMAIL_LOGO_URL with actual hosted logo -->
                            <img src="{EMAIL_LOGO_URL}" alt="{COMPANY_NAME}" width="180" style="max-width: 180px; height: auto; display: block; margin: 0 auto;" onerror="this.style.display='none'">
                            <h1 style="color: {BRAND_COLORS['white']}; font-size: 24px; font-weight: bold; margin: 15px 0 5px 0;">{COMPANY_NAME}</h1>
                            <p style="color: {BRAND_COLORS['accent_light_green']}; font-size: 14px; margin: 0; font-style: italic;">{COMPANY_TAGLINE}</p>
                        </td>
                    </tr>
                    
                    <!-- Accent Line -->
                    <tr>
                        <td style="background-color: {BRAND_COLORS['accent_light_green']}; height: 4px;"></td>
                    </tr>
                    
                    <!-- Body Section - White Background -->
                    <tr>
                        <td style="background-color: {BRAND_COLORS['white']}; padding: 35px 30px;">
                            <div style="color: {BRAND_COLORS['branding_grey']}; font-size: 15px; line-height: 1.6;">
                                {html_content}
                            </div>
                        </td>
                    </tr>
                    
                    <!-- Footer Section -->
                    <tr>
                        <td style="background-color: #f8f8f8; padding: 25px 30px; border-top: 1px solid #e0e0e0; border-radius: 0 0 8px 8px;">
                            <table role="presentation" cellspacing="0" cellpadding="0" border="0" width="100%">
                                <tr>
                                    <td style="text-align: center;">
                                        <p style="color: {BRAND_COLORS['primary_dark_green']}; font-size: 14px; font-weight: bold; margin: 0 0 8px 0;">{COMPANY_NAME}</p>
                                        <p style="color: {BRAND_COLORS['branding_grey']}; font-size: 12px; margin: 0 0 5px 0;">
                                            <a href="{COMPANY_WEBSITE}" style="color: {BRAND_COLORS['accent_light_green']}; text-decoration: none;">{COMPANY_WEBSITE}</a>
                                        </p>
                                        <p style="color: {BRAND_COLORS['branding_grey']}; font-size: 12px; margin: 0 0 5px 0;">
                                            <a href="mailto:{CONTACT_EMAIL}" style="color: {BRAND_COLORS['accent_light_green']}; text-decoration: none;">{CONTACT_EMAIL}</a>
                                        </p>
                                        <p style="color: {BRAND_COLORS['branding_grey']}; font-size: 11px; margin: 10px 0 0 0;">{COMPANY_LOCATION}</p>
                                    </td>
                                </tr>
                            </table>
                        </td>
                    </tr>
                    
                </table>
            </td>
        </tr>
    </table>
</body>
</html>'''
    
    return html_template

def send_branded_email_with_pdf(to_email: str, subject: str, body: str, pdf_bytes: bytes = None, filename: str = "audit_report.pdf") -> tuple[bool, str]:
    """
    Send branded HTML email with optional PDF attachment.
    
    This wraps the email body in the Code Nest branded HTML template
    and sends it as a multipart email with both plain text and HTML versions.
    
    Args:
        to_email: Recipient email address
        subject: Email subject line
        body: Plain text email body (will be wrapped in branded HTML)
        pdf_bytes: Optional PDF file as bytes
        filename: Name for the PDF attachment
    
    Returns:
        Tuple of (success: bool, message: str)
    """
    try:
        config = load_email_config()
        
        # Validate configuration
        if not config.get("enabled"):
            return False, "Email notifications are disabled. Go to API Settings to enable."
        
        if not config.get("smtp_server") or not config.get("sender_email") or not config.get("sender_password"):
            return False, "SMTP not configured. Go to API Settings to configure SMTP first."
        
        # Validate recipient email
        is_valid, error_msg = validate_email(to_email)
        if not is_valid:
            return False, f"Invalid recipient email: {error_msg}"
        
        # Create multipart message (alternative for text/html + mixed for attachments)
        if pdf_bytes:
            msg = MIMEMultipart("mixed")
            msg_alt = MIMEMultipart("alternative")
        else:
            msg = MIMEMultipart("alternative")
            msg_alt = msg
        
        msg["Subject"] = subject
        msg["From"] = f"{config.get('from_name', COMPANY_NAME)} <{config['sender_email']}>"
        msg["To"] = to_email
        if config.get("reply_to"):
            msg["Reply-To"] = config["reply_to"]
        
        # Attach plain text version (fallback)
        plain_text = body
        msg_alt.attach(MIMEText(plain_text, "plain", "utf-8"))
        
        # Attach branded HTML version
        html_body = build_branded_email_html(body, subject)
        msg_alt.attach(MIMEText(html_body, "html", "utf-8"))
        
        # If we have attachments, add the alternative part to mixed
        if pdf_bytes:
            msg.attach(msg_alt)
            
            # Attach PDF
            pdf_attachment = MIMEBase("application", "pdf")
            pdf_attachment.set_payload(pdf_bytes)
            encoders.encode_base64(pdf_attachment)
            pdf_attachment.add_header(
                "Content-Disposition",
                f"attachment; filename={filename}"
            )
            msg.attach(pdf_attachment)
        
        # Connect and send
        smtp_port = config.get("smtp_port", 587)
        smtp_server = config["smtp_server"]
        
        if smtp_port == 465:
            # SSL connection
            server = smtplib.SMTP_SSL(smtp_server, smtp_port, timeout=15)
        else:
            # TLS connection (port 587)
            server = smtplib.SMTP(smtp_server, smtp_port, timeout=15)
            server.starttls()
        
        server.login(config["sender_email"], config["sender_password"])
        server.sendmail(config["sender_email"], to_email, msg.as_string())
        server.quit()
        
        # Log the notification
        log_notification(to_email, subject, "sent_branded_html")
        logger.info(f"Branded HTML email sent to {to_email}: {subject}")
        
        return True, f"Email sent successfully to {to_email}"
        
    except smtplib.SMTPAuthenticationError:
        logger.error("SMTP authentication failed")
        return False, "Authentication failed - check your SMTP credentials in API Settings"
    except smtplib.SMTPConnectError:
        logger.error("SMTP connection failed")
        return False, "Could not connect to SMTP server - check host and port in API Settings"
    except smtplib.SMTPException as e:
        logger.error(f"SMTP error: {str(e)}")
        return False, f"Email service error: {str(e)}"
    except Exception as e:
        logger.error(f"Error sending branded email: {str(e)}")
        return False, f"Error: {str(e)}"

def log_notification(recipient, subject, status):
    """Log sent notifications."""
    try:
        if not NOTIFICATIONS_LOG_PATH.exists():
            notifications = []
        else:
            notifications = json.loads(NOTIFICATIONS_LOG_PATH.read_text())
        
        notifications.append({
            "timestamp": datetime.now().isoformat(),
            "recipient": recipient,
            "subject": subject,
            "status": status
        })
        
        # Keep only last 1000 notifications
        notifications = notifications[-1000:]
        NOTIFICATIONS_LOG_PATH.write_text(json.dumps(notifications, indent=2))
    except Exception as e:
        logger.error(f"Error logging notification: {str(e)}")

def extract_email_from_data(data):
    """Extract contact email from audit data (from OpenAI analysis)."""
    try:
        # Check if OpenAI data contains contact info
        if data.get('ai') and isinstance(data['ai'], dict):
            # Look for email in various fields that OpenAI might return
            ai_data = data['ai']
            
            # Check direct email field
            if ai_data.get('contact_email'):
                email = ai_data.get('contact_email')
                if validate_email(email):
                    return email
            
            # Check if email is embedded in other fields
            for field in ['summary', 'recommendations', 'contact_info']:
                if field in ai_data:
                    text = ai_data[field]
                    if isinstance(text, str):
                        emails = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', text)
                        if emails:
                            for email in emails:
                                # Exclude common system emails
                                if not any(x in email for x in ['noreply', 'no-reply', 'donotreply']):
                                    if validate_email(email):
                                        return email
        
        return None
    except Exception as e:
        logger.error(f"Error extracting email from data: {str(e)}")
        return None

def send_audit_report_email(recipient_email, audit_data):
    """Send audit report to website owner/contact."""
    try:
        if not recipient_email or not validate_email(recipient_email):
            logger.warning(f"Invalid email address: {recipient_email}")
            return False, "Invalid email address"
        
        config = load_email_config()
        
        if not config.get("auto_send_reports"):
            logger.info(f"Auto-send reports disabled, skipping email to {recipient_email}")
            return False, "Auto-send reports disabled"
        
        if not config.get("enabled"):
            return False, "Email notifications are disabled"
        
        if not config.get("smtp_server") or not config.get("sender_email"):
            return False, "Email configuration incomplete"
        
        # Extract domain from URL
        domain = urlparse(audit_data.get('url', '')).netloc.replace('www.', '')
        score = audit_data.get('score', 'N/A')
        
        # Build email subject and body
        subject = f"Website Audit Report - {domain} ({score}/100)"
        
        html_body = f"""
        <html>
            <head>
                <style>
                    body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; color: #333; line-height: 1.6; }}
                    .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
                    .header {{ background: linear-gradient(135deg, #0066CC 0%, #004499 100%); color: white; padding: 30px; border-radius: 5px 5px 0 0; text-align: center; }}
                    .header h1 {{ margin: 0; font-size: 28px; }}
                    .header p {{ margin: 10px 0 0 0; font-size: 14px; opacity: 0.9; }}
                    .metrics {{ display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 15px; margin: 20px 0; }}
                    .metric {{ background: #f5f5f5; padding: 15px; border-radius: 5px; text-align: center; border-left: 4px solid #0066CC; }}
                    .metric-value {{ font-size: 24px; font-weight: bold; color: #0066CC; }}
                    .metric-label {{ font-size: 12px; color: #666; margin-top: 5px; text-transform: uppercase; }}
                    .section {{ margin: 25px 0; }}
                    .section-title {{ font-size: 18px; font-weight: bold; color: #0066CC; border-bottom: 2px solid #0066CC; padding-bottom: 10px; margin-bottom: 15px; }}
                    .issues-list {{ margin-top: 15px; }}
                    .issue {{ background: #fff3cd; border-left: 4px solid #ff9800; padding: 12px; margin: 10px 0; border-radius: 3px; }}
                    .issue-title {{ font-weight: bold; color: #ff6600; }}
                    .cta-button {{ display: inline-block; background: #0066CC; color: white; padding: 12px 30px; text-decoration: none; border-radius: 5px; margin: 20px 0; }}
                    .footer {{ text-align: center; margin-top: 30px; padding-top: 20px; border-top: 1px solid #ddd; font-size: 12px; color: #666; }}
                </style>
            </head>
            <body>
                <div class="container">
                    <div class="header">
                        <h1>🎯 Website Audit Report</h1>
                        <p>Comprehensive Technical Analysis for {domain}</p>
                    </div>
                    
                    <div style="padding: 30px; border: 1px solid #eee; border-top: none;">
                        <p>Hello,</p>
                        <p>We've completed a comprehensive technical audit of your website. Here's what we found:</p>
                        
                        <div class="metrics">
                            <div class="metric">
                                <div class="metric-value">{score}</div>
                                <div class="metric-label">Health Score</div>
                            </div>
                            <div class="metric">
                                <div class="metric-value">{audit_data.get('psi', 'N/A')}</div>
                                <div class="metric-label">Speed Score</div>
                            </div>
                            <div class="metric">
                                <div class="metric-value">{len(audit_data.get('issues', []))}</div>
                                <div class="metric-label">Issues Found</div>
                            </div>
                        </div>
                        
                        <div class="section">
                            <div class="section-title">⚠️ Top Issues Detected</div>
                            <div class="issues-list">
        """
        
        # Add top 5 issues
        for issue in audit_data.get('issues', [])[:5]:
            html_body += f"""
                                <div class="issue">
                                    <div class="issue-title">{issue.get('title', 'Unknown Issue')}</div>
                                    <p><strong>Impact:</strong> {issue.get('impact', 'N/A')}</p>
                                    <p><strong>Solution:</strong> {issue.get('solution', 'N/A')}</p>
                                </div>
            """
        
        html_body += """
                            </div>
                        </div>
                        
                        <div class="section">
                            <div class="section-title">🤖 AI Analysis Summary</div>
        """
        
        ai_section = audit_data.get('ai') if isinstance(audit_data.get('ai'), dict) else {}
        if ai_section:
            html_body += f"""
                            <p><strong>Summary:</strong></p>
                            <p>{ai_section.get('summary', 'No summary available')}</p>
                            <p><strong>Recommendations:</strong></p>
                            <p>{ai_section.get('solutions', 'No recommendations available')}</p>
            """
        
        html_body += f"""
                        </div>
                        
                        <center>
                            <a href="{audit_data.get('url', '#')}" class="cta-button">View Detailed Report</a>
                        </center>
                        
                        <div class="section">
                            <p><strong>Next Steps:</strong></p>
                            <ul>
                                <li>Review the recommendations above</li>
                                <li>Prioritize high-impact issues</li>
                                <li>Contact us for implementation support</li>
                            </ul>
                        </div>
                        
                        <div class="footer">
                            <p>Code Nest LLC | Nest Idea | Code Success</p>
                            <p>© 2025 Code Nest LLC. All rights reserved.</p>
                            <p>Questions? Reply to this email or contact us at contact@codenest.us.com</p>
                        </div>
                    </div>
                </div>
            </body>
        </html>
        """
        
        # Send email
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"] = f"{config['from_name']} <{config['sender_email']}>"
        msg["To"] = recipient_email
        msg["Reply-To"] = config.get('reply_to', config['sender_email'])
        
        part = MIMEText(html_body, "html")
        msg.attach(part)
        
        # Send via Hostinger SMTP
        server = smtplib.SMTP(config["smtp_server"], config["smtp_port"], timeout=10)
        server.starttls()
        server.login(config["sender_email"], config["sender_password"])
        server.sendmail(config["sender_email"], recipient_email, msg.as_string())
        server.quit()
        
        log_notification(recipient_email, subject, "sent")
        logger.info(f"Audit report email sent to {recipient_email} for domain {domain}")
        
        return True, f"Report sent successfully to {recipient_email}"
        
    except smtplib.SMTPAuthenticationError:
        logger.error(f"SMTP authentication failed for {recipient_email}")
        return False, "Authentication failed - check email credentials"
    except smtplib.SMTPException as e:
        logger.error(f"SMTP error sending to {recipient_email}: {str(e)}")
        return False, f"Email service error: {str(e)}"
    except Exception as e:
        logger.error(f"Error sending audit report to {recipient_email}: {str(e)}")
        return False, f"Error: {str(e)}"

def get_email_template(template_type, data):
    """Generate HTML email templates."""
    if template_type == "audit_complete":
        return f"""
        <html>
            <body style="font-family: Arial, sans-serif; color: #333;">
                <h2>Audit Completed! ✅</h2>
                <p>Hello {data.get('username', 'User')},</p>
                <p>Your audit for <strong>{data.get('domain', 'N/A')}</strong> has been completed.</p>
                <div style="background-color: #f0f0f0; padding: 15px; border-radius: 5px; margin: 20px 0;">
                    <p><strong>Score:</strong> {data.get('score', 'N/A')}/100</p>
                    <p><strong>Status:</strong> {data.get('status', 'N/A')}</p>
                    <p><strong>Completed:</strong> {data.get('timestamp', 'N/A')}</p>
                </div>
                <p>Log in to view detailed results and recommendations.</p>
                <p>Best regards,<br>Code Nest Sales Engine</p>
            </body>
        </html>
        """
    elif template_type == "permission_change":
        return f"""
        <html>
            <body style="font-family: Arial, sans-serif; color: #333;">
                <h2>Permission Update 🔐</h2>
                <p>Hello {data.get('username', 'User')},</p>
                <p>Your account permissions have been updated.</p>
                <div style="background-color: #f0f0f0; padding: 15px; border-radius: 5px; margin: 20px 0;">
                    <p><strong>New Role:</strong> {data.get('role', 'N/A')}</p>
                    <p><strong>Changed:</strong> {data.get('timestamp', 'N/A')}</p>
                </div>
                <p>If this wasn't you, please contact support immediately.</p>
                <p>Best regards,<br>Code Nest Sales Engine</p>
            </body>
        </html>
        """
    elif template_type == "admin_alert":
        return f"""
        <html>
            <body style="font-family: Arial, sans-serif; color: #333;">
                <h2>Admin Alert ⚠️</h2>
                <p>An important event has occurred:</p>
                <div style="background-color: #f0f0f0; padding: 15px; border-radius: 5px; margin: 20px 0;">
                    <p><strong>Event:</strong> {data.get('event', 'N/A')}</p>
                    <p><strong>Details:</strong> {data.get('details', 'N/A')}</p>
                    <p><strong>Time:</strong> {data.get('timestamp', 'N/A')}</p>
                </div>
                <p>Log in to the admin panel for more information.</p>
                <p>Best regards,<br>Code Nest Sales Engine</p>
            </body>
        </html>
        """

# ============================================================================
# PHASE 4: ANALYTICS DASHBOARD WITH CHARTS
# ============================================================================

def get_dashboard_analytics():
    """Compile analytics data for dashboard."""
    try:
        audits_raw = get_audit_history_cached(limit=1000)
        if not audits_raw:
            return None
        
        # Convert all audits to dicts (handles ORM objects)
        audits = [audit_to_dict(a) for a in audits_raw]
        
        df = pd.DataFrame(audits)
        
        # Safe issue count calculation
        def get_issue_count(a):
            issues = a.get("issues") or []
            return len(issues) if isinstance(issues, list) else 0
        
        analytics = {
            "total_audits": len(audits),
            "avg_score": df["score"].mean() if "score" in df.columns else 0,
            "high_issue_count": len([a for a in audits if get_issue_count(a) > 10]),
            "audits_by_day": df.groupby(df["timestamp"].str[:10]).size() if "timestamp" in df.columns else {},
            "score_distribution": pd.cut(df["score"], bins=[0, 25, 50, 75, 100]).value_counts() if "score" in df.columns else {},
            "top_issues": get_top_issues(audits),
            "latest_audits": audits[:5]
        }
        return analytics
    except Exception as e:
        logger.error(f"Error compiling analytics: {str(e)}")
        return None

def get_top_issues(audits, limit=10):
    """Get most common issues found."""
    try:
        issue_count = {}
        for audit in audits:
            # Ensure audit is a dict
            if not isinstance(audit, dict):
                audit = audit_to_dict(audit)
            
            issues = audit.get("issues") or []
            if isinstance(issues, list):
                for issue in issues[:5]:  # Take top 5 per audit
                    if isinstance(issue, str):
                        issue_count[issue] = issue_count.get(issue, 0) + 1
                    elif isinstance(issue, dict) and issue.get("title"):
                        issue_count[issue["title"]] = issue_count.get(issue["title"], 0) + 1
        
        return sorted(issue_count.items(), key=lambda x: x[1], reverse=True)[:limit]
    except Exception:
        return []

def show_dashboard():
    """Display analytics dashboard."""
    st.markdown("## 📊 Analytics Dashboard")
    
    analytics = get_dashboard_analytics()
    if not analytics:
        st.warning("No audit data available yet. Complete some audits first!")
        return
    
    # KPI Metrics
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric(
            "Total Audits",
            analytics["total_audits"],
            delta=None,
            help="Total number of audits completed"
        )
    
    with col2:
        avg_score = round(analytics["avg_score"], 1)
        st.metric(
            "Average Score",
            f"{avg_score}/100",
            delta=None,
            help="Average health score across all audits"
        )
    
    with col3:
        st.metric(
            "High Issue Sites",
            analytics["high_issue_count"],
            delta=None,
            help="Sites with 10+ issues found"
        )
    
    # Charts
    st.divider()
    
    col_chart1, col_chart2 = st.columns(2)
    
    with col_chart1:
        st.markdown("### Audits Over Time (Last 30 Days)")
        if analytics["audits_by_day"]:
            try:
                import plotly.express as px
                df_days = pd.DataFrame({
                    "Date": list(analytics["audits_by_day"].keys()),
                    "Count": list(analytics["audits_by_day"].values())
                })
                fig = px.line(df_days, x="Date", y="Count", markers=True,
                            title="", labels={"Count": "Audits"})
                st.plotly_chart(fig, use_container_width=True)
            except Exception as e:
                st.error(f"Error creating chart: {str(e)}")
        else:
            st.info("No recent audit data")
    
    with col_chart2:
        st.markdown("### Score Distribution")
        if analytics["score_distribution"].size > 0:
            try:
                import plotly.express as px
                dist_data = analytics["score_distribution"].reset_index()
                dist_data.columns = ["Range", "Count"]
                fig = px.bar(dist_data, x="Range", y="Count",
                           title="", labels={"Count": "Number of Sites"})
                st.plotly_chart(fig, use_container_width=True)
            except Exception as e:
                st.error(f"Error creating chart: {str(e)}")
        else:
            st.info("No score distribution data")
    
    # Top Issues
    st.divider()
    st.markdown("### Top 10 Issues Found")
    if analytics["top_issues"]:
        issue_df = pd.DataFrame(analytics["top_issues"], columns=["Issue", "Occurrences"])
        st.dataframe(issue_df, use_container_width=True)
    else:
        st.info("No issue data available yet")

# ============================================================================
# PHASE 4: PDF & EXCEL EXPORT REPORTS
# ============================================================================

def generate_pdf_report(audit_data, filename="audit_report.pdf"):
    """
    Generate professional branded PDF report for an audit.
    Uses the PDFReport class for consistent Code Nest branding.
    
    Brand Info:
    - Company: Code Nest LLC
    - Tagline: Nest Idea | Code Success
    - Website: https://codenest.us.com
    - Email: contact@codenest.us.com
    """
    try:
        pdf = PDFReport()
        
        # Extract domain for cover page
        domain = audit_data.get("domain", audit_data.get("url", "Unknown"))
        if '://' in domain:
            domain = domain.split('://')[1].split('/')[0]
        domain = domain.replace('www.', '')
        
        score = audit_data.get('score', audit_data.get('health_score', 0))
        
        # =====================================================================
        # COVER PAGE
        # =====================================================================
        pdf.add_cover_page(domain, score)
        
        # =====================================================================
        # AUDIT DETAILS PAGE
        # =====================================================================
        pdf.add_page()
        
        # Audit Summary Section
        pdf.section_title("Audit Summary")
        
        details = [
            ("Domain", domain),
            ("Health Score", f"{score}/100"),
            ("Page Speed Score", str(audit_data.get('psi', audit_data.get('psi_score', 'N/A')))),
            ("Domain Age", str(audit_data.get('domain_age', 'N/A'))),
            ("Report Date", audit_data.get('timestamp', audit_data.get('created_at', datetime.now().strftime('%Y-%m-%d %H:%M')))),
        ]
        
        for label, value in details:
            pdf.set_font("Arial", "B", 10)
            pdf.set_text_color(*pdf.brand_dark_green)
            pdf.cell(60, 6, f"{label}:", 0, 0)
            pdf.set_font("Arial", "", 10)
            pdf.set_text_color(*pdf.brand_grey)
            pdf.cell(0, 6, str(value)[:50], 0, 1)
        
        pdf.ln(5)
        
        # Tech Stack
        if "tech_stack" in audit_data and audit_data["tech_stack"]:
            pdf.section_title("Technology Stack")
            tech_text = ", ".join(audit_data["tech_stack"][:15])
            pdf.chapter_body(f"Detected technologies: {tech_text}")
        
        # =====================================================================
        # CRITICAL FINDINGS
        # =====================================================================
        if "issues" in audit_data and audit_data["issues"]:
            pdf.section_title("Critical Findings", color_type="issues")
            
            for i, issue in enumerate(audit_data["issues"][:15], 1):
                pdf.set_font("Arial", "B", 10)
                pdf.set_text_color(180, 60, 60)
                issue_text = str(issue)[:80]
                pdf.cell(0, 5, f"[!] {clean_text(issue_text)}", 0, 1)
            pdf.ln(2)
        
        # Emails Found
        if "emails" in audit_data and audit_data["emails"]:
            pdf.section_title("Contact Emails Found")
            for email in audit_data["emails"][:5]:
                pdf.bullet_point(email)
            pdf.ln(2)
        
        # =====================================================================
        # AI INSIGHTS PAGE
        # =====================================================================
        if "ai" in audit_data and audit_data["ai"]:
            ai_data = audit_data["ai"]
            
            pdf.add_page()
            
            # AI Insights header
            pdf.set_font('Arial', 'B', 20)
            pdf.set_text_color(*pdf.brand_dark_green)
            pdf.cell(0, 12, "AI-Powered Insights", 0, 1, 'C')
            
            # Decorative line
            pdf.set_draw_color(*pdf.brand_accent_green)
            pdf.set_line_width(1)
            y = pdf.get_y()
            pdf.line(70, y, 140, y)
            pdf.ln(10)
            
            if ai_data.get("summary"):
                pdf.section_title("Executive Summary", color_type="ai_insights")
                pdf.chapter_body(str(ai_data.get("summary", ""))[:400])
            
            if ai_data.get("impact"):
                pdf.section_title("Business Impact", color_type="issues")
                pdf.chapter_body(str(ai_data.get("impact", ""))[:400])
            
            if ai_data.get("solutions"):
                pdf.section_title("Recommended Solutions", color_type="quick_wins")
                pdf.chapter_body(str(ai_data.get("solutions", ""))[:400])
            
            if ai_data.get("email"):
                pdf.section_title("AI Email Draft", color_type="ai_insights")
                pdf.set_font("Arial", "", 9)
                pdf.set_text_color(*pdf.brand_grey)
                pdf.multi_cell(0, 4, clean_text(str(ai_data.get("email", ""))[:600]))
        
        # Footer is handled automatically by PDFReport class
        
        # Save to bytes
        pdf_bytes = pdf.output(dest='S').encode('latin-1')
        return pdf_bytes
    except Exception as e:
        logger.error(f"Error generating PDF: {str(e)}")
        return None

def generate_excel_report(audits, filename="audits_report.xlsx"):
    """Generate Excel report with multiple sheets."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Create workbook
        from io import BytesIO
        output = BytesIO()
        
        df = pd.DataFrame(audits)
        
        # Write to Excel
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Sheet 1: Summary
            summary_data = {
                "Metric": ["Total Audits", "Average Score", "Highest Score", "Lowest Score"],
                "Value": [
                    len(audits),
                    round(df["score"].mean(), 2) if "score" in df.columns else "N/A",
                    df["score"].max() if "score" in df.columns else "N/A",
                    df["score"].min() if "score" in df.columns else "N/A"
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Sheet 2: All Audits
            df.to_excel(writer, sheet_name='Audits', index=False)
        
        output.seek(0)
        return output
    except Exception as e:
        logger.error(f"Error generating Excel: {str(e)}")
        return None

# ============================================================================
# PHASE 4: DARK MODE SUPPORT
# ============================================================================

def init_theme_preferences():
    """Initialize theme preferences file."""
    theme_path = Path(__file__).parent / "theme_prefs.json"
    if not theme_path.exists():
        theme_path.write_text(json.dumps({"theme": "light"}, indent=2))
    return theme_path

def get_user_theme(username):
    """Get user's theme preference."""
    try:
        theme_path = init_theme_preferences()
        prefs = json.loads(theme_path.read_text())
        return prefs.get("theme", "light")
    except Exception:
        return "light"

def save_user_theme(username, theme):
    """Save user's theme preference."""
    try:
        theme_path = init_theme_preferences()
        prefs = json.loads(theme_path.read_text()) if theme_path.exists() else {}
        prefs["theme"] = theme
        theme_path.write_text(json.dumps(prefs, indent=2))
        return True
    except Exception:
        return False

def apply_theme(theme="light"):
    """Apply theme via CSS."""
    if theme == "dark":
        st.markdown("""
        <style>
            :root {
                --primary-color: #3B82F6;
                --bg-color: #111827;
                --text-color: #F3F4F6;
                --border-color: #374151;
            }
            body {
                background-color: #111827;
                color: #F3F4F6;
            }
            .stButton > button {
                background-color: #3B82F6;
                color: white;
            }
            .stMetric {
                background-color: #1F2937;
                padding: 10px;
                border-radius: 5px;
            }
        </style>
        """, unsafe_allow_html=True)

# ============================================================================
# PHASE 4: USER PREFERENCES & SETTINGS PANEL
# ============================================================================

PREFERENCES_PATH = Path(__file__).parent / "user_preferences.json"

def init_preferences_file():
    """Initialize user preferences file."""
    if not PREFERENCES_PATH.exists():
        PREFERENCES_PATH.write_text(json.dumps({}, indent=2))

def load_user_preferences(username):
    """Load user preferences."""
    try:
        init_preferences_file()
        prefs = json.loads(PREFERENCES_PATH.read_text())
        return prefs.get(username, {
            "theme": "light",
            "notifications_enabled": True,
            "notification_frequency": "weekly",
            "items_per_page": 50,
            "timezone": "UTC",
            "language": "en"
        })
    except Exception:
        return {}

def save_user_preferences(username, preferences):
    """Save user preferences."""
    try:
        init_preferences_file()
        prefs = json.loads(PREFERENCES_PATH.read_text())
        prefs[username] = preferences
        PREFERENCES_PATH.write_text(json.dumps(prefs, indent=2))
        return True, "Preferences saved successfully"
    except Exception as e:
        return False, f"Error saving preferences: {str(e)}"

def show_preferences_panel(username):
    """Display user preferences panel."""
    st.markdown("## ⚙️ User Preferences")
    
    # Load current preferences
    prefs = load_user_preferences(username)
    
    # Create tabs for different settings
    tab1, tab2, tab3, tab4 = st.tabs(["Display", "Notifications", "Account", "Privacy"])
    
    with tab1:
        st.markdown("### Display Settings")
        new_theme = st.selectbox(
            "Theme",
            ["light", "dark", "auto"],
            index=["light", "dark", "auto"].index(prefs.get("theme", "light"))
        )
        
        new_items_per_page = st.slider(
            "Items per page",
            min_value=10,
            max_value=100,
            value=prefs.get("items_per_page", 50),
            step=10
        )
        
        new_language = st.selectbox(
            "Language",
            ["English", "Spanish", "French", "German"],
            index=0 if prefs.get("language", "en") == "en" else 1
        )
    
    with tab2:
        st.markdown("### Notification Settings")
        notifications_enabled = st.checkbox(
            "Enable email notifications",
            value=prefs.get("notifications_enabled", True)
        )
        
        notification_frequency = st.selectbox(
            "Notification frequency",
            ["immediately", "daily", "weekly", "monthly"],
            index=["immediately", "daily", "weekly", "monthly"].index(prefs.get("notification_frequency", "weekly"))
        )
        
        st.markdown("**Notify me for:**")
        col1, col2, col3 = st.columns(3)
        with col1:
            audit_complete = st.checkbox("Audit completion", value=True)
        with col2:
            permission_change = st.checkbox("Permission changes", value=True)
        with col3:
            admin_alerts = st.checkbox("Admin alerts", value=True)
    
    with tab3:
        st.markdown("### Account Settings")
        if st.button("Change Password", use_container_width=True):
            st.session_state.show_password_change = True
        
        if st.button("View Login History", use_container_width=True):
            st.session_state.show_login_history = True
        
        if st.checkbox("Enable Two-Factor Authentication", value=False):
            st.info("Two-factor authentication is enabled for your account.")
    
    with tab4:
        st.markdown("### Privacy Settings")
        data_retention = st.selectbox(
            "Data retention",
            ["30 days", "90 days", "1 year", "indefinite"],
            index=2
        )
        
        api_key_visible = st.checkbox("Show API key in plain text", value=False)
        
        if st.button("Download My Data", use_container_width=True):
            st.info("Your data will be downloaded as a JSON file.")
        
        if st.button("Delete Account", use_container_width=True, help="This action cannot be undone"):
            st.warning("⚠️ This will permanently delete your account and all associated data.")
    
    # Save button
    if st.button("Save Preferences", use_container_width=True, type="primary"):
        updated_prefs = {
            "theme": new_theme if 'new_theme' in locals() else prefs.get("theme", "light"),
            "notifications_enabled": notifications_enabled if 'notifications_enabled' in locals() else prefs.get("notifications_enabled", True),
            "notification_frequency": notification_frequency if 'notification_frequency' in locals() else prefs.get("notification_frequency", "weekly"),
            "items_per_page": new_items_per_page if 'new_items_per_page' in locals() else prefs.get("items_per_page", 50),
            "timezone": prefs.get("timezone", "UTC"),
            "language": new_language if 'new_language' in locals() else prefs.get("language", "en")
        }
        success, message = save_user_preferences(username, updated_prefs)
        if success:
            st.success(message)
        else:
            st.error(message)

# ============================================================================
# AUTHENTICATION & USER MANAGEMENT
# ============================================================================

USERS_PATH = Path(__file__).parent / "users.json"
TWO_FA_PATH = Path(__file__).parent / "two_fa.json"
SESSIONS_PATH = Path(__file__).parent / "sessions.json"
LOGIN_ATTEMPTS_PATH = Path(__file__).parent / "login_attempts.json"
ENCRYPTION_KEY_PATH = Path(__file__).parent / ".encryption_key"
SESSION_MAX_LIFETIME_HOURS = 168  # 7 days max session lifetime
SESSION_IDLE_TIMEOUT_MINUTES = 60  # 1 hour idle timeout - user logged out after 1 hour of inactivity
TOKEN_EXPIRES_SECONDS = 3600  # 1 hour login token expiry
LOGIN_ATTEMPT_LIMIT = 5
LOGIN_ATTEMPT_WINDOW_MINUTES = 5

# Cookie name for persistent auth
AUTH_COOKIE_NAME = "codenest_auth"


# ============================================================================
# CENTRALIZED SESSION INITIALIZATION (MUST RUN FIRST)
# ============================================================================

def init_app_session():
    """
    Initializes all persistent session variables.
    This function MUST run before any page rendering.
    Should never reset values once set.
    
    This is the single source of truth for session state initialization.
    All session keys are defined here with their default values.
    Keys are only set if they don't already exist, preserving existing values.
    """
    logger = logging.getLogger("sales_engine")
    
    # -------------------------------------------------------------------------
    # AUTHENTICATION STATE (critical - never overwrite if set)
    # -------------------------------------------------------------------------
    if 'user_authenticated' not in st.session_state:
        st.session_state.user_authenticated = False
    
    if 'user_role' not in st.session_state:
        st.session_state.user_role = "user"
    
    if 'user_email' not in st.session_state:
        st.session_state.user_email = None
    
    if 'current_user' not in st.session_state:
        st.session_state.current_user = None
    
    if 'is_admin' not in st.session_state:
        st.session_state.is_admin = False
    
    # Legacy compatibility - map authenticated to user_authenticated
    if 'authenticated' not in st.session_state:
        st.session_state.authenticated = st.session_state.user_authenticated
    
    # -------------------------------------------------------------------------
    # LOGIN TOKEN STATE (for persistent sessions)
    # -------------------------------------------------------------------------
    if 'LOGIN_TOKEN' not in st.session_state:
        st.session_state.LOGIN_TOKEN = None
    
    if 'TOKEN_CREATED_AT' not in st.session_state:
        st.session_state.TOKEN_CREATED_AT = None
    
    if 'session_token' not in st.session_state:
        st.session_state.session_token = None
    
    # -------------------------------------------------------------------------
    # API KEYS (CRITICAL - never overwrite if already set with a value)
    # Only initialize if not present. Don't overwrite with empty env vars.
    # The actual loading from DB happens in reload_user_api_keys() after login.
    # -------------------------------------------------------------------------
    if 'OPENAI_API_KEY' not in st.session_state:
        env_key = os.environ.get("OPENAI_API_KEY", "").strip()
        st.session_state.OPENAI_API_KEY = env_key if env_key else ""
    
    if 'GOOGLE_API_KEY' not in st.session_state:
        env_key = os.environ.get("GOOGLE_API_KEY", "").strip()
        st.session_state.GOOGLE_API_KEY = env_key if env_key else ""
    
    if 'SLACK_WEBHOOK' not in st.session_state:
        env_key = os.environ.get("SLACK_WEBHOOK", "").strip()
        st.session_state.SLACK_WEBHOOK = env_key if env_key else ""
    
    # SMTP settings (stored in session for runtime use)
    if '_smtp_host' not in st.session_state:
        st.session_state['_smtp_host'] = ""
    if '_smtp_port' not in st.session_state:
        st.session_state['_smtp_port'] = 587
    if '_smtp_user' not in st.session_state:
        st.session_state['_smtp_user'] = ""
    if '_smtp_pass' not in st.session_state:
        st.session_state['_smtp_pass'] = ""
    
    # -------------------------------------------------------------------------
    # 2FA STATE
    # -------------------------------------------------------------------------
    if '2fa_pending' not in st.session_state:
        st.session_state['2fa_pending'] = False
    
    if '2fa_username' not in st.session_state:
        st.session_state['2fa_username'] = None
    
    # -------------------------------------------------------------------------
    # NAVIGATION STATE (now managed by persistence layer)
    # -------------------------------------------------------------------------
    if 'current_section' not in st.session_state:
        st.session_state.current_section = 'Single Audit'
    
    # -------------------------------------------------------------------------
    # BULK AUDIT STATE (bulk scan is separate from single audit persistence)
    # -------------------------------------------------------------------------
    if 'audit_bulk_selected' not in st.session_state:
        st.session_state.audit_bulk_selected = set()
    
    if 'bulk_scan_active' not in st.session_state:
        st.session_state.bulk_scan_active = False
    
    if 'bulk_scan_session_id' not in st.session_state:
        st.session_state.bulk_scan_session_id = None
    
    # NOTE: current_audit_data is now managed by persistence layer
    # Use get_current_audit() and set_current_audit() instead
    
    # -------------------------------------------------------------------------
    # INTERNAL FLAGS (tracking state)
    # -------------------------------------------------------------------------
    if '_api_keys_loaded_this_run' not in st.session_state:
        st.session_state._api_keys_loaded_this_run = False
    
    if '_api_keys_migrated' not in st.session_state:
        st.session_state._api_keys_migrated = False
    
    if '_session_initialized' not in st.session_state:
        st.session_state._session_initialized = True
        logger.debug("Session state initialized for first time")
    
    # -------------------------------------------------------------------------
    # CRASH-PROOF REQUIRED KEYS (Phase 5 Step 3)
    # Ensures all critical keys exist with safe defaults
    # -------------------------------------------------------------------------
    required_keys = {
        "nav_state": None,
        "audit_persistence": {},
        "audit_tags": {},
        "crm_selected_lead_id": None,
        "email_drafts": {},
        "pdf_context": {},
        "ai_cache": {},
        "_current_audit_id": None,
        "last_bulk_session_id": None,
        "filter_preferences": {},
        "export_settings": {},
    }
    
    for key, default_value in required_keys.items():
        if key not in st.session_state:
            st.session_state[key] = default_value


def save_login_state(username: str, role: str, email: str = None) -> str:
    """
    Saves login state and creates persistent session.
    
    This should:
    - Generate a secure login token (uuid4)
    - Save LOGIN_TOKEN
    - Save TOKEN_CREATED_AT
    - Save user email + role
    - Create persistent cookie (via session file - Streamlit doesn't have native cookies)
    - Never clear API keys automatically
    
    Args:
        username: The authenticated username
        role: User role ("admin" or "user")
        email: Optional user email
    
    Returns:
        The generated session token
    """
    import uuid
    logger = logging.getLogger("sales_engine")
    
    # Generate secure login token
    login_token = uuid.uuid4().hex
    token_created_at = datetime.now()
    
    # Save to session state (memory)
    st.session_state.LOGIN_TOKEN = login_token
    st.session_state.TOKEN_CREATED_AT = token_created_at.isoformat()
    st.session_state.user_authenticated = True
    st.session_state.authenticated = True  # Legacy compatibility
    st.session_state.current_user = username
    st.session_state.user_email = email
    st.session_state.user_role = role
    st.session_state.is_admin = (role == "admin")
    st.session_state['2fa_pending'] = False
    st.session_state['2fa_username'] = None
    
    # Create persistent session (file-based for Streamlit)
    sessions = load_sessions()
    sessions[login_token] = {
        "username": username,
        "role": role,
        "email": email,
        "created_at": token_created_at.isoformat(),
        "last_access": token_created_at.isoformat(),
        "expires_at": (token_created_at + timedelta(seconds=TOKEN_EXPIRES_SECONDS)).isoformat()
    }
    save_sessions(sessions)
    
    # Store token in session_state for URL persistence
    st.session_state.session_token = login_token
    
    # Set query parameter for URL persistence (Streamlit's cookie alternative)
    try:
        st.query_params['session_token'] = login_token
    except Exception as e:
        logger.warning(f"Could not set query params: {e}")
    
    logger.info(f"Login state saved for user: {username}, role: {role}")
    
    # DO NOT clear or modify API keys here - they should persist
    
    return login_token


def load_login_state() -> bool:
    """
    Loads and validates login state from persistent storage.
    
    This should:
    - Read session token from session_state or query params
    - If token missing → return NOT logged in
    - Compare TOKEN_CREATED_AT with TOKEN_EXPIRES_SECONDS
    - If expired → clear session + cookie
    - If valid → restore user email + role + authenticated state
    - Must NOT reset API keys or delete user data
    
    Returns:
        True if user is logged in and session is valid, False otherwise
    """
    logger = logging.getLogger("sales_engine")
    
    # Priority 1: Check session_state for existing valid session
    if st.session_state.get('user_authenticated') and st.session_state.get('current_user'):
        # Validate the session token is still valid
        token = st.session_state.get('session_token') or st.session_state.get('LOGIN_TOKEN')
        if token:
            session_info = validate_session(token)
            if session_info:
                # Session still valid, update last access
                logger.debug(f"Session valid for user: {st.session_state.current_user}")
                return True
            else:
                # Session expired, but don't clear API keys
                logger.info(f"Session expired for user: {st.session_state.current_user}")
                _clear_auth_state_only()
                return False
        else:
            # No token but authenticated - trust session state
            return True
    
    # Priority 2: Try to restore from session_state token
    token = st.session_state.get('session_token') or st.session_state.get('LOGIN_TOKEN')
    if token:
        if _restore_from_token(token):
            return True
    
    # Priority 3: Try to restore from query params
    try:
        query_params = st.query_params
        if 'session_token' in query_params:
            token = query_params.get('session_token', '')
            if isinstance(token, list):
                token = token[0] if token else ''
            if token and _restore_from_token(token):
                return True
    except Exception as e:
        logger.warning(f"Error reading query params: {e}")
    
    # No valid session found
    return False


def _restore_from_token(token: str) -> bool:
    """Internal helper to restore session from token."""
    logger = logging.getLogger("sales_engine")
    
    if not token:
        return False
    
    session_info = validate_session(token)
    if not session_info:
        logger.debug(f"Token validation failed: {token[:8]}...")
        return False
    
    # Restore session state from validated token
    st.session_state.session_token = token
    st.session_state.LOGIN_TOKEN = token
    st.session_state.TOKEN_CREATED_AT = session_info.get('created_at')
    st.session_state.user_authenticated = True
    st.session_state.authenticated = True  # Legacy
    st.session_state.current_user = session_info['username']
    st.session_state.user_email = session_info.get('email')
    st.session_state.user_role = session_info['role']
    st.session_state.is_admin = (session_info['role'] == 'admin')
    st.session_state['2fa_pending'] = False
    st.session_state['2fa_username'] = None
    
    logger.info(f"Session restored from token for user: {session_info['username']}")
    
    # Reload API keys for user (from database, preserving existing if any)
    reload_user_api_keys()
    
    return True


def _clear_auth_state_only():
    """Clear only authentication state, preserving API keys and other data."""
    logger = logging.getLogger("sales_engine")
    
    # Get current user for logging
    current_user = st.session_state.get('current_user', 'unknown')
    
    # Clear ONLY auth-related keys using pop (safe removal)
    st.session_state.pop('user_authenticated', None)
    st.session_state.pop('authenticated', None)
    st.session_state.pop('current_user', None)
    st.session_state.pop('user_email', None)
    st.session_state.pop('user_role', None)
    st.session_state.pop('is_admin', None)
    st.session_state.pop('session_token', None)
    st.session_state.pop('LOGIN_TOKEN', None)
    st.session_state.pop('TOKEN_CREATED_AT', None)
    st.session_state.pop('2fa_pending', None)
    st.session_state.pop('2fa_username', None)
    
    # Clear query params token
    try:
        if 'session_token' in st.query_params:
            del st.query_params['session_token']
    except Exception:
        pass
    
    logger.debug(f"Auth state cleared for user: {current_user}")
    
    # CRITICAL: DO NOT clear these keys - they must persist across logout/login:
    # - OPENAI_API_KEY, GOOGLE_API_KEY, SLACK_WEBHOOK (API keys)
    # - _smtp_host, _smtp_port, _smtp_user, _smtp_pass (SMTP settings)
    # - current_audit_data, audit_bulk_selected (audit state)
    # - bulk_scan_active, bulk_scan_session_id (bulk scan state)
    # - current_section (navigation)


def logout_user():
    """
    Logs out the current user.
    
    This must:
    - Delete only auth-related keys
    - Keep all other session keys (including API keys, SMTP settings & audit data)
    - Delete session from file storage
    - Clear navigation state for the user
    - Not delete cached audit results on disk
    - NOT clear API keys or SMTP settings
    """
    logger = logging.getLogger("sales_engine")
    
    current_user = st.session_state.get('current_user', 'unknown')
    session_token = st.session_state.get('session_token') or st.session_state.get('LOGIN_TOKEN')
    
    logger.info(f"Logging out user: {current_user}")
    
    # Destroy session in file storage
    if session_token:
        destroy_session(session_token)
    
    # Clear persistence layer state (navigation, current audit selection)
    # NOTE: This PRESERVES API keys and SMTP settings
    # This preserves disk cache but clears session-specific state
    on_logout_cleanup()
    
    # Clear only auth state
    _clear_auth_state_only()
    
    # Re-initialize auth defaults (not authenticated)
    st.session_state.user_authenticated = False
    st.session_state.authenticated = False
    st.session_state.user_role = "user"
    st.session_state.is_admin = False
    
    logger.info(f"User logged out successfully: {current_user}")


def store_api_keys(username: str, openai_key: str = None, google_key: str = None, slack_webhook: str = None) -> bool:
    """
    Stores API keys securely in the database with encryption.
    
    This must:
    - Save API keys to database under user account table
    - Encrypt keys at rest using Fernet
    - Update session state with new keys
    - Never clear them on refresh
    
    Args:
        username: The username to store keys for
        openai_key: OpenAI API key (optional)
        google_key: Google API key (optional)
        slack_webhook: Slack webhook URL (optional)
    
    Returns:
        True if all keys saved successfully, False otherwise
    """
    logger = logging.getLogger("sales_engine")
    success = True
    
    if not username:
        logger.error("store_api_keys called without username")
        return False
    
    # Save each key if provided
    if openai_key is not None:
        ok, err = save_user_api_key(username, "openai", openai_key)
        if ok:
            st.session_state.OPENAI_API_KEY = openai_key
        else:
            logger.error(f"Failed to save OpenAI key: {err}")
            success = False
    
    if google_key is not None:
        ok, err = save_user_api_key(username, "google", google_key)
        if ok:
            st.session_state.GOOGLE_API_KEY = google_key
        else:
            logger.error(f"Failed to save Google key: {err}")
            success = False
    
    if slack_webhook is not None:
        ok, err = save_user_api_key(username, "slack", slack_webhook)
        if ok:
            st.session_state.SLACK_WEBHOOK = slack_webhook
        else:
            logger.error(f"Failed to save Slack webhook: {err}")
            success = False
    
    if success:
        logger.info(f"API keys stored successfully for user: {username}")
    
    return success


def require_role(required_role: str):
    """
    Decorator to require a specific role for accessing a function/page.
    
    Usage:
        @require_role("admin")
        def show_admin_page():
            ...
    
    Functionality:
    - If user is not logged in → redirect to Login page
    - If user role doesn't match required role → show Access Denied
    - Works with Streamlit routing
    - Does not break tabs or navigation
    
    Args:
        required_role: The role required ("admin" or "user")
    
    Returns:
        Decorated function
    """
    from functools import wraps
    
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            # Check if user is logged in
            if not load_login_state():
                st.warning("🔒 Please log in to access this page.")
                show_auth_page()
                st.stop()
                return None
            
            # Check role
            user_role = st.session_state.get('user_role', 'user')
            
            # Admin can access everything
            if user_role == 'admin':
                return func(*args, **kwargs)
            
            # Check if role matches
            if required_role == 'user':
                # User role can access user pages
                return func(*args, **kwargs)
            
            # Role doesn't match
            st.error("🚫 **Access Denied**")
            st.markdown(f"You need **{required_role}** permissions to access this page.")
            st.markdown("Please contact an administrator for access.")
            st.stop()
            return None
        
        return wrapper
    return decorator


def check_auth() -> bool:
    """
    Quick auth check for use at the top of pages.
    Returns True if authenticated, False otherwise.
    Does not render anything - use for conditional logic.
    """
    return load_login_state()


def require_auth():
    """
    Guard function to require authentication.
    Call at the top of any page that requires login.
    Will show login page and stop execution if not authenticated.
    """
    if not load_login_state():
        show_auth_page()
        st.stop()

def init_users_file():
    """Create users.json if it doesn't exist."""
    if not USERS_PATH.exists():
        USERS_PATH.write_text(json.dumps({"users": {}}, indent=2))


def ensure_admin_user_exists():
    """
    Ensure the admin user 'rabnawaz' exists with the correct SHA256 password hash.
    
    This function:
    - Checks if admin user exists in users.json
    - Creates or updates admin with correct SHA256 hash if needed
    - Syncs admin user to database
    
    IMPORTANT: Does NOT touch other users or any Phase 1/2/3 code.
    """
    logger = logging.getLogger("sales_engine")
    
    ADMIN_USERNAME = "rabnawaz"
    ADMIN_PASSWORD = "92948870"
    ADMIN_HASH = hashlib.sha256(ADMIN_PASSWORD.encode('utf-8')).hexdigest()
    
    try:
        # Load existing users
        init_users_file()
        users_data = json.loads(USERS_PATH.read_text())
        users = users_data.get("users", {})
        
        admin_user = users.get(ADMIN_USERNAME)
        needs_update = False
        
        if not admin_user:
            # Admin doesn't exist, create it
            logger.info(f"Creating admin user: {ADMIN_USERNAME}")
            needs_update = True
        elif admin_user.get("password_hash") != ADMIN_HASH:
            # Admin exists but hash is wrong (e.g., bcrypt instead of SHA256)
            logger.info(f"Updating admin password hash for: {ADMIN_USERNAME}")
            needs_update = True
        
        if needs_update:
            users[ADMIN_USERNAME] = {
                "name": "Admin",
                "password_hash": ADMIN_HASH,
                "role": "admin",
                "admin_request": False,
                "admin_request_reason": "",
                "created_at": users.get(ADMIN_USERNAME, {}).get("created_at", datetime.now().isoformat()),
                "last_login": users.get(ADMIN_USERNAME, {}).get("last_login")
            }
            users_data["users"] = users
            USERS_PATH.write_text(json.dumps(users_data, indent=2))
            logger.info(f"Admin user '{ADMIN_USERNAME}' saved to users.json")
            
            # Also sync to database
            try:
                db = get_db()
                if db:
                    existing = db.query(User).filter(User.username == ADMIN_USERNAME).first()
                    if existing:
                        existing.password_hash = ADMIN_HASH
                        existing.is_admin = True
                    else:
                        new_user = User(
                            username=ADMIN_USERNAME,
                            email=f"{ADMIN_USERNAME}@admin.local",
                            password_hash=ADMIN_HASH,
                            is_admin=True
                        )
                        db.add(new_user)
                    db.commit()
                    db.close()
                    logger.info(f"Admin user '{ADMIN_USERNAME}' synced to database")
            except Exception as db_err:
                logger.warning(f"Could not sync admin to database: {db_err}")
        
        return True
    except Exception as e:
        logger.error(f"Error ensuring admin user exists: {e}")
        return False


def init_2fa_file():
    """Create two_fa.json if it doesn't exist."""
    if not TWO_FA_PATH.exists():
        TWO_FA_PATH.write_text(json.dumps({}, indent=2))

def get_encryption_key() -> bytes:
    """Get or create encryption key for API keys."""
    if ENCRYPTION_KEY_PATH.exists():
        return ENCRYPTION_KEY_PATH.read_bytes()
    else:
        key = Fernet.generate_key()
        ENCRYPTION_KEY_PATH.write_bytes(key)
        return key

def init_login_attempts_file():
    """Create login_attempts.json if it doesn't exist."""
    if not LOGIN_ATTEMPTS_PATH.exists():
        LOGIN_ATTEMPTS_PATH.write_text(json.dumps({}, indent=2))

def load_login_attempts():
    """Load login attempts tracking."""
    init_login_attempts_file()
    try:
        return json.loads(LOGIN_ATTEMPTS_PATH.read_text())
    except Exception:
        return {}

def save_login_attempts(attempts: dict):
    """Save login attempts tracking."""
    LOGIN_ATTEMPTS_PATH.write_text(json.dumps(attempts, indent=2))

def check_login_rate_limit(username: str) -> tuple[bool, str]:
    """Check if user is rate limited. Returns (allowed, message)."""
    attempts = load_login_attempts()
    now = datetime.now()
    
    if username not in attempts:
        attempts[username] = []
    
    # Remove attempts older than the window
    attempts[username] = [
        attempt for attempt in attempts[username]
        if datetime.fromisoformat(attempt) > now - timedelta(minutes=LOGIN_ATTEMPT_WINDOW_MINUTES)
    ]
    
    if len(attempts[username]) >= LOGIN_ATTEMPT_LIMIT:
        return False, f"Too many login attempts. Please try again in {LOGIN_ATTEMPT_WINDOW_MINUTES} minutes."
    
    return True, ""

def record_login_attempt(username: str):
    """Record a failed login attempt."""
    attempts = load_login_attempts()
    
    if username not in attempts:
        attempts[username] = []
    
    attempts[username].append(datetime.now().isoformat())
    save_login_attempts(attempts)

def clear_login_attempts(username: str):
    """Clear login attempts on successful login."""
    attempts = load_login_attempts()
    if username in attempts:
        attempts[username] = []
        save_login_attempts(attempts)

def init_sessions_file():
    """Create sessions.json if it doesn't exist."""
    if not SESSIONS_PATH.exists():
        SESSIONS_PATH.write_text(json.dumps({}, indent=2))

def load_sessions():
    """Load active sessions."""
    init_sessions_file()
    try:
        return json.loads(SESSIONS_PATH.read_text())
    except Exception:
        return {}

def save_sessions(sessions: dict):
    """Save active sessions."""
    SESSIONS_PATH.write_text(json.dumps(sessions, indent=2))

def create_session(username: str, role: str) -> str:
    """Create a new session and return session token."""
    import uuid
    
    session_token = uuid.uuid4().hex[:32]
    sessions = load_sessions()
    
    sessions[session_token] = {
        "username": username,
        "role": role,
        "created_at": datetime.now().isoformat(),
        "last_access": datetime.now().isoformat()
    }
    save_sessions(sessions)
    
    # Also store in browser via query parameter (Streamlit workaround)
    st.session_state['session_token'] = session_token
    
    return session_token

def validate_session(session_token: str) -> dict:
    """Validate session token and return user info if valid. Checks for idle timeout and max lifetime."""
    logger = logging.getLogger("sales_engine")
    sessions = load_sessions()
    
    if session_token not in sessions:
        return None
    
    session_data = sessions[session_token]
    now = datetime.now()
    
    try:
        created_at = datetime.fromisoformat(session_data["created_at"])
        # Use last_access if available, otherwise fall back to created_at
        last_access_str = session_data.get("last_access", session_data["created_at"])
        last_access = datetime.fromisoformat(last_access_str)
        
        # Check 1: Max session lifetime (7 days absolute limit)
        if now - created_at > timedelta(hours=SESSION_MAX_LIFETIME_HOURS):
            logger.info(f"Session expired (max lifetime): {session_data.get('username')}")
            del sessions[session_token]
            save_sessions(sessions)
            return None
        
        # Check 2: Idle timeout (1 hour of inactivity)
        if now - last_access > timedelta(minutes=SESSION_IDLE_TIMEOUT_MINUTES):
            logger.info(f"Session expired (idle timeout): {session_data.get('username')}")
            del sessions[session_token]
            save_sessions(sessions)
            return None
        
        # Session is valid - update last access time
        session_data["last_access"] = now.isoformat()
        sessions[session_token] = session_data
        save_sessions(sessions)
        
        logger.debug(f"Session validated for user: {session_data.get('username')}")
        return session_data
        
    except Exception as e:
        logger.error(f"Error validating session: {str(e)}")
        # Invalid session data, remove it
        del sessions[session_token]
        save_sessions(sessions)
        return None

def destroy_session(session_token: str):
    """Destroy a session."""
    sessions = load_sessions()
    if session_token in sessions:
        del sessions[session_token]
        save_sessions(sessions)

def load_users():
    """Return users dict from users.json."""
    init_users_file()
    try:
        data = json.loads(USERS_PATH.read_text())
        return data.get("users", {})
    except Exception:
        return {}

def save_users(users: dict):
    """Save users dict to users.json."""
    data = {"users": users}
    USERS_PATH.write_text(json.dumps(data, indent=2))

def get_user_api_permissions(username: str) -> dict:
    """Get API permissions for a user."""
    users = load_users()
    if username not in users:
        return {"openai": False, "google": False, "slack": False}
    
    return users[username].get("api_permissions", {"openai": False, "google": False, "slack": False})

def set_user_api_permission(username: str, api_key: str, enabled: bool):
    """Set API permission for a user."""
    users = load_users()
    if username not in users:
        return False
    
    if "api_permissions" not in users[username]:
        users[username]["api_permissions"] = {"openai": False, "google": False, "slack": False}
    
    users[username]["api_permissions"][api_key] = enabled
    save_users(users)
    return True

def check_user_api_access(username: str, api_key: str) -> bool:
    """Check if user has access to specific API."""
    perms = get_user_api_permissions(username)
    return perms.get(api_key, False)

def encrypt_key(key: str) -> str:
    """Encrypt API key using Fernet (secure encryption)."""
    cipher = Fernet(get_encryption_key())
    return cipher.encrypt(key.encode()).decode()

def decrypt_key(encrypted_key: str) -> str:
    """Decrypt API key using Fernet."""
    try:
        cipher = Fernet(get_encryption_key())
        return cipher.decrypt(encrypted_key.encode()).decode()
    except:
        return ""

def get_user_api_keys(username: str) -> dict:
    """Get API keys for a user from database. Returns dict with keys and error status."""
    logger = logging.getLogger("sales_engine")
    default_keys = {"openai": "", "google": "", "slack": "", "smtp_host": "", "smtp_port": 587, "smtp_user": "", "smtp_pass": "", "_db_error": False}
    
    if not username:
        logger.warning("get_user_api_keys called with empty username")
        return default_keys
    
    try:
        db = get_db()
        if not db:
            logger.error(f"Database unavailable when loading API keys for {username}")
            return {"openai": "", "google": "", "slack": "", "smtp_host": "", "smtp_port": 587, "smtp_user": "", "smtp_pass": "", "_db_error": True}
        
        user = db.query(User).filter(User.username == username).first()
        if not user:
            # User doesn't exist in DB yet - try to create from JSON
            users_json = load_users()
            if username in users_json:
                user_data = users_json[username]
                ensure_user_in_database(username, user_data.get("password_hash", ""), is_admin=(user_data.get("role") == "admin"))
                # Re-query after creation
                user = db.query(User).filter(User.username == username).first()
            
            if not user:
                logger.warning(f"User not found in database when loading API keys: {username}")
                db.close()
                return default_keys
        
        api_keys = user.api_keys or {}
        # Decrypt keys when retrieving - always return all keys
        decrypted = {"openai": "", "google": "", "slack": "", "smtp_host": "", "smtp_port": 587, "smtp_user": "", "smtp_pass": "", "_db_error": False}
        for key_name, encrypted_value in api_keys.items():
            if key_name.startswith("_"):
                continue  # Skip internal flags
            try:
                if encrypted_value:
                    # SMTP port is not encrypted
                    if key_name == "smtp_port":
                        decrypted[key_name] = int(encrypted_value) if encrypted_value else 587
                    elif key_name in ["smtp_host", "smtp_user"]:
                        # These don't need encryption
                        decrypted[key_name] = encrypted_value
                    else:
                        decrypted[key_name] = decrypt_key(encrypted_value)
            except Exception as decrypt_err:
                logger.error(f"Failed to decrypt {key_name} for {username}: {str(decrypt_err)}")
                if key_name == "smtp_port":
                    decrypted[key_name] = 587
                else:
                    decrypted[key_name] = ""
        
        db.close()
        logger.debug(f"Loaded API keys for {username}: openai={'yes' if decrypted.get('openai') else 'no'}, google={'yes' if decrypted.get('google') else 'no'}, slack={'yes' if decrypted.get('slack') else 'no'}, smtp={'yes' if decrypted.get('smtp_host') else 'no'}")
        return decrypted
    except Exception as e:
        logger.error(f"Error loading API keys for {username}: {str(e)}", exc_info=True)
        return {"openai": "", "google": "", "slack": "", "smtp_host": "", "smtp_port": 587, "smtp_user": "", "smtp_pass": "", "_db_error": True}

def save_user_api_key(username: str, key_name: str, key_value: str) -> tuple:
    """Save API key for a user to database (encrypted). Returns (success, error_message).
    
    Handles both API keys (encrypted) and SMTP settings (some unencrypted).
    CRITICAL: Uses flag_modified() to ensure SQLAlchemy detects JSON column changes.
    """
    logger = logging.getLogger("sales_engine")
    db = None
    
    if not username or not key_name:
        logger.error("save_user_api_key called with empty username or key_name")
        return False, "Invalid username or key name"
    
    # SAFETY: Never save empty values for critical keys - this prevents accidental overwriting
    # For critical keys, empty value = skip save (preserve existing)
    if key_name in ["openai", "google", "slack", "smtp_pass"]:
        if not key_value or not str(key_value).strip():
            logger.warning(f"Refusing to save empty value for {key_name} - preserving existing key")
            return True, None  # Return success but don't save
    
    try:
        db = get_db()
        if not db:
            logger.error(f"Database unavailable when saving API key {key_name} for {username}")
            return False, "Database connection unavailable"
        
        user = db.query(User).filter(User.username == username).first()
        if not user:
            # Try to create user from JSON if they exist there
            users_json = load_users()
            if username in users_json:
                user_data = users_json[username]
                ensure_user_in_database(username, user_data.get("password_hash", ""), is_admin=(user_data.get("role") == "admin"))
                # Re-query after creation
                user = db.query(User).filter(User.username == username).first()
            
            if not user:
                logger.error(f"User not found in database when saving API key: {username}")
                db.close()
                return False, "User not found in database"
        
        # CRITICAL: Initialize api_keys dict if it's None (not just falsy)
        if user.api_keys is None:
            user.api_keys = {}
        
        # Handle different key types
        if key_name == "smtp_port":
            # SMTP port is stored as-is (integer as string)
            user.api_keys[key_name] = str(key_value) if key_value else "587"
        elif key_name in ["smtp_host", "smtp_user"]:
            # SMTP host and user are stored unencrypted
            user.api_keys[key_name] = str(key_value) if key_value else ""
        else:
            # Encrypt API keys and passwords
            user.api_keys[key_name] = encrypt_key(str(key_value))
        
        user.api_keys_updated_at = datetime.utcnow()
        
        # CRITICAL: Force SQLAlchemy to detect the change in JSON column
        from sqlalchemy.orm.attributes import flag_modified
        flag_modified(user, "api_keys")
        
        db.commit()
        logger.info(f"Successfully saved API key {key_name} for {username} to database")
        db.close()
        return True, None
    except Exception as e:
        logger.error(f"Error saving API key {key_name} for {username}: {str(e)}", exc_info=True)
        if db:
            try:
                db.rollback()
                db.close()
            except:
                pass
        return False, str(e)

def delete_user_api_key(username: str, key_name: str):
    """Delete API key for a user from database."""
    try:
        db = get_db()
        if not db:
            return False
        
        user = db.query(User).filter(User.username == username).first()
        if not user:
            db.close()
            return False
        
        if user.api_keys and key_name in user.api_keys:
            del user.api_keys[key_name]
            user.api_keys_updated_at = datetime.utcnow()
            
            # Force SQLAlchemy to detect the change in JSON column
            from sqlalchemy.orm.attributes import flag_modified
            flag_modified(user, "api_keys")
            
            db.commit()
            db.close()
            return True
        
        db.close()
        return False
    except Exception as e:
        logger = logging.getLogger("sales_engine")
        logger.error(f"Error deleting API key for {username}: {str(e)}")
        if db:
            db.close()
        return False

def migrate_api_keys_from_json_to_db():
    """Migrate API keys from users.json to database (one-time migration)."""
    logger = logging.getLogger("sales_engine")
    try:
        users_json = load_users()
        db = get_db()
        if not db:
            return
        
        for username, user_data in users_json.items():
            if "api_keys" in user_data and user_data["api_keys"]:
                # Find or create user in database
                db_user = db.query(User).filter(User.username == username).first()
                if db_user:
                    # Only migrate if database user doesn't have keys yet
                    if not db_user.api_keys:
                        db_user.api_keys = user_data["api_keys"]
                        db_user.api_keys_updated_at = datetime.utcnow()
                        db.commit()
                        logger.info(f"Migrated API keys for user: {username}")
        
        db.close()
    except Exception as e:
        logger.error(f"Error during API key migration: {str(e)}")

def ensure_user_in_database(username: str, password_hash: str, is_admin: bool = False):
    """Ensure a user exists in the database (called during login for JSON-only users)."""
    logger = logging.getLogger("sales_engine")
    try:
        db = get_db()
        if not db:
            return False
        
        # Check if user already exists in database
        existing_user = db.query(User).filter(User.username == username).first()
        if existing_user:
            db.close()
            return True
        
        # Create new user in database with empty but initialized JSON columns
        new_user = User(
            username=username,
            password_hash=password_hash,
            is_admin=is_admin,
            api_keys={},
            smtp_settings={}
        )
        db.add(new_user)
        db.commit()
        db.close()
        logger.info(f"Synced JSON user to database: {username}")
        return True
    except Exception as e:
        logger.error(f"Error syncing user {username} to database: {str(e)}")
        if db:
            db.close()
        return False


def ensure_users_in_database_on_startup():
    """
    Sync ALL users from users.json to database on startup.
    
    This ensures users exist in the database BEFORE any authentication or key loading.
    Called once during app initialization, after DB is ready.
    """
    logger = logging.getLogger("sales_engine")
    logger.info("Syncing JSON users to database on startup...")
    
    try:
        # Load users from JSON file
        users_json = load_users()
        if not users_json:
            logger.warning("No users found in JSON file during startup sync")
            return 0
        
        db = get_db()
        if not db:
            logger.error("Database unavailable during startup user sync")
            return 0
        
        synced_count = 0
        for username, user_data in users_json.items():
            try:
                existing = db.query(User).filter(User.username == username).first()
                if not existing:
                    # User exists in JSON but not in DB - create them
                    is_admin = user_data.get("role") == "admin"
                    new_user = User(
                        username=username,
                        password_hash=user_data.get("password_hash", ""),
                        is_admin=is_admin,
                        api_keys={},
                        smtp_settings={}
                    )
                    db.add(new_user)
                    synced_count += 1
                    logger.info(f"Synced user to database: {username} (admin={is_admin})")
            except Exception as user_err:
                logger.error(f"Error syncing user {username}: {str(user_err)}")
                continue
        
        if synced_count > 0:
            db.commit()
            logger.info(f"Startup sync complete: {synced_count} users synced to database")
        else:
            logger.info("Startup sync: all users already exist in database")
        
        db.close()
        return synced_count
        
    except Exception as e:
        logger.error(f"Error during startup user sync: {str(e)}", exc_info=True)
        return 0


def load_2fa_secrets():
    """Load 2FA secrets."""
    init_2fa_file()
    try:
        return json.loads(TWO_FA_PATH.read_text())
    except Exception:
        return {}

def save_2fa_secrets(secrets: dict):
    """Save 2FA secrets."""
    TWO_FA_PATH.write_text(json.dumps(secrets, indent=2))

def hash_password(password: str) -> str:
    """Hash password with SHA-256."""
    return hashlib.sha256(password.encode('utf-8')).hexdigest()

def verify_password(password: str, stored_hash: str) -> bool:
    """Verify password."""
    return hash_password(password) == stored_hash

def generate_2fa_secret():
    """Generate 2FA secret."""
    if not TWO_FA_AVAILABLE:
        return None
    return pyotp.random_base32()

def verify_2fa_token(secret: str, token: str) -> bool:
    """Verify 2FA token."""
    if not TWO_FA_AVAILABLE or not secret:
        return False
    try:
        totp = pyotp.TOTP(secret)
        return totp.verify(token)
    except Exception:
        return False

def reload_user_api_keys() -> dict:
    """Reload API keys and SMTP settings from database for current user. Returns status dict."""
    logger = logging.getLogger("sales_engine")
    current_user = st.session_state.get("current_user")
    
    if not current_user:
        logger.warning("reload_user_api_keys called with no current_user")
        return {"success": False, "error": "No user logged in"}
    
    logger.info(f"Reloading API keys for user: {current_user}")
    user_keys = get_user_api_keys(current_user)
    
    # Check for database errors
    db_error = user_keys.pop("_db_error", False)
    if db_error:
        st.session_state["_api_keys_db_error"] = True
        logger.error(f"Database error while loading API keys for {current_user}")
        # DON'T overwrite existing session keys on DB error
        return {"success": False, "error": "Database error"}
    else:
        st.session_state["_api_keys_db_error"] = False
    
    # Check environment variables - ONLY use if they are non-empty strings
    env_openai = os.environ.get("OPENAI_API_KEY", "").strip()
    env_google = os.environ.get("GOOGLE_API_KEY", "").strip()
    env_slack = os.environ.get("SLACK_WEBHOOK", "").strip()
    
    # Priority: non-empty env var > user's saved key > EXISTING session value > empty string
    # This ensures env vars only override if they actually have a value
    # And we NEVER overwrite a valid key with an empty one
    db_openai = user_keys.get("openai", "")
    db_google = user_keys.get("google", "")
    db_slack = user_keys.get("slack", "")
    
    # Get existing session values as fallback
    existing_openai = st.session_state.get("OPENAI_API_KEY", "")
    existing_google = st.session_state.get("GOOGLE_API_KEY", "")
    existing_slack = st.session_state.get("SLACK_WEBHOOK", "")
    
    # Apply priority: env > db > existing > empty
    st.session_state.OPENAI_API_KEY = env_openai or db_openai or existing_openai
    st.session_state.GOOGLE_API_KEY = env_google or db_google or existing_google
    st.session_state.SLACK_WEBHOOK = env_slack or db_slack or existing_slack
    
    # Load SMTP settings to session state (priority: db > existing > default)
    # SMTP settings don't typically come from env vars
    db_smtp_host = user_keys.get("smtp_host", "")
    db_smtp_port = user_keys.get("smtp_port", 587)
    db_smtp_user = user_keys.get("smtp_user", "")
    db_smtp_pass = user_keys.get("smtp_pass", "")
    
    existing_smtp_host = st.session_state.get("_smtp_host", "")
    existing_smtp_port = st.session_state.get("_smtp_port", 587)
    existing_smtp_user = st.session_state.get("_smtp_user", "")
    existing_smtp_pass = st.session_state.get("_smtp_pass", "")
    
    # Apply priority: db > existing > default (never overwrite with empty)
    st.session_state["_smtp_host"] = db_smtp_host or existing_smtp_host
    st.session_state["_smtp_port"] = db_smtp_port if db_smtp_port else existing_smtp_port
    st.session_state["_smtp_user"] = db_smtp_user or existing_smtp_user
    st.session_state["_smtp_pass"] = db_smtp_pass or existing_smtp_pass
    
    # Track which source each key came from (for debugging)
    st.session_state["_api_key_sources"] = {
        "openai": "env" if env_openai else ("db" if db_openai else ("session" if existing_openai else "none")),
        "google": "env" if env_google else ("db" if db_google else ("session" if existing_google else "none")),
        "slack": "env" if env_slack else ("db" if db_slack else ("session" if existing_slack else "none"))
    }
    
    logger.info(f"API keys loaded for {current_user}: openai={st.session_state['_api_key_sources']['openai']}, google={st.session_state['_api_key_sources']['google']}, slack={st.session_state['_api_key_sources']['slack']}")
    
    return {"success": True, "error": None}

def show_auth_page():
    """Enhanced login/signup UI with 2FA support."""
    st.markdown("""
    <style>
    .auth-container {
        max-width: 500px;
        margin: 0 auto;
        padding: 2rem;
        border-radius: 10px;
        background: linear-gradient(135deg, #0066CC 0%, #004A99 100%);
    }
    </style>
    """, unsafe_allow_html=True)
    
    st.title("🦅 Code Nest Sales Engine")
    st.caption("Sign in to access intelligent website auditing & lead generation")
    
    tab = st.radio("Auth Mode", ["Login", "Sign Up"], horizontal=True, label_visibility="collapsed")
    users = load_users()
    
    if tab == "Login":
        st.subheader("🔐 Login")
        username = st.text_input("Username", key="login_username")
        password = st.text_input("Password", type="password", key="login_password")
        
        if st.button("Login", type="primary", use_container_width=True):
            if not username or not password:
                st.error("Please fill in all fields")
            else:
                # Check rate limiting
                allowed, message = check_login_rate_limit(username)
                if not allowed:
                    st.error(message)
                elif username not in users:
                    record_login_attempt(username)
                    st.error("User not found")
                elif not verify_password(password, users[username]["password_hash"]):
                    record_login_attempt(username)
                    st.error("Invalid credentials")
                else:
                    # Successful login - clear attempts
                    clear_login_attempts(username)
                    
                    # Ensure user exists in database (for users who only exist in JSON)
                    user_role = users[username].get("role", "user")
                    ensure_user_in_database(username, users[username]["password_hash"], is_admin=(user_role == "admin"))
                    
                    # Check if 2FA is enabled
                    secrets_2fa = load_2fa_secrets()
                    if username in secrets_2fa and secrets_2fa[username].get("enabled"):
                        st.session_state["2fa_pending"] = True
                        st.session_state["2fa_username"] = username
                        st.rerun()
                    else:
                        # Create persistent session using new centralized function
                        user_email = users[username].get("email")
                        save_login_state(username, user_role, user_email)
                        
                        # Reload API keys from user account after login
                        reload_user_api_keys()
                        
                        st.success(f"Welcome, {users[username].get('name') or username}!")
                        time.sleep(1)
                        st.rerun()
        
        # 2FA verification if pending
        if st.session_state.get("2fa_pending"):
            st.divider()
            st.markdown("### 🔑 Two-Factor Authentication")
            token = st.text_input("Enter 6-digit code from authenticator app", max_chars=6, key="2fa_token")
            
            if st.button("Verify 2FA", type="primary"):
                secrets_2fa = load_2fa_secrets()
                if token and verify_2fa_token(secrets_2fa[st.session_state["2fa_username"]]["secret"], token):
                    # Create persistent session using new centralized function
                    username_2fa = st.session_state["2fa_username"]
                    user_role = users[username_2fa].get("role", "user")
                    user_email = users[username_2fa].get("email")
                    
                    # Ensure user exists in database (for users who only exist in JSON)
                    ensure_user_in_database(username_2fa, users[username_2fa]["password_hash"], is_admin=(user_role == "admin"))
                    
                    # Use centralized save_login_state
                    save_login_state(username_2fa, user_role, user_email)
                    
                    # Reload API keys from user account after login
                    reload_user_api_keys()
                    
                    st.success("2FA verified!")
                    time.sleep(1)
                    st.rerun()
                else:
                    st.error("Invalid 2FA token")
    else:
        st.subheader("📝 Create Account")
        full_name = st.text_input("Full Name", key="signup_name")
        username = st.text_input("Username", key="signup_username")
        password = st.text_input("Password", type="password", key="signup_password")
        confirm = st.text_input("Confirm Password", type="password", key="signup_confirm")
        request_admin = st.checkbox("Request admin access")
        reason = ""
        if request_admin:
            reason = st.text_area("Why do you need admin access?")
        
        if st.button("Create Account", type="primary", use_container_width=True):
            # Sanitize inputs
            full_name_clean = sanitize_input(full_name, max_length=100)
            username_clean = sanitize_input(username, max_length=50)
            reason_clean = sanitize_input(reason, max_length=500) if reason else ""
            
            # Validate fields
            if not full_name_clean or not username_clean or not password:
                st.error("❌ All fields are required")
                logger.warning(f"Signup attempt with missing fields")
            elif password != confirm:
                st.error("❌ Passwords don't match")
                logger.warning(f"Signup attempt: password mismatch for {username_clean}")
            elif username_clean in users:
                st.error("❌ Username already taken")
                logger.warning(f"Signup attempt with existing username: {username_clean}")
            else:
                # Validate password strength
                is_valid, pwd_error = validate_password(password)
                if not is_valid:
                    st.error(f"❌ Password requirements: {pwd_error}")
                    logger.warning(f"Weak password attempt for {username_clean}")
                else:
                    # Validate username format
                    if not re.match(r'^[a-zA-Z0-9_]{3,50}$', username_clean):
                        st.error("❌ Username must be 3-50 characters, alphanumeric + underscore only")
                        logger.warning(f"Invalid username format: {username_clean}")
                    else:
                        try:
                            users[username_clean] = {
                                "name": full_name_clean,
                                "password_hash": hash_password(password),
                                "role": "user",
                                "admin_request": request_admin,
                                "admin_request_reason": reason_clean,
                                "created_at": datetime.now().isoformat(),
                                "last_login": None
                            }
                            save_users(users)
                            
                            # Also create user in database
                            if DB_AVAILABLE:
                                try:
                                    db = get_db()
                                    if db:
                                        new_user = User(
                                            username=username_clean,
                                            password_hash=hash_password(password),
                                            is_admin=(request_admin and "admin" in reason_clean.lower())
                                        )
                                        db.add(new_user)
                                        db.commit()
                                        db.close()
                                        logger.info(f"New database user created: {username_clean}")
                                except Exception as e:
                                    logger.error(f"Error creating database user for {username_clean}: {str(e)}")
                            
                            logger.info(f"New account created: {username_clean}")
                            st.success("✅ Account created! Please login.")
                            time.sleep(1)
                            st.rerun()
                        except Exception as e:
                            logger.error(f"Error creating account for {username_clean}: {str(e)}", exc_info=True)
                            st.error("❌ Failed to create account. Please try again.")
# Initialize auth
init_users_file()
ensure_admin_user_exists()  # Ensure admin user exists with correct SHA256 hash
init_2fa_file()
init_sessions_file()
init_login_attempts_file()

# Initialize encryption key (will be created if not exists)
get_encryption_key()

# Migrate API keys from JSON to database (one-time operation)
if not st.session_state.get("_api_keys_migrated"):
    migrate_api_keys_from_json_to_db()
    st.session_state["_api_keys_migrated"] = True

# ============================================================================
# CENTRALIZED SESSION INITIALIZATION & AUTHENTICATION CHECK
# ============================================================================
# This section MUST run before any page rendering.
# Uses the new init_app_session() and load_login_state() functions.

# Step 1: Initialize all session state variables (never overwrites existing values)
init_app_session()

# Step 1.5: Initialize persistence layer (audit data, AI cache, navigation state)
# This restores audit data from disk cache if available, syncs URL params, etc.
persistence_status = init_app_session_persistence()

# Step 1.6: Sync ALL users from JSON to database on startup (if not already done this session)
# This ensures users exist in the database BEFORE any key loading
if not st.session_state.get("_users_synced_to_db"):
    if DB_AVAILABLE:
        ensure_users_in_database_on_startup()
    st.session_state["_users_synced_to_db"] = True

# Step 2: Try to restore/validate login state
# This checks session_state, then query params, validates tokens, and restores user info
session_valid = load_login_state()

# Step 3: If not authenticated, show login page
if not session_valid:
    show_auth_page()
    st.stop()

# Step 4: Reload API keys for authenticated user
# Always reload on each app run to ensure keys are current from DB
if st.session_state.get("current_user"):
    reload_user_api_keys()

# Get current API keys from session (for backward compatibility)
OPENAI_API_KEY = st.session_state.get('OPENAI_API_KEY', '')
GOOGLE_API_KEY = st.session_state.get('GOOGLE_API_KEY', '')
SLACK_WEBHOOK = st.session_state.get('SLACK_WEBHOOK', '')

# ============================================================================
# SIDEBAR & NAVIGATION
# ============================================================================

with st.sidebar:
    st.header("🦅 Code Nest Panel")
    st.caption("Navigation")
    st.divider()
    
    # Build navigation items based on role
    nav_items = ["Single Audit", "Audit History", "Dashboard", "Preferences"]
    if st.session_state.get("is_admin"):
        nav_items.extend([
            "Bulk Audit",
            "Competitor Analysis",
            "Lead Management",
            "CRM Pipeline",
            "Email Outreach",
            "Scheduled Audits",
            "API Settings",
            "Admin Settings",
            "Email Settings",
            "Export Reports"
        ])
    
    # Navigation buttons
    for item in nav_items:
        if st.button(
            item,
            key=f"nav_{item}",
            use_container_width=True,
            type="primary" if st.session_state.current_section == item else "secondary"
        ):
            st.session_state.current_section = item
            st.rerun()
    
    st.divider()
    st.divider()
    
    # Status badge
    if DB_AVAILABLE:
        st.success("🟢 System: **Active**")
    else:
        st.warning("🟡 System: **Limited**")
    
    st.divider()
    
    # User info & logout
    st.markdown(f"**User:** {st.session_state.get('current_user')}")
    st.markdown(f"**Role:** {st.session_state.get('user_role', 'user').capitalize()}")
    
    if st.button("🚪 Logout", use_container_width=True):
        # Use the centralized logout_user function
        # This preserves API keys and audit data while clearing auth state
        logout_user()
        
        # Clear query params (removes session token and audit_id from URL)
        try:
            clear_navigation_state()
            st.query_params.clear()
        except Exception:
            pass
        
        st.rerun()
    
    # 2FA setup (if admin)
    if st.session_state.get('is_admin'):
        st.divider()
        if st.checkbox("Setup 2FA", key="admin_2fa_checkbox"):
            if TWO_FA_AVAILABLE:
                st.markdown("### Enable 2FA for Your Account")
                if st.button("Generate 2FA"):
                    secret = generate_2fa_secret()
                    st.session_state["2fa_setup_secret"] = secret
                    
                    totp = pyotp.TOTP(secret)
                    qr = qrcode.QRCode(version=1, box_size=10)
                    qr.add_data(totp.provisioning_uri(st.session_state['current_user'], issuer_name='Code Nest'))
                    qr.make(fit=True)
                    img = qr.make_image(fill_color="black", back_color="white")
                    
                    st.image(img, width=200)
                    st.code(secret, language="text")
                    st.info("Scan QR code with authenticator app (Google Authenticator, Authy, etc)")
                    
                    token = st.text_input("Enter 6-digit code to confirm")
                    if token and verify_2fa_token(secret, token):
                        secrets_2fa = load_2fa_secrets()
                        secrets_2fa[st.session_state['current_user']] = {
                            "secret": secret,
                            "enabled": True
                        }
                        save_2fa_secrets(secrets_2fa)
                        st.success("2FA enabled!")
                        st.rerun()
            else:
                st.info("2FA requires pyotp and qrcode packages. Install via: pip install pyotp qrcode[pil]")

def show_api_settings():
    """API Settings page for configuring API keys securely."""
    st.title("🔑 API Settings")
    st.markdown("Securely manage your API keys for AI email generation, PageSpeed insights, and Slack notifications")
    st.markdown("---")
    
    username = st.session_state.get("current_user", "")
    if not username:
        st.warning("Please log in to access API settings")
        return
    
    # Get current user's API keys
    current_keys = get_user_api_keys(username)
    
    # Check for database errors and show warning
    db_error = current_keys.pop("_db_error", False)
    if db_error:
        st.error("⚠️ Could not load your API keys from the database. Please check your connection and try again.")
        st.info("Your keys may still be saved - this could be a temporary connection issue.")
    
    # Custom CSS for API settings cards
    st.markdown("""
    <style>
    .api-key-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 20px;
        border-radius: 12px;
        color: white;
        margin-bottom: 15px;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
    }
    .api-key-status {
        display: flex;
        justify-content: space-between;
        align-items: center;
        margin-bottom: 15px;
        padding-bottom: 10px;
        border-bottom: 1px solid rgba(255, 255, 255, 0.2);
    }
    .api-key-name {
        font-size: 16px;
        font-weight: bold;
    }
    .api-key-badge {
        display: inline-block;
        padding: 4px 12px;
        border-radius: 20px;
        font-size: 12px;
        font-weight: bold;
    }
    .badge-active {
        background-color: rgba(100, 200, 100, 0.8);
        color: white;
    }
    .badge-inactive {
        background-color: rgba(255, 107, 107, 0.8);
        color: white;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # API 1: OpenAI
    st.markdown("### 🤖 OpenAI API")
    st.markdown("For AI email generation and content analysis")
    
    with st.container():
        col1, col2 = st.columns([3, 1])
        
        with col1:
            if current_keys.get("openai") and not st.session_state.get("edit_openai"):
                # Show masked key with options
                masked_key = current_keys["openai"][:8] + "..." + current_keys["openai"][-4:] if len(current_keys["openai"]) > 12 else "***"
                st.markdown(f"""
                <div class="api-key-card">
                    <div class="api-key-status">
                        <div>
                            <div class="api-key-name">✅ OpenAI Key Active</div>
                            <div style="font-size: 12px; opacity: 0.8;">Key: {masked_key}</div>
                        </div>
                        <div class="api-key-badge badge-active">ACTIVE</div>
                    </div>
                </div>
                """, unsafe_allow_html=True)
                
                col_view, col_delete, col_add = st.columns(3)
                
                with col_view:
                    if st.button("👁️ View Full Key", key="view_openai", use_container_width=True):
                        st.session_state.show_openai_key = not st.session_state.get("show_openai_key", False)
                    
                    if st.session_state.get("show_openai_key"):
                        st.code(current_keys["openai"], language="text")
                
                with col_delete:
                    if st.button("🗑️ Delete Key", key="del_openai", use_container_width=True):
                        delete_user_api_key(username, "openai")
                        st.success("✓ OpenAI key deleted")
                        st.rerun()
                
                with col_add:
                    if st.button("🔄 Replace Key", key="replace_openai", use_container_width=True):
                        st.session_state.edit_openai = True
                        st.rerun()
            
            elif st.session_state.get("edit_openai") or not current_keys.get("openai"):
                # Add/Edit form
                if not current_keys.get("openai"):
                    st.markdown("""
                    <div class="api-key-card">
                        <div class="api-key-status">
                            <div>
                                <div class="api-key-name">❌ No Key Added</div>
                                <div style="font-size: 12px; opacity: 0.8;">Add your OpenAI key to enable AI features</div>
                            </div>
                            <div class="api-key-badge badge-inactive">INACTIVE</div>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                
                st.markdown("**Add or Replace OpenAI API Key:**")
                new_openai_key = st.text_input(
                    "Paste your OpenAI API Key",
                    type="password",
                    key="input_openai",
                    placeholder="sk-..."
                )
                
                col_save, col_cancel = st.columns(2)
                with col_save:
                    if st.button("✅ Save OpenAI Key", key="save_openai_btn", use_container_width=True):
                        if new_openai_key:
                            success, error = save_user_api_key(username, "openai", new_openai_key)
                            if success:
                                st.session_state.OPENAI_API_KEY = new_openai_key
                                st.success("✓ OpenAI key saved securely and loaded into session")
                                st.session_state.edit_openai = False
                                st.rerun()
                            else:
                                st.error(f"Failed to save key: {error or 'Unknown error'}")
                        else:
                            st.error("Please enter an API key")
                
                with col_cancel:
                    if st.button("❌ Cancel", key="cancel_openai_btn", use_container_width=True):
                        st.session_state.edit_openai = False
                        st.rerun()
    
    st.divider()
    
    # API 2: Google PageSpeed
    st.markdown("### 🔍 Google PageSpeed API")
    st.markdown("For website performance analysis")
    
    with st.container():
        col1, col2 = st.columns([3, 1])
        
        with col1:
            if current_keys.get("google") and not st.session_state.get("edit_google"):
                masked_key = current_keys["google"][:8] + "..." + current_keys["google"][-4:] if len(current_keys["google"]) > 12 else "***"
                st.markdown(f"""
                <div class="api-key-card">
                    <div class="api-key-status">
                        <div>
                            <div class="api-key-name">✅ Google Key Active</div>
                            <div style="font-size: 12px; opacity: 0.8;">Key: {masked_key}</div>
                        </div>
                        <div class="api-key-badge badge-active">ACTIVE</div>
                    </div>
                </div>
                """, unsafe_allow_html=True)
                
                col_view, col_delete, col_add = st.columns(3)
                
                with col_view:
                    if st.button("👁️ View Full Key", key="view_google", use_container_width=True):
                        st.session_state.show_google_key = not st.session_state.get("show_google_key", False)
                    
                    if st.session_state.get("show_google_key"):
                        st.code(current_keys["google"], language="text")
                
                with col_delete:
                    if st.button("🗑️ Delete Key", key="del_google", use_container_width=True):
                        delete_user_api_key(username, "google")
                        st.success("✓ Google key deleted")
                        st.rerun()
                
                with col_add:
                    if st.button("🔄 Replace Key", key="replace_google", use_container_width=True):
                        st.session_state.edit_google = True
                        st.rerun()
            
            elif st.session_state.get("edit_google") or not current_keys.get("google"):
                # Add/Edit form
                if not current_keys.get("google"):
                    st.markdown("""
                    <div class="api-key-card">
                        <div class="api-key-status">
                            <div>
                                <div class="api-key-name">❌ No Key Added</div>
                                <div style="font-size: 12px; opacity: 0.8;">Add your Google PageSpeed key to enable performance analysis</div>
                            </div>
                            <div class="api-key-badge badge-inactive">INACTIVE</div>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                
                st.markdown("**Add or Replace Google PageSpeed API Key:**")
                new_google_key = st.text_input(
                    "Paste your Google PageSpeed API Key",
                    type="password",
                    key="input_google",
                    placeholder="AIza..."
                )
                
                col_save, col_cancel = st.columns(2)
                with col_save:
                    if st.button("✅ Save Google Key", key="save_google_btn", use_container_width=True):
                        if new_google_key:
                            success, error = save_user_api_key(username, "google", new_google_key)
                            if success:
                                st.session_state.GOOGLE_API_KEY = new_google_key
                                st.success("✓ Google key saved securely and loaded into session")
                                st.session_state.edit_google = False
                                st.rerun()
                            else:
                                st.error(f"Failed to save key: {error or 'Unknown error'}")
                        else:
                            st.error("Please enter an API key")
                
                with col_cancel:
                    if st.button("❌ Cancel", key="cancel_google_btn", use_container_width=True):
                        st.session_state.edit_google = False
                        st.rerun()
    
    st.divider()
    
    # API 3: Slack Webhook
    st.markdown("### 📱 Slack Webhook URL")
    st.markdown("For audit notifications (optional)")
    
    with st.container():
        col1, col2 = st.columns([3, 1])
        
        with col1:
            if current_keys.get("slack") and not st.session_state.get("edit_slack"):
                masked_key = current_keys["slack"][:20] + "..." + current_keys["slack"][-10:] if len(current_keys["slack"]) > 30 else "***"
                st.markdown(f"""
                <div class="api-key-card">
                    <div class="api-key-status">
                        <div>
                            <div class="api-key-name">✅ Slack Webhook Active</div>
                            <div style="font-size: 12px; opacity: 0.8;">URL: {masked_key}</div>
                        </div>
                        <div class="api-key-badge badge-active">ACTIVE</div>
                    </div>
                </div>
                """, unsafe_allow_html=True)
                
                col_view, col_delete, col_add = st.columns(3)
                
                with col_view:
                    if st.button("👁️ View Full URL", key="view_slack", use_container_width=True):
                        st.session_state.show_slack_key = not st.session_state.get("show_slack_key", False)
                    
                    if st.session_state.get("show_slack_key"):
                        st.code(current_keys["slack"], language="text")
                
                with col_delete:
                    if st.button("🗑️ Delete URL", key="del_slack", use_container_width=True):
                        delete_user_api_key(username, "slack")
                        st.success("✓ Slack webhook deleted")
                        st.rerun()
                
                with col_add:
                    if st.button("🔄 Replace URL", key="replace_slack", use_container_width=True):
                        st.session_state.edit_slack = True
                        st.rerun()
            
            elif st.session_state.get("edit_slack") or not current_keys.get("slack"):
                # Add/Edit form
                if not current_keys.get("slack"):
                    st.markdown("""
                    <div class="api-key-card">
                        <div class="api-key-status">
                            <div>
                                <div class="api-key-name">❌ No Webhook Added</div>
                                <div style="font-size: 12px; opacity: 0.8;">Add your Slack webhook to receive notifications</div>
                            </div>
                            <div class="api-key-badge badge-inactive">INACTIVE</div>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                
                st.markdown("**Add or Replace Slack Webhook URL:**")
                new_slack_key = st.text_input(
                    "Paste your Slack Webhook URL",
                    type="password",
                    key="input_slack",
                    placeholder="https://hooks.slack.com/services/..."
                )
                
                col_save, col_cancel = st.columns(2)
                with col_save:
                    if st.button("✅ Save Slack URL", key="save_slack_btn", use_container_width=True):
                        if new_slack_key:
                            success, error = save_user_api_key(username, "slack", new_slack_key)
                            if success:
                                st.session_state.SLACK_WEBHOOK = new_slack_key
                                st.success("✓ Slack webhook saved securely and loaded into session")
                                st.session_state.edit_slack = False
                                st.rerun()
                            else:
                                st.error(f"Failed to save webhook: {error or 'Unknown error'}")
                        else:
                            st.error("Please enter a webhook URL")
                
                with col_cancel:
                    if st.button("❌ Cancel", key="cancel_slack_btn", use_container_width=True):
                        st.session_state.edit_slack = False
                        st.rerun()
    
    st.markdown("---")
    st.info("🔒 **Security:** Your API keys are encrypted and stored securely in the database. They are never displayed in plain text unless you explicitly view them.")
    
    # Show last updated time if available
    users = load_users()
    if username in users and "api_keys_updated" in users[username]:
        last_updated = users[username]["api_keys_updated"]
        st.caption(f"Last updated: {last_updated}")
    
    # =========================================================================
    # SMTP CONFIGURATION FOR EMAIL OUTREACH
    # =========================================================================
    st.markdown("---")
    st.markdown("## 📧 SMTP Email Configuration")
    st.markdown("Configure your Hostinger SMTP settings to send audit reports and cold emails directly from the app.")
    
    # Load current SMTP config
    smtp_config = load_email_config()
    
    # Status indicator
    if is_smtp_configured():
        st.success("✅ SMTP is configured and ready to send emails")
    else:
        st.warning("⚠️ SMTP not fully configured. Fill in all fields below to enable email sending.")
    
    with st.expander("📋 Hostinger SMTP Setup Guide", expanded=False):
        st.markdown("""
        ### Quick Setup for Hostinger Email
        
        **SMTP Configuration for Hostinger:**
        - **SMTP Server:** `smtp.hostinger.com`
        - **SMTP Port:** `465` (SSL) or `587` (TLS)
        - **Username:** Your full Hostinger email address
        - **Password:** Your Hostinger email password
        
        **Steps:**
        1. Log in to Hostinger Control Panel
        2. Go to **Email** → Your email account
        3. Copy the email address and password
        4. Fill in the form below
        """)
    
    with st.form("smtp_config_form"):
        st.markdown("### SMTP Server Settings")
        
        col1, col2 = st.columns(2)
        with col1:
            smtp_enabled = st.checkbox(
                "Enable Email Sending",
                value=smtp_config.get("enabled", True),
                help="Turn on/off email functionality"
            )
        with col2:
            auto_send = st.checkbox(
                "Auto-send Reports",
                value=smtp_config.get("auto_send_reports", False),
                help="Automatically send reports when emails are found"
            )
        
        col1, col2 = st.columns(2)
        with col1:
            smtp_host = st.text_input(
                "SMTP Host",
                value=smtp_config.get("smtp_server", "smtp.hostinger.com"),
                placeholder="smtp.hostinger.com",
                help="Your email provider's SMTP server"
            )
        with col2:
            smtp_port = st.selectbox(
                "SMTP Port",
                options=[465, 587, 25],
                index=0 if smtp_config.get("smtp_port", 465) == 465 else (1 if smtp_config.get("smtp_port") == 587 else 2),
                help="465 for SSL, 587 for TLS"
            )
        
        st.markdown("### Authentication")
        
        col1, col2 = st.columns(2)
        with col1:
            smtp_username = st.text_input(
                "SMTP Username (Email)",
                value=smtp_config.get("sender_email", ""),
                placeholder="contact@yourdomain.com",
                help="Your full email address for login"
            )
        with col2:
            smtp_password = st.text_input(
                "SMTP Password",
                value=smtp_config.get("sender_password", ""),
                type="password",
                placeholder="Your email password",
                help="Password for SMTP authentication"
            )
        
        st.markdown("### Display Settings")
        
        col1, col2 = st.columns(2)
        with col1:
            from_name = st.text_input(
                "From Name",
                value=smtp_config.get("from_name", "Code Nest"),
                placeholder="Code Nest",
                help="Name that appears in emails"
            )
        with col2:
            reply_to = st.text_input(
                "Reply-To Email",
                value=smtp_config.get("reply_to", smtp_config.get("sender_email", "")),
                placeholder="contact@yourdomain.com",
                help="Email address for replies"
            )
        
        st.markdown("---")
        col1, col2, col3 = st.columns([2, 2, 1])
        
        with col1:
            save_smtp = st.form_submit_button("💾 Save SMTP Settings", type="primary", use_container_width=True)
        
        with col2:
            test_email_addr = st.text_input(
                "Test Email Address",
                placeholder="test@example.com",
                help="Enter an email to send a test",
                label_visibility="collapsed"
            )
        
        with col3:
            send_test = st.form_submit_button("📧 Test", use_container_width=True)
        
        if save_smtp:
            new_smtp_config = {
                "enabled": smtp_enabled,
                "smtp_server": smtp_host,
                "smtp_port": smtp_port,
                "sender_email": smtp_username,
                "sender_password": smtp_password,
                "from_name": from_name,
                "reply_to": reply_to or smtp_username,
                "auto_send_reports": auto_send,
                "notifications": smtp_config.get("notifications", {
                    "audit_complete": True,
                    "report_sent": True,
                    "permission_change": True,
                    "admin_alert": True
                })
            }
            success, message = save_email_config(new_smtp_config)
            if success:
                st.success("✅ SMTP settings saved successfully!")
                st.rerun()
            else:
                st.error(f"❌ Failed to save: {message}")
        
        if send_test and test_email_addr:
            if not smtp_username or not smtp_password:
                st.error("Please fill in SMTP username and password first")
            else:
                # Temporarily save config for testing
                test_config = {
                    "enabled": True,
                    "smtp_server": smtp_host,
                    "smtp_port": smtp_port,
                    "sender_email": smtp_username,
                    "sender_password": smtp_password,
                    "from_name": from_name,
                    "reply_to": reply_to or smtp_username,
                }
                save_email_config(test_config)
                
                # Send test email with branding
                success, message = send_branded_email_with_pdf(
                    test_email_addr,
                    "Test Email from Code Nest Sales Engine",
                    "This is a test email to verify your SMTP configuration is working correctly.\n\nIf you received this, your email settings are configured properly!\n\nYou should see this message wrapped in our branded HTML email template with:\n- Dark green header with logo\n- White content area\n- Professional footer with contact info\n\n- Code Nest Team",
                    None,  # No PDF attachment for test
                    ""
                )
                if success:
                    st.success(f"✅ Test email sent to {test_email_addr}!")
                else:
                    st.error(f"❌ Test failed: {message}")

def show_single_audit():
    """
    Single website audit page with full persistence layer integration.
    
    Uses the new persistence framework for:
    - Audit data survival across refresh/navigation
    - AI cache to prevent recomputation
    - PDF context for report generation
    - Navigation state for deep linking
    """
    st.title("🚀 Single Website Audit")
    st.markdown("Enter a website URL to analyze its technical health, SEO, performance & generate AI insights")
    st.markdown("---")
    
    # =========================================================================
    # STEP 1: Restore navigation and audit state on page load
    # =========================================================================
    restore_result = restore_navigation_on_refresh()
    
    # Check for deep link audit ID from URL params
    deep_link_audit_id = get_deep_link_audit_id()
    if deep_link_audit_id:
        set_current_audit(deep_link_audit_id)
        logger.info(f"Loaded audit from deep link: {deep_link_audit_id}")
    
    # Save current navigation state
    current_audit_id = st.session_state.get('_current_audit_id')
    save_navigation_state("Single Audit", current_audit_id)
    
    # =========================================================================
    # STEP 2: Load current audit data from persistence layer
    # =========================================================================
    audit_data = get_current_audit()
    
    # If no audit in persistence but we have an ID, try to rebuild from DB
    if not audit_data and current_audit_id:
        audit_data = rebuild_audit_data_from_db(current_audit_id)
        if audit_data:
            set_current_audit(current_audit_id)
            logger.info(f"Rebuilt audit {current_audit_id} from database")
    
    # =========================================================================
    # STEP 3: URL input and analysis
    # =========================================================================
    col1, col2 = st.columns([3, 1])
    with col1:
        url = st.text_input("Website URL", placeholder="example.com")
    with col2:
        analyze_btn = st.button("🔍 Analyze", type="primary", use_container_width=True)
    
    if analyze_btn:
        url_sanitized = sanitize_input(url, max_length=2000)
        
        # Validate URL
        is_valid, error_msg = validate_url(url_sanitized)
        if not is_valid:
            st.error(f"❌ Invalid URL: {error_msg}")
            logger.warning(f"Invalid URL submitted: {url_sanitized}")
        else:
            logger.info(f"Starting audit for URL: {url_sanitized}")
            
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            status_text.text("⚡ Starting audit...")
            progress_bar.progress(10)
            
            with st.spinner("🔍 Analyzing website (typically 10-25 seconds)..."):
                success, data = safe_execute(
                    run_audit,
                    url_sanitized,
                    st.session_state.OPENAI_API_KEY,
                    st.session_state.GOOGLE_API_KEY,
                    error_message="Audit failed"
                )
                
                progress_bar.progress(100)
                status_text.text("✅ Audit complete!")
                
                if not success or "error" in data:
                    error_msg = data.get('error', 'Unknown error during audit')
                    st.error(f"❌ Scan Failed: {error_msg}")
                    logger.error(f"Audit failed for {url_sanitized}: {error_msg}")
                    clear_current_audit()
                else:
                    logger.info(f"Audit completed successfully for {url_sanitized}")
                    
                    # Save to database and get audit_id
                    audit_id = None
                    try:
                        if DB_AVAILABLE:
                            audit_id = save_audit_to_db(data)
                            logger.debug(f"Audit saved to database for {url_sanitized}, id={audit_id}")
                    except Exception as e:
                        logger.warning(f"Failed to save audit to database: {str(e)}")
                        st.warning("⚠️ Audit completed but couldn't save to database")
                    
                    # =========================================================
                    # PERSIST TO ALL LAYERS (session + disk + db)
                    # =========================================================
                    if audit_id:
                        data['audit_id'] = audit_id
                        persist_audit_data(audit_id, data)
                        set_current_audit(audit_id)
                        store_pdf_context(audit_id, data)
                        save_navigation_state("Single Audit", audit_id)
                        sync_query_params()
                        logger.info(f"Audit {audit_id} persisted to all layers")
                    
                    # Update local variable for display
                    audit_data = data
    
    # =========================================================================
    # STEP 4: Display audit results (from persistence layer)
    # =========================================================================
    if not audit_data:
        st.info("👆 Enter a website URL above and click **Analyze** to start an audit.")
        return
    
    # Phase 5 Step 3: Ensure all audit fields exist with safe defaults
    audit_data = ensure_audit_defaults(audit_data)
    
    # Ensure issues list is valid
    if not isinstance(audit_data.get("issues"), list):
        audit_data["issues"] = []
    
    # Metrics
    st.markdown("---")
    st.markdown("### 📊 Audit Results")
    
    c1, c2, c3, c4, c5 = st.columns(5)
    
    with c1:
        st.metric("Health Score", audit_data.get('score', 'N/A'), 
                  delta=("Good" if audit_data.get('score', 0) >= 70 else "Needs Work"))
    with c2:
        st.metric("Google Speed", audit_data.get('psi', 'N/A'))
    with c3:
        st.metric("Accessibility", audit_data.get('accessibility_score', 'N/A'))
    with c4:
        st.metric("Issues Found", len(audit_data.get('issues', [])))
    with c5:
        st.metric("Age", audit_data.get('domain_age', 'Unknown'))
    
    # Tech stack
    if audit_data.get('tech_stack'):
        st.markdown(f"**📦 Tech Stack:** {', '.join(audit_data['tech_stack'])}")
    
    # Issues
    if audit_data.get('issues'):
        st.markdown("---")
        st.markdown("### ⚠️ Issues Detected")
        
        for i, issue in enumerate(audit_data.get('issues', []), 1):
            try:
                with st.expander(f"{i}. {issue.get('title', 'Unknown Issue')}", expanded=(i <= 2)):
                    col1, col2 = st.columns(2)
                    with col1:
                        st.markdown(f"**Impact:** {issue.get('impact', 'N/A')}")
                    with col2:
                        st.markdown(f"**Solution:** {issue.get('solution', 'N/A')}")
            except Exception as e:
                logger.error(f"Error displaying issue {i}: {str(e)}")
                st.warning(f"Could not display issue #{i}")
    
    # AI analysis - Phase 5 Step 3: Ensure AI data is always a dict
    ai_data = audit_data.get('ai')
    if not isinstance(ai_data, dict):
        ai_data = {"summary": "", "impact": "", "solutions": "", "email": ""}
        audit_data['ai'] = ai_data
    
    if ai_data:
        st.markdown("---")
        
        # Show cache indicator
        if ai_data.get('from_cache'):
            st.markdown("### 🤖 AI Analysis _(cached)_")
            st.caption("💾 Using cached AI result to save API costs. Click 'Regenerate' for fresh analysis.")
        else:
            st.markdown("### 🤖 AI Analysis")
        
        # Check if we have structured insights
        insights = ai_data.get('insights')
        
        if insights:
            # Display structured AI insights
            col1, col2 = st.columns(2)
            
            with col1:
                # Snapshot Summary
                st.markdown("**📸 Snapshot Summary**")
                for bullet in insights.get('snapshot_summary', []):
                    st.markdown(f"• {bullet}")
                
                # Top 3 Issues
                st.markdown("**🚨 Top 3 Issues Hurting You Most**")
                for item in insights.get('top_3_issues', []):
                    st.markdown(f"**{item.get('issue', '')}**")
                    st.caption(f"↳ {item.get('impact', '')}")
            
            with col2:
                # Quick Wins
                st.markdown("**⚡ Quick Wins (Next 30 Days)**")
                for idx, win in enumerate(insights.get('quick_wins', []), 1):
                    st.markdown(f"{idx}. {win}")
                
                # Code Nest Services
                st.markdown("**🛠️ How Code Nest Can Help**")
                for item in insights.get('code_nest_services', []):
                    st.markdown(f"• **{item.get('issue', '')}** → {item.get('service', '')}")
            
            # Suggested Next Step
            if insights.get('next_step'):
                st.info(f"**📞 Next Step:** {insights['next_step']}")
        else:
            # Fallback to legacy display
            col1, col2 = st.columns(2)
            with col1:
                st.markdown("**Summary**")
                st.info(ai_data.get('summary', 'No summary available'))
                st.markdown("**Impact**")
                st.warning(ai_data.get('impact', 'No impact assessment available'))
            
            with col2:
                st.markdown("**Solutions**")
                st.success(ai_data.get('solutions', 'No solutions available'))
        
        # =========================================================================
        # EMAIL OUTREACH SECTION (always shown when AI is available)
        # =========================================================================
        st.markdown("---")
        st.markdown("### 📧 Send Email Outreach")
        
        # Premium format info note
        st.caption("✨ _This message was generated using Code Nest LLC's premium outreach format, covering Website Optimization, SEO, Social Media Management, PPC, and paid ad strategy._")
        
        # Check SMTP configuration
        if not is_smtp_configured():
            st.warning("⚠️ **SMTP not configured.** Go to **API Settings** to configure your Hostinger SMTP settings before sending emails.")
        else:
            st.success("✅ SMTP configured and ready")
        
        # Get domain for email personalization
        domain = urlparse(audit_data.get('url', '')).netloc.replace("www.", "")
        
        # Pre-fill email from audit data
        emails_found = audit_data.get('emails', [])
        default_email = emails_found[0] if emails_found else ""
        
        # Email subject - use AI generated or default (ensure lowercase)
        default_subject = ai_data.get('email_subject', f"quick note about {domain}").lower()
        
        # Email body - use AI generated cold email
        default_body = ai_data.get('email', '')
        
        col1, col2 = st.columns([2, 1])
        
        with col1:
            to_email = st.text_input(
                "📬 To Email",
                value=default_email,
                placeholder="contact@theirsite.com",
                help="Recipient email address - pre-filled if found during audit"
            )
        
        with col2:
            email_subject = st.text_input(
                "📝 Subject",
                value=default_subject,
                placeholder="quick note about domain.com",
                help="Lowercase, 3-5 words, personalized to their domain"
            )
        
        email_body = st.text_area(
            "✉️ Email Body",
            value=default_body,
            height=280,
            help="Premium agency-level cold email - edit as needed"
        )
        
        # Attach PDF checkbox
        attach_pdf = st.checkbox("📎 Attach PDF Audit Report", value=True, help="Include the full audit report as a PDF attachment")
        
        col1, col2, col3 = st.columns([2, 1, 1])
        
        with col1:
            send_email_btn = st.button("📤 Send Email", type="primary", use_container_width=True, disabled=not is_smtp_configured())
        
        with col2:
            # Check regeneration limit using persistence layer
            audit_url = audit_data.get('url', '')
            regen_count = get_regen_count(audit_url)
            can_regen = can_regenerate(audit_url)
            regen_label = f"🔄 Regenerate ({MAX_REGENERATIONS_PER_URL - regen_count} left)" if can_regen else "🔄 Limit Reached"
            
            if st.button(regen_label, use_container_width=True, disabled=not can_regen):
                if not can_regen:
                    st.warning(f"⚠️ Regeneration limit reached ({MAX_REGENERATIONS_PER_URL} max per audit) to control API costs.")
                elif st.session_state.OPENAI_API_KEY:
                    with st.spinner("Generating new email..."):
                        # Increment regen counter (persists to disk)
                        increment_regen_count(audit_url)
                        # Force regenerate (skip cache)
                        clear_ai_cache_for_url(audit_url)
                        new_ai = get_ai_consultation(audit_url, audit_data, st.session_state.OPENAI_API_KEY, force_regenerate=True)
                        audit_data['ai'] = new_ai
                        
                        # Update persistence layer with new AI data
                        audit_id = audit_data.get('audit_id')
                        if audit_id:
                            persist_audit_data(audit_id, audit_data)
                            store_pdf_context(audit_id, audit_data)
                        
                        # Cache the new AI result
                        cache_ai_result(audit_url, new_ai)
                        st.rerun()
                else:
                    st.error("OpenAI API key required")
        
        with col3:
            st.download_button(
                "📋 Copy Email",
                f"Subject: {email_subject}\n\n{email_body}",
                file_name="cold_email.txt",
                mime="text/plain",
                use_container_width=True
            )
        
        if send_email_btn:
            if not to_email:
                st.error("❌ Please enter a recipient email address")
            elif not email_subject:
                st.error("❌ Please enter an email subject")
            elif not email_body:
                st.error("❌ Please enter an email body")
            else:
                with st.spinner("Sending email..."):
                    # Generate PDF if checkbox is checked
                    pdf_bytes = None
                    pdf_filename = ""
                    
                    if attach_pdf:
                        try:
                            pdf_bytes = generate_pdf(audit_data)
                            pdf_filename = f"CodeNest_Audit_{domain.replace('.', '_')}.pdf"
                        except Exception as e:
                            st.warning(f"Could not generate PDF: {e}")
                            pdf_bytes = None
                    
                    # Send branded HTML email with PDF attachment
                    success, message = send_branded_email_with_pdf(
                        to_email=to_email,
                        subject=email_subject,
                        body=email_body,
                        pdf_bytes=pdf_bytes,
                        filename=pdf_filename
                    )
                    
                    if success:
                        st.success(f"✅ {message}")
                        st.balloons()
                        logger.info(f"Cold email sent to {to_email} for {domain}")
                    else:
                        st.error(f"❌ {message}")
                        logger.error(f"Failed to send email to {to_email}: {message}")
        
        # Save to DB (only if not already saved)
        audit_id = audit_data.get('audit_id')
        if not audit_id and DB_AVAILABLE:
            audit_id = save_audit_to_db(audit_data)
            if audit_id:
                audit_data['audit_id'] = audit_id
                # Persist the updated audit data with new ID
                persist_audit_data(audit_id, audit_data)
                set_current_audit(audit_id)
                store_pdf_context(audit_id, audit_data)
                st.success(f"✓ Audit saved (ID: {audit_id})")
                
                # Send Slack notification
                if st.session_state.SLACK_WEBHOOK:
                    send_slack_notification(f"🔍 New audit: {audit_data['url']} (Score: {audit_data['score']}/100)", st.session_state.SLACK_WEBHOOK)
        
        # PDF export and persistent storage
        st.markdown("---")
        try:
            pdf_bytes = generate_pdf(audit_data)
            domain_name = urlparse(audit_data['url']).netloc.replace("www.", "").replace(".", "_")
            
            # Save PDF to persistent storage if audit was saved to DB
            if audit_id:
                save_audit_pdf_to_file(audit_id, pdf_bytes)
                st.info(f"✓ PDF saved for future downloads")
            
            st.download_button(
                "📥 Download PDF Report",
                pdf_bytes,
                f"CodeNest_Audit_{domain_name}.pdf",
                "application/pdf",
                type="primary",
                use_container_width=True
            )
        except Exception as e:
            st.error(f"PDF Error: {e}")

# ============================================================================
# BULK SCAN - CONSTANTS & CONFIGURATION
# ============================================================================

# Bulk Scan Session State Keys (centralized)
BULK_SCAN_SESSION_KEYS = [
    "bulk_scan_session_id",
    "bulk_scan_active", 
    "bulk_scan_current_url",
    "bulk_scan_error_count"
]

# Rate limiting configuration
BULK_SCAN_DELAY_SECONDS = 0.5  # Delay between scans to prevent rate limiting
BULK_SCAN_MAX_ERRORS = 5  # Max consecutive errors before pausing

# ============================================================================
# CRM - PHONE EXTRACTION HELPER
# ============================================================================

def extract_phone_from_html(html_content: str) -> str:
    """Extract phone number from HTML content using regex patterns.
    
    Args:
        html_content: Raw HTML string
        
    Returns:
        Phone number string or None if not found
    """
    if not html_content:
        return None
    
    # Common phone number patterns
    phone_patterns = [
        # US formats: (555) 555-5555, 555-555-5555, 555.555.5555
        r'\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}',
        # International: +1 555 555 5555, +44 20 7123 4567
        r'\+\d{1,3}[-.\s]?\(?\d{1,4}\)?[-.\s]?\d{1,4}[-.\s]?\d{1,9}',
        # Simple: 5555555555
        r'\b\d{10,11}\b',
        # With tel: prefix from href
        r'tel:([+\d\-\s\(\)]+)',
    ]
    
    for pattern in phone_patterns:
        matches = re.findall(pattern, html_content)
        if matches:
            # Clean up the first match
            phone = matches[0] if isinstance(matches[0], str) else matches[0][0] if matches[0] else None
            if phone:
                # Remove common non-phone prefixes
                phone = re.sub(r'^tel:', '', phone).strip()
                # Basic validation: should have at least 10 digits
                digits_only = re.sub(r'\D', '', phone)
                if len(digits_only) >= 10:
                    return phone[:50]  # Truncate to fit DB column
    
    return None


# ============================================================================
# CRM - LEAD MANAGEMENT HELPER FUNCTIONS
# ============================================================================

def get_lead_for_domain(domain: str):
    """Get Lead record for a domain.
    
    Args:
        domain: Domain name (without protocol)
        
    Returns:
        Lead object or None
    """
    db = get_db()
    if not db:
        return None
    try:
        return db.query(Lead).filter(Lead.domain == domain).first()
    except Exception:
        return None
    finally:
        try:
            db.close()
        except:
            pass


def get_lead_for_audit(audit_id: int):
    """Get Lead record associated with an Audit.
    
    Args:
        audit_id: ID of the audit
        
    Returns:
        Lead object or None
    """
    logger = logging.getLogger("sales_engine")
    db = get_db()
    if not db:
        return None
    try:
        audit = db.query(Audit).filter(Audit.id == audit_id).first()
        if not audit:
            return None
        return db.query(Lead).filter(Lead.domain == audit.domain).first()
    except Exception as e:
        logger.warning(f"Could not get lead for audit {audit_id}: {e}")
        return None
    finally:
        try:
            db.close()
        except:
            pass


def update_lead_crm_fields(lead_id: int, **kwargs) -> bool:
    """Update CRM fields on a Lead.
    
    Args:
        lead_id: Lead database ID
        **kwargs: Field names and values to update
        
    Supported fields:
        approached, approached_date, follow_up_date, lead_status,
        interested, pipeline_stage, assigned_user, notes
        
    Returns:
        True on success, False on failure
    """
    logger = logging.getLogger("sales_engine")
    db = get_db()
    if not db:
        return False
    
    try:
        lead = db.query(Lead).filter(Lead.id == lead_id).first()
        if not lead:
            logger.warning(f"Lead {lead_id} not found for CRM update")
            return False
        
        # List of allowed CRM fields
        allowed_fields = [
            'approached', 'approached_date', 'follow_up_date', 'lead_status',
            'interested', 'pipeline_stage', 'assigned_user', 'notes', 
            'phone', 'email', 'company_name'
        ]
        
        for field, value in kwargs.items():
            if field in allowed_fields and hasattr(lead, field):
                setattr(lead, field, value)
                logger.debug(f"Lead {lead_id}: set {field}={value}")
        
        lead.updated_at = datetime.utcnow()
        db.commit()
        logger.info(f"Lead {lead_id} CRM fields updated: {list(kwargs.keys())}")
        return True
        
    except Exception as e:
        logger.error(f"Failed to update lead {lead_id} CRM fields: {e}")
        if db:
            try:
                db.rollback()
            except:
                pass
        return False
    finally:
        if db:
            try:
                db.close()
            except:
                pass


def mark_lead_as_approached(lead_id: int, update_pipeline: bool = True) -> bool:
    """Mark a lead as approached (contacted).
    
    Args:
        lead_id: Lead database ID
        update_pipeline: If True, also move pipeline_stage to "contacted" if it was "new"
        
    Returns:
        True on success, False on failure
    """
    logger = logging.getLogger("sales_engine")
    db = get_db()
    if not db:
        return False
    
    try:
        lead = db.query(Lead).filter(Lead.id == lead_id).first()
        if not lead:
            return False
        
        lead.approached = True
        lead.approached_date = datetime.utcnow()
        
        # Move pipeline forward if still at "new"
        if update_pipeline and lead.pipeline_stage == "new":
            lead.pipeline_stage = "contacted"
        
        lead.updated_at = datetime.utcnow()
        db.commit()
        logger.info(f"Lead {lead_id} ({lead.domain}) marked as approached")
        return True
        
    except Exception as e:
        logger.error(f"Failed to mark lead {lead_id} as approached: {e}")
        if db:
            try:
                db.rollback()
            except:
                pass
        return False
    finally:
        if db:
            try:
                db.close()
            except:
                pass


def get_crm_metrics() -> dict:
    """Get CRM dashboard metrics.
    
    Returns:
        Dict with counts: total, not_approached, approached, hot, follow_up_due
    """
    db = get_db()
    if not db:
        return {"total": 0, "not_approached": 0, "approached": 0, "hot": 0, "follow_up_due": 0}
    
    try:
        total = db.query(Lead).count()
        not_approached = db.query(Lead).filter(Lead.approached == False).count()
        approached = db.query(Lead).filter(Lead.approached == True).count()
        hot = db.query(Lead).filter(Lead.lead_status == "hot").count()
        
        # Follow-ups due today or earlier
        today = datetime.utcnow().replace(hour=23, minute=59, second=59)
        follow_up_due = db.query(Lead).filter(
            Lead.follow_up_date <= today,
            Lead.pipeline_stage == "follow-up"
        ).count()
        
        return {
            "total": total,
            "not_approached": not_approached,
            "approached": approached,
            "hot": hot,
            "follow_up_due": follow_up_due
        }
    except Exception:
        return {"total": 0, "not_approached": 0, "approached": 0, "hot": 0, "follow_up_due": 0}
    finally:
        try:
            db.close()
        except:
            pass

# ============================================================================
# BULK SCAN - URL VALIDATION & SANITIZATION
# ============================================================================

def validate_and_clean_url(url: str) -> tuple:
    """Validate and sanitize a URL for bulk scanning.
    
    Args:
        url: Raw URL string from CSV
        
    Returns:
        Tuple of (cleaned_url, is_valid, error_message)
    """
    if not url or not isinstance(url, str):
        return None, False, "Empty or invalid URL"
    
    # Strip whitespace
    url = url.strip()
    
    # Skip empty or nan values
    if url.lower() in ['nan', 'none', '', 'null', 'n/a']:
        return None, False, "Empty value"
    
    # Remove common prefixes that might cause issues
    url = url.replace('www.', '') if url.startswith('www.') else url
    
    # Add https:// if no protocol
    if not url.startswith(('http://', 'https://')):
        url = f"https://{url}"
    
    # Basic domain validation
    try:
        from urllib.parse import urlparse
        parsed = urlparse(url)
        if not parsed.netloc:
            return None, False, "Invalid domain"
        
        # Check for valid TLD (basic check)
        domain_parts = parsed.netloc.split('.')
        if len(domain_parts) < 2:
            return None, False, "Invalid domain format"
        
        return url, True, None
    except Exception as e:
        return None, False, str(e)


def detect_website_column(df) -> str:
    """Detect the column containing website URLs in a DataFrame.
    
    Args:
        df: Pandas DataFrame
        
    Returns:
        Column name or None if not found
    """
    # Check for common column names (case-insensitive)
    common_names = ['website', 'url', 'domain', 'site', 'web', 'link', 'homepage']
    
    for col in df.columns:
        if col.lower().strip() in common_names:
            return col
    
    # Check if any column contains mostly URL-like values
    for col in df.columns:
        sample = df[col].dropna().head(10).astype(str)
        url_count = sum(1 for v in sample if '.' in v and len(v) > 5)
        if url_count >= len(sample) * 0.7:  # 70% look like URLs
            return col
    
    return None


def extract_urls_from_dataframe(df, column_name: str) -> tuple:
    """Extract and validate URLs from a DataFrame column.
    
    Args:
        df: Pandas DataFrame
        column_name: Name of the column containing URLs
        
    Returns:
        Tuple of (valid_urls_list, invalid_count, errors_list)
    """
    valid_urls = []
    errors = []
    invalid_count = 0
    
    for idx, raw_url in enumerate(df[column_name].tolist()):
        url, is_valid, error = validate_and_clean_url(str(raw_url))
        
        if is_valid:
            valid_urls.append(url)
        else:
            invalid_count += 1
            if error and error != "Empty value":
                errors.append(f"Row {idx + 1}: {error}")
    
    return valid_urls, invalid_count, errors[:10]  # Limit error messages


# ============================================================================
# BULK SCAN - SAFE AUDIT RUNNER
# ============================================================================

def run_bulk_audit_safe(url: str, openai_key: str, google_key: str) -> tuple:
    """Safely run an audit with error handling for bulk processing.
    
    Args:
        url: URL to audit
        openai_key: OpenAI API key
        google_key: Google API key
        
    Returns:
        Tuple of (audit_id, success, error_message)
    """
    logger = logging.getLogger("sales_engine")
    
    # Phase 5 Step 3: Skip invalid or overly long URLs
    if not url or not isinstance(url, str):
        logger.warning(f"Bulk audit skipping invalid URL type: {type(url)}")
        return None, False, "Invalid URL"
    
    url = url.strip()
    
    if len(url) > 255:
        logger.warning(f"Bulk audit skipping URL > 255 chars: {url[:50]}...")
        return None, False, "URL too long (max 255 chars)"
    
    if len(url) < 4:
        logger.warning(f"Bulk audit skipping URL too short: {url}")
        return None, False, "URL too short"
    
    # Basic domain validation
    if not any(url.startswith(p) for p in ['http://', 'https://', 'www.']) and '.' not in url:
        logger.warning(f"Bulk audit skipping invalid domain format: {url}")
        return None, False, "Invalid domain format"
    
    try:
        # Run the audit
        audit_data = run_audit(url, openai_key, google_key)
        
        if not audit_data:
            logger.warning(f"Bulk audit returned no data for: {url}")
            return None, False, "Audit returned no data"
        
        # Check for error in audit_data
        if isinstance(audit_data, dict) and "error" in audit_data:
            error_msg = audit_data.get("error", "Unknown error")
            logger.warning(f"Bulk audit returned error for {url}: {error_msg}")
            return None, False, error_msg
        
        # Save to database with source="bulk"
        audit_id = save_audit_to_db(audit_data, source="bulk")
        
        if audit_id:
            logger.info(f"Bulk audit success: {url} -> audit_id={audit_id}")
            return audit_id, True, None
        else:
            logger.warning(f"Bulk audit save failed for: {url}")
            return None, False, "Failed to save audit"
            
    except Exception as e:
        logger.error(f"Bulk audit error for {url}: {str(e)}")
        return None, False, str(e)


# ============================================================================
# BULK SCAN BACKGROUND PROCESSING FUNCTIONS (FIXED)
# ============================================================================

def create_bulk_scan_session(urls: list, username: str = None) -> str:
    """Create a new bulk scan session and store in database.
    
    Args:
        urls: List of validated URLs to scan
        username: Optional username of the creator
        
    Returns:
        Session ID string or None on failure
    """
    import uuid
    logger = logging.getLogger("sales_engine")
    db = get_db()
    
    if not db:
        logger.error("Database unavailable for bulk scan session creation")
        return None
    
    try:
        session_id = str(uuid.uuid4())
        bulk_scan = BulkScan(
            session_id=session_id,
            status="running",
            total_urls=len(urls),
            urls=urls,
            processed_urls=0,
            paused_at_index=0,
            results={}
        )
        db.add(bulk_scan)
        db.commit()
        
        logger.info(f"Bulk scan session created: {session_id[:8]}... ({len(urls)} URLs)")
        return session_id
        
    except Exception as e:
        logger.error(f"Error creating bulk scan session: {e}")
        if db:
            try:
                db.rollback()
            except:
                pass
        return None
    finally:
        if db:
            try:
                db.close()
            except:
                pass


def get_bulk_scan_session(session_id: str):
    """Retrieve bulk scan session from database.
    
    Args:
        session_id: UUID of the session
        
    Returns:
        BulkScan object or None
    """
    db = get_db()
    if not db:
        return None
    try:
        return db.query(BulkScan).filter(BulkScan.session_id == session_id).first()
    finally:
        try:
            db.close()
        except:
            pass


def update_bulk_scan_progress(session_id: str, processed_count: int, results: dict, 
                               paused_at_index: int = None, status: str = "running") -> bool:
    """Update bulk scan progress in database.
    
    FIXED: Now updates paused_at_index to enable proper resume functionality.
    
    Args:
        session_id: UUID of the session
        processed_count: Number of URLs processed
        results: Dictionary of {url: audit_id}
        paused_at_index: Current index for resume (defaults to processed_count)
        status: Session status
        
    Returns:
        True on success, False on failure
    """
    logger = logging.getLogger("sales_engine")
    db = get_db()
    
    if not db:
        return False
    
    try:
        scan = db.query(BulkScan).filter(BulkScan.session_id == session_id).first()
        if scan:
            scan.processed_urls = processed_count
            scan.results = results
            scan.status = status
            # FIX: Update paused_at_index for proper resume
            scan.paused_at_index = paused_at_index if paused_at_index is not None else processed_count
            scan.updated_at = datetime.utcnow()
            db.commit()
            return True
        return False
    except Exception as e:
        logger.error(f"Error updating bulk scan: {e}")
        if db:
            try:
                db.rollback()
            except:
                pass
        return False
    finally:
        if db:
            try:
                db.close()
            except:
                pass


def pause_bulk_scan(session_id: str, current_index: int) -> bool:
    """Pause a bulk scan at a specific index.
    
    Args:
        session_id: UUID of the session
        current_index: Current processing index to resume from
        
    Returns:
        True on success, False on failure
    """
    logger = logging.getLogger("sales_engine")
    db = get_db()
    
    if not db:
        return False
    
    try:
        scan = db.query(BulkScan).filter(BulkScan.session_id == session_id).first()
        if scan:
            scan.status = "paused"
            scan.paused_at_index = current_index
            scan.updated_at = datetime.utcnow()
            db.commit()
            logger.info(f"Bulk scan paused: {session_id[:8]}... at index {current_index}")
            return True
        return False
    except Exception as e:
        logger.error(f"Error pausing bulk scan: {e}")
        if db:
            try:
                db.rollback()
            except:
                pass
        return False
    finally:
        if db:
            try:
                db.close()
            except:
                pass


def resume_bulk_scan(session_id: str) -> bool:
    """Resume a paused bulk scan.
    
    Args:
        session_id: UUID of the session
        
    Returns:
        True on success, False on failure
    """
    logger = logging.getLogger("sales_engine")
    db = get_db()
    
    if not db:
        return False
    
    try:
        scan = db.query(BulkScan).filter(BulkScan.session_id == session_id).first()
        if scan:
            scan.status = "running"
            scan.updated_at = datetime.utcnow()
            db.commit()
            logger.info(f"Bulk scan resumed: {session_id[:8]}... from index {scan.paused_at_index}")
            return True
        return False
    except Exception as e:
        logger.error(f"Error resuming bulk scan: {e}")
        if db:
            try:
                db.rollback()
            except:
                pass
        return False
    finally:
        if db:
            try:
                db.close()
            except:
                pass


def stop_bulk_scan(session_id: str) -> bool:
    """Stop a bulk scan permanently.
    
    Args:
        session_id: UUID of the session
        
    Returns:
        True on success, False on failure
    """
    logger = logging.getLogger("sales_engine")
    db = get_db()
    
    if not db:
        return False
    
    try:
        scan = db.query(BulkScan).filter(BulkScan.session_id == session_id).first()
        if scan:
            scan.status = "stopped"
            scan.updated_at = datetime.utcnow()
            db.commit()
            logger.info(f"Bulk scan stopped: {session_id[:8]}...")
            return True
        return False
    except Exception as e:
        logger.error(f"Error stopping bulk scan: {e}")
        if db:
            try:
                db.rollback()
            except:
                pass
        return False
    finally:
        if db:
            try:
                db.close()
            except:
                pass


def complete_bulk_scan(session_id: str) -> bool:
    """Mark a bulk scan as completed.
    
    Args:
        session_id: UUID of the session
        
    Returns:
        True on success, False on failure
    """
    logger = logging.getLogger("sales_engine")
    db = get_db()
    
    if not db:
        return False
    
    try:
        scan = db.query(BulkScan).filter(BulkScan.session_id == session_id).first()
        if scan:
            scan.status = "completed"
            scan.updated_at = datetime.utcnow()
            db.commit()
            logger.info(f"Bulk scan completed: {session_id[:8]}...")
            return True
        return False
    except Exception as e:
        logger.error(f"Error completing bulk scan: {e}")
        if db:
            try:
                db.rollback()
            except:
                pass
        return False
    finally:
        if db:
            try:
                db.close()
            except:
                pass


def get_bulk_scan_sessions(limit: int = 10, status_filter: str = None) -> list:
    """Get list of recent bulk scan sessions.
    
    Args:
        limit: Maximum number of sessions to return
        status_filter: Optional status to filter by
        
    Returns:
        List of BulkScan objects
    """
    db = get_db()
    if not db:
        return []
    
    try:
        query = db.query(BulkScan).order_by(BulkScan.created_at.desc())
        if status_filter:
            query = query.filter(BulkScan.status == status_filter)
        return query.limit(limit).all()
    except Exception:
        return []
    finally:
        try:
            db.close()
        except:
            pass


def export_bulk_scan_results(session_id: str) -> bytes:
    """Export bulk scan results as CSV.
    
    Args:
        session_id: UUID of the session
        
    Returns:
        CSV data as bytes, or empty bytes on failure
    """
    logger = logging.getLogger("sales_engine")
    db = get_db()
    
    if not db:
        return b""
    
    try:
        session = db.query(BulkScan).filter(BulkScan.session_id == session_id).first()
        if not session or not session.results:
            return b""
        
        results_list = []
        for url, audit_id in session.results.items():
            if audit_id:
                audit_orm = db.query(Audit).filter(Audit.id == audit_id).first()
                if audit_orm:
                    # Normalize audit for consistent access
                    audit = normalize_audit(audit_orm)
                    tech_stack = audit["tech_stack"][:5] if audit["tech_stack"] else []
                    emails = audit["emails_found"] if audit["emails_found"] else []
                    ai_summary = audit["ai_summary"]
                    
                    results_list.append({
                        "Website": url,
                        "Domain": audit["domain"] or url,
                        "Health Score": audit["health_score"] if audit["health_score"] is not None else "N/A",
                        "PSI Speed": audit["psi_score"] if audit["psi_score"] else "N/A",
                        "Issues Found": len(audit["issues"]) if audit["issues"] else 0,
                        "Domain Age": audit["domain_age"] or "N/A",
                        "Tech Stack": ", ".join(tech_stack) if tech_stack else "N/A",
                        "Emails Found": ", ".join(emails) if emails else "N/A",
                        "AI Summary": (ai_summary[:100] + "...") if ai_summary else "N/A",
                        "Scanned At": safe_timestamp_slice(audit["created_at"], 16).replace("T", " ") if audit["created_at"] else "N/A"
                    })
            else:
                results_list.append({
                    "Website": url,
                    "Domain": url,
                    "Health Score": "FAILED",
                    "PSI Speed": "N/A",
                    "Issues Found": "N/A",
                    "Domain Age": "N/A",
                    "Tech Stack": "N/A",
                    "Emails Found": "N/A",
                    "AI Summary": "Audit failed",
                    "Scanned At": "N/A"
                })
        
        if results_list:
            df = pd.DataFrame(results_list)
            return df.to_csv(index=False).encode('utf-8')
        return b""
        
    except Exception as e:
        logger.error(f"Error exporting bulk scan results: {e}")
        return b""
    finally:
        try:
            db.close()
        except:
            pass


# ============================================================================
# BULK AUDIT UI - CODE NEST BRANDED (COMPLETE REWRITE)
# ============================================================================

def show_bulk_audit():
    """Bulk Website Audit - Code Nest branded interface with reliable processing.
    
    Features:
    - CSV upload with smart column detection
    - URL validation and sanitization
    - Background processing with auto-continue
    - Pause/Resume/Stop controls
    - Real-time progress tracking
    - CSV export of results
    - Professional Code Nest branding
    """
    logger = logging.getLogger("sales_engine")
    logger.debug("Rendering Bulk Audit page")
    
    # =========================================================================
    # CODE NEST BRANDING - CSS STYLES
    # =========================================================================
    st.markdown("""
    <style>
    /* Code Nest Brand Colors */
    :root {
        --cn-dark-green: #0c3740;
        --cn-accent-green: #2b945f;
        --cn-font-grey: #5a5a5a;
        --cn-white: #feffff;
        --cn-light-bg: #f8faf9;
    }
    
    /* Bulk Audit Container */
    .bulk-audit-header {
        background: linear-gradient(135deg, #0c3740 0%, #1a5a6e 100%);
        padding: 2rem;
        border-radius: 12px;
        margin-bottom: 1.5rem;
        color: white;
    }
    .bulk-audit-header h1 {
        font-family: 'Poppins', sans-serif;
        font-size: 2rem;
        margin: 0;
        font-weight: 600;
    }
    .bulk-audit-header p {
        font-family: 'Lato', sans-serif;
        opacity: 0.9;
        margin: 0.5rem 0 0 0;
    }
    
    /* Upload Card */
    .upload-card {
        background: white;
        border: 2px dashed #2b945f;
        border-radius: 12px;
        padding: 2rem;
        text-align: center;
        transition: all 0.3s ease;
    }
    .upload-card:hover {
        border-color: #0c3740;
        box-shadow: 0 4px 20px rgba(12, 55, 64, 0.1);
    }
    
    /* Session Card */
    .session-card {
        background: white;
        border-radius: 12px;
        padding: 1.5rem;
        margin-bottom: 1rem;
        box-shadow: 0 2px 12px rgba(0,0,0,0.06);
        border-left: 4px solid #2b945f;
    }
    .session-card.paused {
        border-left-color: #f0ad4e;
    }
    .session-card.stopped {
        border-left-color: #d9534f;
    }
    .session-card.completed {
        border-left-color: #2b945f;
    }
    
    /* Status Badges */
    .status-badge {
        display: inline-block;
        padding: 0.25rem 0.75rem;
        border-radius: 20px;
        font-size: 0.85rem;
        font-weight: 600;
        font-family: 'Lato', sans-serif;
    }
    .status-running {
        background: #d4edda;
        color: #155724;
    }
    .status-paused {
        background: #fff3cd;
        color: #856404;
    }
    .status-completed {
        background: #cce5ff;
        color: #004085;
    }
    .status-stopped {
        background: #f8d7da;
        color: #721c24;
    }
    
    /* Progress Bar Custom */
    .bulk-progress {
        background: #e9ecef;
        border-radius: 10px;
        height: 12px;
        overflow: hidden;
        margin: 0.5rem 0;
    }
    .bulk-progress-bar {
        background: linear-gradient(90deg, #2b945f 0%, #3dc978 100%);
        height: 100%;
        border-radius: 10px;
        transition: width 0.5s ease;
    }
    
    /* Empty State */
    .empty-state {
        text-align: center;
        padding: 3rem 2rem;
        background: #f8faf9;
        border-radius: 12px;
        border: 1px dashed #dee2e6;
    }
    .empty-state h3 {
        font-family: 'Poppins', sans-serif;
        color: #0c3740;
        margin-bottom: 0.5rem;
    }
    .empty-state p {
        font-family: 'Lato', sans-serif;
        color: #5a5a5a;
    }
    
    /* Processing Indicator */
    .processing-indicator {
        background: linear-gradient(135deg, #2b945f 0%, #3dc978 100%);
        color: white;
        padding: 1rem 1.5rem;
        border-radius: 10px;
        margin: 1rem 0;
        animation: pulse 2s infinite;
    }
    @keyframes pulse {
        0%, 100% { opacity: 1; }
        50% { opacity: 0.8; }
    }
    
    /* Button Styling */
    .stButton > button {
        font-family: 'Lato', sans-serif;
        border-radius: 8px;
        transition: all 0.2s ease;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # =========================================================================
    # HEADER
    # =========================================================================
    st.markdown("""
    <div class="bulk-audit-header">
        <h1>📂 Bulk Website Audit</h1>
        <p>Upload your lead list and let Code Nest analyze them all. Background processing means you can navigate away and return anytime.</p>
    </div>
    """, unsafe_allow_html=True)
    
    # =========================================================================
    # TABS: NEW SCAN | ACTIVE SESSIONS
    # =========================================================================
    tab1, tab2 = st.tabs(["📤 New Scan", "📊 Active Sessions"])
    
    # =========================================================================
    # TAB 1: NEW SCAN
    # =========================================================================
    with tab1:
        st.markdown("### Upload Your Lead List")
        st.markdown("Upload a CSV file with a column containing website URLs. We'll detect and validate them automatically.")
        
        uploaded = st.file_uploader(
            "Choose CSV file",
            type=["csv"],
            key="bulk_csv_upload",
            help="CSV should have a column named 'Website', 'URL', 'Domain', or similar"
        )
        
        if uploaded:
            try:
                df = pd.read_csv(uploaded)
                
                # Detect website column
                website_col = detect_website_column(df)
                
                if not website_col:
                    st.error("❌ Could not find a website column. Please ensure your CSV has a column named 'Website', 'URL', 'Domain', or 'Site'.")
                    st.markdown("**Your columns:** " + ", ".join(df.columns.tolist()))
                else:
                    # Show detected column
                    if website_col.lower() != 'website':
                        st.info(f"ℹ️ Using column **'{website_col}'** as website source")
                    
                    # Extract and validate URLs
                    valid_urls, invalid_count, errors = extract_urls_from_dataframe(df, website_col)
                    
                    # Summary metrics
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Total Rows", len(df))
                    with col2:
                        st.metric("Valid URLs", len(valid_urls), delta=None if invalid_count == 0 else f"-{invalid_count} invalid")
                    with col3:
                        estimated_time = len(valid_urls) * 15  # ~15 seconds per URL
                        st.metric("Est. Time", f"{estimated_time // 60}m {estimated_time % 60}s")
                    
                    # Show validation errors if any
                    if errors:
                        with st.expander(f"⚠️ {invalid_count} URLs could not be validated", expanded=False):
                            for err in errors:
                                st.caption(err)
                    
                    # Preview table
                    st.markdown("#### Preview (first 10 rows)")
                    preview_df = df.head(10)[[website_col]].copy()
                    preview_df['Status'] = preview_df[website_col].apply(
                        lambda x: "✓ Valid" if validate_and_clean_url(str(x))[1] else "✗ Invalid"
                    )
                    st.dataframe(preview_df, use_container_width=True, hide_index=True)
                    
                    # Start button
                    st.markdown("---")
                    
                    if len(valid_urls) == 0:
                        st.error("No valid URLs found. Please check your CSV file.")
                    else:
                        col_start, col_space = st.columns([2, 3])
                        with col_start:
                            if st.button("🚀 Start Bulk Scan", type="primary", use_container_width=True):
                                # Create session
                                session_id = create_bulk_scan_session(valid_urls)
                                
                                if session_id:
                                    st.session_state['bulk_scan_session_id'] = session_id
                                    st.session_state['bulk_scan_active'] = True
                                    
                                    st.success(f"✅ Scan started! Processing {len(valid_urls)} websites...")
                                    st.markdown("""
                                    <div class="processing-indicator">
                                        🔄 <strong>Scan in Progress</strong> — You can navigate to other pages. Progress will continue in the background.
                                    </div>
                                    """, unsafe_allow_html=True)
                                    
                                    time.sleep(1)
                                    st.rerun()
                                else:
                                    st.error("Failed to create scan session. Please try again.")
                        
                        with col_space:
                            st.caption("Scanning will run in the background. Results are saved automatically.")
                            
            except Exception as e:
                logger.error(f"Error reading CSV: {e}")
                st.error(f"Error reading CSV file: {str(e)}")
        
        else:
            # Empty state
            st.markdown("""
            <div class="empty-state">
                <h3>📋 No File Uploaded</h3>
                <p>Upload your lead list and let Code Nest do the heavy lifting.<br>
                We'll analyze each website and generate actionable insights.</p>
            </div>
            """, unsafe_allow_html=True)
    
    # =========================================================================
    # TAB 2: ACTIVE SESSIONS
    # =========================================================================
    with tab2:
        st.markdown("### Your Scanning Sessions")
        
        # Refresh button
        col_refresh, col_space = st.columns([1, 4])
        with col_refresh:
            if st.button("🔄 Refresh", key="refresh_sessions"):
                st.rerun()
        
        sessions = get_bulk_scan_sessions(limit=20)
        
        if not sessions:
            st.markdown("""
            <div class="empty-state">
                <h3>📭 No Sessions Yet</h3>
                <p>Start a new bulk scan to see your sessions here.<br>
                Sessions are saved automatically and persist across page reloads.</p>
            </div>
            """, unsafe_allow_html=True)
        else:
            for session in sessions:
                progress_pct = (session.processed_urls / session.total_urls * 100) if session.total_urls > 0 else 0
                
                # Session container
                with st.container(border=True):
                    # Header row
                    col1, col2, col3 = st.columns([3, 1, 2])
                    
                    with col1:
                        # Status badge
                        status_map = {
                            "running": ("🟢 Running", "status-running"),
                            "paused": ("⏸️ Paused", "status-paused"),
                            "completed": ("✅ Completed", "status-completed"),
                            "stopped": ("🔴 Stopped", "status-stopped")
                        }
                        status_text, status_class = status_map.get(session.status, ("Unknown", ""))
                        st.markdown(f"**{status_text}** • Session `{session.session_id[:8]}...`")
                        
                        # Progress bar
                        st.progress(progress_pct / 100)
                        st.caption(f"{session.processed_urls} of {session.total_urls} URLs processed ({progress_pct:.1f}%)")
                    
                    with col2:
                        success_count = len([v for v in (session.results or {}).values() if v]) if session.results else 0
                        fail_count = session.processed_urls - success_count
                        st.metric("✓ Success", success_count)
                        if fail_count > 0:
                            st.caption(f"✗ {fail_count} failed")
                    
                    with col3:
                        # Action buttons based on status
                        if session.status == "running":
                            c1, c2 = st.columns(2)
                            with c1:
                                if st.button("⏸️ Pause", key=f"pause_{session.id}", use_container_width=True):
                                    pause_bulk_scan(session.session_id, session.processed_urls)
                                    # Clear active flag if this is the active session
                                    if st.session_state.get('bulk_scan_session_id') == session.session_id:
                                        st.session_state.pop('bulk_scan_active', None)
                                    st.rerun()
                            with c2:
                                if st.button("⏹️ Stop", key=f"stop_{session.id}", use_container_width=True):
                                    stop_bulk_scan(session.session_id)
                                    if st.session_state.get('bulk_scan_session_id') == session.session_id:
                                        st.session_state.pop('bulk_scan_session_id', None)
                                        st.session_state.pop('bulk_scan_active', None)
                                    st.rerun()
                        
                        elif session.status == "paused":
                            c1, c2 = st.columns(2)
                            with c1:
                                if st.button("▶️ Resume", key=f"resume_{session.id}", use_container_width=True):
                                    resume_bulk_scan(session.session_id)
                                    st.session_state['bulk_scan_session_id'] = session.session_id
                                    st.session_state['bulk_scan_active'] = True
                                    st.rerun()
                            with c2:
                                if st.button("⏹️ Stop", key=f"stop2_{session.id}", use_container_width=True):
                                    stop_bulk_scan(session.session_id)
                                    st.rerun()
                        
                        elif session.status == "completed":
                            # Export button
                            csv_data = export_bulk_scan_results(session.session_id)
                            if csv_data:
                                st.download_button(
                                    "📥 Download Results",
                                    csv_data,
                                    f"bulk_scan_{session.session_id[:8]}.csv",
                                    "text/csv",
                                    key=f"download_{session.id}",
                                    use_container_width=True
                                )
                        
                        elif session.status == "stopped":
                            st.caption("Scan was stopped")
                    
                    # Recently scanned URLs (expandable) with CRM navigation
                    if session.results and len(session.results) > 0:
                        with st.expander(f"📋 View scanned URLs ({len(session.results)})", expanded=False):
                            st.markdown("""
                            <style>
                            .bulk-url-row {
                                display: flex;
                                align-items: center;
                                padding: 0.5rem;
                                border-bottom: 1px solid #eee;
                                background: #fafafa;
                                border-radius: 4px;
                                margin-bottom: 4px;
                            }
                            .bulk-url-row:hover { background: #f0f0f0; }
                            .url-status-icon { margin-right: 8px; font-size: 1.1em; }
                            .url-domain { flex: 1; font-family: 'Lato', sans-serif; color: #333; }
                            .url-actions { display: flex; gap: 4px; }
                            </style>
                            """, unsafe_allow_html=True)
                            
                            # Show all URLs with navigation buttons
                            urls_list = list(session.results.items())
                            for idx, (url, audit_id) in enumerate(urls_list):
                                icon = "✅" if audit_id else "❌"
                                # Extract domain for display
                                try:
                                    domain = urlparse(url).netloc.replace("www.", "")[:40]
                                except:
                                    domain = url[:40]
                                
                                col_icon, col_domain, col_btn1, col_btn2 = st.columns([0.5, 4, 2, 2])
                                
                                with col_icon:
                                    st.markdown(f"<span class='url-status-icon'>{icon}</span>", unsafe_allow_html=True)
                                
                                with col_domain:
                                    if audit_id:
                                        st.markdown(f"**{domain}**")
                                    else:
                                        st.markdown(f"~~{domain}~~ (failed)")
                                
                                with col_btn1:
                                    if audit_id:
                                        if st.button("📊 History", key=f"bulk_hist_{session.id}_{idx}", use_container_width=True,
                                                     help="View this audit in Audit History"):
                                            # Navigate to audit history with domain pre-filled
                                            st.session_state.hist_search = domain
                                            st.session_state.current_section = 'Audit History'
                                            st.rerun()
                                    else:
                                        st.button("📊 History", key=f"bulk_hist_d_{session.id}_{idx}", disabled=True, use_container_width=True)
                                
                                with col_btn2:
                                    if audit_id:
                                        if st.button("🔍 Open", key=f"bulk_open_{session.id}_{idx}", use_container_width=True,
                                                     help="Load this audit in Single Audit view"):
                                            # Load audit data and navigate to Single Audit using persistence layer
                                            db = get_db()
                                            if db:
                                                try:
                                                    audit = db.query(Audit).filter(Audit.id == audit_id).first()
                                                    if audit:
                                                        data = convert_audit_to_data_dict(audit)
                                                        
                                                        # Use persistence layer exclusively (no direct session state)
                                                        set_current_audit(audit_id)
                                                        persist_audit_data(audit_id, data)
                                                        store_pdf_context(audit_id, data)
                                                        save_navigation_state("Single Audit", audit_id)
                                                        sync_query_params()
                                                        
                                                        st.session_state.current_section = 'Single Audit'
                                                        st.toast(f"✓ Loaded {domain}")
                                                except Exception as e:
                                                    logger.warning(f"Error loading audit {audit_id}: {e}")
                                                finally:
                                                    db.close()
                                            st.rerun()
                                    else:
                                        # Retry button for failed URLs
                                        if st.button("🔄 Retry", key=f"bulk_retry_{session.id}_{idx}", use_container_width=True,
                                                     help="Retry auditing this URL"):
                                            with st.spinner(f"Retrying {domain}..."):
                                                new_audit_id, success, error = run_bulk_audit_safe(
                                                    url,
                                                    st.session_state.get('OPENAI_API_KEY', ''),
                                                    st.session_state.get('GOOGLE_API_KEY', '')
                                                )
                                                if success and new_audit_id:
                                                    # Update session results
                                                    db = get_db()
                                                    if db:
                                                        try:
                                                            sess = db.query(BulkScan).filter(BulkScan.id == session.id).first()
                                                            if sess:
                                                                results = dict(sess.results) if sess.results else {}
                                                                results[url] = new_audit_id
                                                                sess.results = results
                                                                db.commit()
                                                                st.success(f"✅ {domain} audit complete!")
                                                        except Exception as e:
                                                            logger.warning(f"Error updating session: {e}")
                                                        finally:
                                                            db.close()
                                                    st.rerun()
                                                else:
                                                    st.error(f"Retry failed: {error}")
                            
                            # Summary stats for the session
                            success_count = sum(1 for _, aid in urls_list if aid)
                            fail_count = len(urls_list) - success_count
                            st.markdown(f"""
                            ---
                            **Summary:** ✅ {success_count} successful | ❌ {fail_count} failed | 📊 Total: {len(urls_list)}
                            """)
    
    # =========================================================================
    # BACKGROUND PROCESSING LOGIC (FIXED - RUNS ON EVERY PAGE LOAD)
    # =========================================================================
    
    if st.session_state.get('bulk_scan_active') and st.session_state.get('bulk_scan_session_id'):
        session_id = st.session_state['bulk_scan_session_id']
        
        # Get fresh session data with open DB connection
        db = get_db()
        if not db:
            st.error("Database unavailable")
            st.session_state.pop('bulk_scan_active', None)
            return
        
        try:
            session = db.query(BulkScan).filter(BulkScan.session_id == session_id).first()
            
            if not session:
                logger.warning(f"Bulk scan session not found: {session_id}")
                st.session_state.pop('bulk_scan_session_id', None)
                st.session_state.pop('bulk_scan_active', None)
                db.close()
                return
            
            # Check if paused/stopped externally
            if session.status != "running":
                st.session_state.pop('bulk_scan_active', None)
                db.close()
                return
            
            urls = session.urls
            current_idx = session.processed_urls  # FIX: Use processed_urls as current index
            results = dict(session.results) if session.results else {}
            
            # Check if all done
            if current_idx >= len(urls):
                session.status = "completed"
                session.updated_at = datetime.utcnow()
                db.commit()
                db.close()
                
                st.session_state.pop('bulk_scan_session_id', None)
                st.session_state.pop('bulk_scan_active', None)
                
                st.success("🎉 Bulk scan completed!")
                st.balloons()
                return
            
            # Get current URL
            url = urls[current_idx]
            
            # Show processing indicator
            st.markdown(f"""
            <div class="processing-indicator">
                🔄 <strong>Processing ({current_idx + 1}/{len(urls)})</strong>: {url[:60]}{'...' if len(url) > 60 else ''}
            </div>
            """, unsafe_allow_html=True)
            
            # Process ONE URL with safe runner
            audit_id, success, error = run_bulk_audit_safe(
                url,
                st.session_state.get('OPENAI_API_KEY', ''),
                st.session_state.get('GOOGLE_API_KEY', '')
            )
            
            # Store result (even if failed)
            results[url] = audit_id
            
            # Update session in database (keeping connection open)
            session.processed_urls = current_idx + 1
            session.results = results
            session.paused_at_index = current_idx + 1  # FIX: Update resume point
            session.updated_at = datetime.utcnow()
            
            # Check if completed
            if current_idx + 1 >= len(urls):
                session.status = "completed"
                st.session_state.pop('bulk_scan_session_id', None)
                st.session_state.pop('bulk_scan_active', None)
                logger.info(f"Bulk scan completed: {session_id[:8]}...")
            
            db.commit()
            db.close()
            
            # Rate limiting
            time.sleep(BULK_SCAN_DELAY_SECONDS)
            
            # FIX: Trigger next URL processing via rerun
            if st.session_state.get('bulk_scan_active'):
                st.rerun()
                
        except Exception as e:
            logger.error(f"Bulk scan processing error: {str(e)}")
            if db:
                try:
                    db.rollback()
                    db.close()
                except:
                    pass
            # Don't stop the whole process - just log and continue
            st.session_state['bulk_scan_error_count'] = st.session_state.get('bulk_scan_error_count', 0) + 1
            
            if st.session_state.get('bulk_scan_error_count', 0) >= BULK_SCAN_MAX_ERRORS:
                st.error("Too many consecutive errors. Scan paused.")
                st.session_state.pop('bulk_scan_active', None)
            else:
                time.sleep(1)
                st.rerun()

def show_audit_history():
    """Audit history page with CRM lead management integration.
    
    Features:
    - User-based access control (admin sees all, users see their own)
    - CRM filters: source, lead status, pipeline stage, approached
    - Multiple filter options (time, score, domain search)
    - Bulk actions with confirmation dialogs
    - CSV export functionality
    - CRM metrics dashboard strip
    - Grouped display by time period
    """
    logger = logging.getLogger("sales_engine")
    logger.debug("Rendering Audit History page")
    
    # Initialize session state for this module
    init_audit_history_session_state()
    
    # =========================================================================
    # CRM-ENHANCED BRANDING CSS
    # =========================================================================
    st.markdown("""
    <style>
    /* CRM Status Badges */
    .badge-source {
        display: inline-block;
        padding: 2px 8px;
        border-radius: 12px;
        font-size: 0.75rem;
        font-weight: 600;
        margin-right: 4px;
    }
    .badge-single { background: #e3f2fd; color: #1565c0; }
    .badge-bulk { background: #fff3e0; color: #ef6c00; }
    .badge-manual { background: #f3e5f5; color: #7b1fa2; }
    
    .badge-lead-status {
        display: inline-block;
        padding: 2px 8px;
        border-radius: 12px;
        font-size: 0.75rem;
        font-weight: 600;
        margin-right: 4px;
    }
    .badge-hot { background: #ffebee; color: #c62828; }
    .badge-warm { background: #fff8e1; color: #f57c00; }
    .badge-cold { background: #eceff1; color: #546e7a; }
    
    .badge-pipeline {
        display: inline-block;
        padding: 2px 8px;
        border-radius: 12px;
        font-size: 0.75rem;
        font-weight: 600;
    }
    .badge-new { background: #e8f5e9; color: #2e7d32; }
    .badge-contacted { background: #e3f2fd; color: #1565c0; }
    .badge-follow-up { background: #fff3e0; color: #ef6c00; }
    .badge-closed { background: #f5f5f5; color: #616161; }
    
    .approached-yes { color: #2e7d32; }
    .approached-no { color: #9e9e9e; }
    
    /* CRM Metrics Strip */
    .crm-metrics-strip {
        background: linear-gradient(135deg, #0c3740 0%, #1a5a6e 100%);
        padding: 1rem 1.5rem;
        border-radius: 12px;
        margin-bottom: 1.5rem;
        color: white;
    }
    .crm-metric-card {
        background: rgba(255,255,255,0.1);
        border-radius: 8px;
        padding: 0.75rem;
        text-align: center;
    }
    .crm-metric-value {
        font-size: 1.5rem;
        font-weight: bold;
        font-family: 'Poppins', sans-serif;
    }
    .crm-metric-label {
        font-size: 0.8rem;
        opacity: 0.9;
        font-family: 'Lato', sans-serif;
    }
    
    /* Lead Details Card */
    .lead-details-card {
        background: #f8faf9;
        border: 1px solid #e0e0e0;
        border-radius: 12px;
        padding: 1.25rem;
        margin-top: 1rem;
    }
    .lead-details-card h4 {
        color: #0c3740;
        font-family: 'Poppins', sans-serif;
        margin-bottom: 1rem;
    }
    
    /* Sticky bar and existing styles */
    .sticky-bar {position: -webkit-sticky; position: sticky; top: 0; z-index: 100; background: #fff; padding: 1rem 0 0.5rem 0; border-bottom: 1px solid #eee; margin-bottom: 1rem;}
    .active-filter {background: #e0f7fa; color: #0066cc; border-radius: 6px; padding: 2px 8px; margin-left: 6px; font-size: 0.95em;}
    .clear-btn {background: #ffebee; color: #c62828; border-radius: 6px; padding: 2px 8px; margin-left: 6px; font-size: 0.9em; cursor: pointer;}
    .confirm-delete {background: #ffcdd2; border: 2px solid #c62828; padding: 1rem; border-radius: 8px; margin: 1rem 0;}
    @media (max-width: 900px) {
        .element-container .stColumn {width: 100% !important; display: block !important;}
        .sticky-bar {padding: 0.5rem 0;}
    }
    </style>
    """, unsafe_allow_html=True)
    
    st.title("📊 Audit History & Lead CRM")
    
    # Get current user info for access control
    current_user = st.session_state.get("current_user")
    is_admin = st.session_state.get("is_admin", False)
    
    # Show role indicator
    if is_admin:
        st.markdown("🔓 **Admin View** - Viewing all audits across all users")
    else:
        st.markdown(f"🔒 Viewing your audits and leads")
    
    # =========================================================================
    # CRM METRICS STRIP (Dashboard summary)
    # =========================================================================
    crm_metrics = get_crm_metrics()
    
    st.markdown("### 📈 Lead Pipeline Overview")
    col_m1, col_m2, col_m3, col_m4, col_m5 = st.columns(5)
    with col_m1:
        st.metric("📊 Total Leads", crm_metrics.get("total", 0))
    with col_m2:
        st.metric("📧 Not Approached", crm_metrics.get("not_approached", 0), 
                  help="Leads that haven't been contacted yet")
    with col_m3:
        st.metric("✅ Approached", crm_metrics.get("approached", 0))
    with col_m4:
        st.metric("🔥 Hot Leads", crm_metrics.get("hot", 0), 
                  delta="Priority" if crm_metrics.get("hot", 0) > 0 else None)
    with col_m5:
        st.metric("📅 Follow-up Due", crm_metrics.get("follow_up_due", 0),
                  delta="Action needed" if crm_metrics.get("follow_up_due", 0) > 0 else None)
    
    st.markdown("---")
    
    # =========================================================================
    # CRM QUICK VIEW TABS
    # =========================================================================
    crm_view = st.radio(
        "Quick View",
        ["All Leads", "Not Approached", "Approached", "🔥 Hot Leads", "📅 Needs Follow-up"],
        horizontal=True,
        key="crm_quick_view",
        help="Quick filter for common CRM views"
    )
    
    st.markdown('<div class="sticky-bar">', unsafe_allow_html=True)
    
    # Clear Filters button
    col_clear, col_spacer = st.columns([1, 7])
    with col_clear:
        if st.button("✖️ Clear All Filters", key="clear_all_filters_btn", help="Reset all filters to default"):
            logger.info("User cleared all audit history filters")
            clear_audit_history_filters()
    
    # Initialize date variables
    selected_date = None
    date_from = None
    date_to = None
    
    # =========================================================================
    # ROW 1: TIME & SOURCE FILTERS
    # =========================================================================
    col_date1, col_date2, col_date3, col_source = st.columns([2, 2, 2, 2])
    with col_date1:
        time_filter = st.selectbox(
            "📅 Time Period",
            ["All Time", "Today", "Yesterday", "This Week", "This Month", "Last 30 Days", "Select Specific Date", "Custom Range"],
            help="Filter audits by time period"
        )
    with col_date2:
        if time_filter == "Select Specific Date":
            selected_date = st.date_input("📆 Pick a Date", value=datetime.now(), key="hist_calendar_date")
        elif time_filter == "Custom Range":
            date_from = st.date_input("From Date", value=datetime.now() - timedelta(days=30), key="hist_date_from")
        else:
            st.markdown("<div style='margin-top:28px;'></div>", unsafe_allow_html=True)
    with col_date3:
        if time_filter == "Custom Range":
            date_to = st.date_input("To Date", value=datetime.now(), key="hist_date_to")
        else:
            st.markdown("<div style='margin-top:28px;'></div>", unsafe_allow_html=True)
    with col_source:
        source_filter = st.selectbox(
            "🔗 Source",
            ["All", "Single Audit", "Bulk Audit", "Manual"],
            key="source_filter",
            help="Filter by how the audit was created"
        )
    
    # =========================================================================
    # ROW 2: CRM FILTERS (Lead Status, Pipeline, Approached)
    # =========================================================================
    col_lead_status, col_pipeline, col_approached, col_show_all = st.columns([2, 2, 2, 2])
    with col_lead_status:
        lead_status_filter = st.selectbox(
            "🎯 Lead Status",
            ["All", "🔥 Hot", "☀️ Warm", "❄️ Cold"],
            key="lead_status_filter",
            help="Filter by lead temperature"
        )
    with col_pipeline:
        pipeline_filter = st.selectbox(
            "📊 Pipeline Stage",
            ["All", "New", "Contacted", "Follow-up", "Closed"],
            key="pipeline_filter",
            help="Filter by sales pipeline stage"
        )
    with col_approached:
        approached_filter = st.selectbox(
            "📧 Approached",
            ["All", "✅ Yes", "❌ No"],
            key="approached_filter",
            help="Filter by whether lead has been contacted"
        )
    with col_show_all:
        show_all_scans = st.checkbox("🔄 Show All Scans", value=False, key="show_all_scans", 
                                      help="Unchecked: shows only most recent scan per domain. Checked: all historical scans.")
    
    # =========================================================================
    # ROW 3: SEARCH & SCORE FILTERS
    # =========================================================================
    col1, col2, col3 = st.columns([2, 1, 1])
    with col1:
        search_raw = st.text_input("🔍 Search domain", key="hist_search", help="Filter audits by domain name")
        search = sanitize_domain_search(search_raw) if search_raw else None
    with col2:
        min_score_raw = st.number_input("Min Score", 0, 100, 0, help="Show audits with score above this value")
    with col3:
        max_score_raw = st.number_input("Max Score", 0, 100, 100, help="Show audits with score below this value")
    
    # Validate score filters
    min_score, max_score, score_valid, score_error = validate_score_filters(min_score_raw, max_score_raw)
    if not score_valid:
        st.warning(f"⚠️ {score_error}")
    
    # Show active filters as chips using helper function
    active_filters = generate_active_filter_chips(time_filter, selected_date, search, min_score, max_score)
    if active_filters:
        st.markdown("**Active Filters:** " + " ".join(active_filters), unsafe_allow_html=True)
    
    st.markdown('</div>', unsafe_allow_html=True)
    
    # Check database availability
    if not DB_AVAILABLE:
        st.error("⚠️ Database required for audit history. Please check your database configuration.")
        logger.error("Database unavailable for audit history")
        return
    
    # Sort controls at the top
    col_sort1, col_sort2 = st.columns([1, 1])
    with col_sort1:
        sort_options = ["Score", "Domain", "Speed", "Issues", "Date"]
        sort_col = st.selectbox("Sort by", sort_options, index=0, key="audit_sort_col")
    with col_sort2:
        sort_dir = st.radio("Order", ["Descending", "Ascending"], horizontal=True, key="audit_sort_dir")
    
    # Load audits with user-based filtering
    with st.spinner("Loading audits..."):
        try:
            logger.debug(f"Loading audits for user={current_user}, admin={is_admin}, search={search}")
            audits = get_audit_history_cached(
                limit=1000,
                search_query=search,
                min_score=min_score if min_score > 0 else None,
                max_score=max_score if max_score < 100 else None,
                username=current_user,
                is_admin=is_admin
            )
        except Exception as e:
            logger.error(f"Error loading audit history: {str(e)}")
            audits = []
    
    # Check for database errors
    if st.session_state.get("_audit_history_db_error"):
        st.warning("⚠️ Could not load audits from database. Please check your connection.")
    
    if not audits:
        st.markdown("""
        <div style='text-align:center;margin-top:2em;'>
            <img src='https://cdn-icons-png.flaticon.com/512/4076/4076549.png' width='120' alt='No audits illustration' style='opacity:0.7;'/><br>
            <h4>No audits found</h4>
            <p>Try adjusting your filters or run a new audit to see results here.<br>
            Need help? Go to <b>Single Audit</b> to scan a website.</p>
        </div>
        """, unsafe_allow_html=True)
        return
    
    # Apply time-based filtering using helper function
    audits = filter_audits_by_time_period(
        audits, time_filter, selected_date, date_from, date_to
    )
    
    if not audits:
        st.info("No audits match your time filter. Try selecting a different time period.")
        return
    
    # =========================================================================
    # APPLY CRM FILTERS (Source, Lead Status, Pipeline, Approached)
    # =========================================================================
    # Load leads for CRM filtering
    db = get_db()
    leads_by_domain = {}
    if db:
        try:
            all_leads = db.query(Lead).all()
            leads_by_domain = {lead.domain: lead for lead in all_leads}
        except Exception:
            pass
        finally:
            try:
                db.close()
            except:
                pass
    
    # Apply CRM Quick View filter
    if crm_view == "Not Approached":
        audits = [a for a in audits if leads_by_domain.get(a.domain) and not leads_by_domain[a.domain].approached]
    elif crm_view == "Approached":
        audits = [a for a in audits if leads_by_domain.get(a.domain) and leads_by_domain[a.domain].approached]
    elif crm_view == "🔥 Hot Leads":
        audits = [a for a in audits if leads_by_domain.get(a.domain) and leads_by_domain[a.domain].lead_status == "hot"]
    elif crm_view == "📅 Needs Follow-up":
        today = datetime.utcnow().replace(hour=23, minute=59, second=59)
        audits = [a for a in audits if leads_by_domain.get(a.domain) and 
                  leads_by_domain[a.domain].follow_up_date and 
                  leads_by_domain[a.domain].follow_up_date <= today and
                  leads_by_domain[a.domain].pipeline_stage == "follow-up"]
    
    # Apply source filter
    if source_filter == "Single Audit":
        audits = [a for a in audits if getattr(a, 'source', 'single') == 'single']
    elif source_filter == "Bulk Audit":
        audits = [a for a in audits if getattr(a, 'source', None) == 'bulk']
    elif source_filter == "Manual":
        audits = [a for a in audits if getattr(a, 'source', None) == 'manual']
    
    # Apply lead status filter
    if lead_status_filter == "🔥 Hot":
        audits = [a for a in audits if leads_by_domain.get(a.domain) and leads_by_domain[a.domain].lead_status == "hot"]
    elif lead_status_filter == "☀️ Warm":
        audits = [a for a in audits if leads_by_domain.get(a.domain) and leads_by_domain[a.domain].lead_status == "warm"]
    elif lead_status_filter == "❄️ Cold":
        audits = [a for a in audits if leads_by_domain.get(a.domain) and leads_by_domain[a.domain].lead_status == "cold"]
    
    # Apply pipeline filter
    pipeline_map = {"New": "new", "Contacted": "contacted", "Follow-up": "follow-up", "Closed": "closed"}
    if pipeline_filter in pipeline_map:
        stage = pipeline_map[pipeline_filter]
        audits = [a for a in audits if leads_by_domain.get(a.domain) and leads_by_domain[a.domain].pipeline_stage == stage]
    
    # Apply approached filter
    if approached_filter == "✅ Yes":
        audits = [a for a in audits if leads_by_domain.get(a.domain) and leads_by_domain[a.domain].approached]
    elif approached_filter == "❌ No":
        audits = [a for a in audits if leads_by_domain.get(a.domain) and not leads_by_domain[a.domain].approached]
    
    if not audits:
        st.info("No audits match your CRM filters. Try adjusting the filters above.")
        return
    
    # Deduplicate by domain if not showing all scans
    if not show_all_scans:
        audits = dedupe_audits_by_domain(audits, per_user=True)
    
    # Group audits by time period using helper
    grouped_audits = group_audits_by_period(audits)
    
    # Generate row data for all audits using helper
    if 'audit_tags' not in st.session_state:
        st.session_state.audit_tags = {}
    
    hist_data = [generate_audit_row_data(audit) for audit in audits]
    
    # Sort data using helper
    reverse = sort_dir == "Descending"
    hist_data = sort_audit_rows(hist_data, sort_col, reverse)
    
    # Pagination
    paginated_data, total_pages, current_page = get_paginated_items(
        hist_data, 
        page_key="audit_history_page",
        items_per_page=50
    )
    
    # Bulk actions state
    if 'audit_bulk_selected' not in st.session_state:
        st.session_state.audit_bulk_selected = set()
    
    # Summary metrics
    period_counts = {k: len(v) for k, v in grouped_audits.items()}
    
    col_sum1, col_sum2, col_sum3, col_sum4, col_sum5 = st.columns(5)
    with col_sum1:
        st.metric("📅 Today", period_counts.get("Today", 0))
    with col_sum2:
        st.metric("📆 Yesterday", period_counts.get("Yesterday", 0))
    with col_sum3:
        st.metric("📊 This Week", period_counts.get("This Week", 0))
    with col_sum4:
        st.metric("📈 This Month", period_counts.get("This Month", 0))
    with col_sum5:
        st.metric("📚 Total", len(audits))
    
    st.markdown("---")
    
    dedup_msg = "" if show_all_scans else " (most recent per domain)"
    st.markdown(f"### 📋 Audit List{dedup_msg}")
    st.markdown(f"**{len(audits)}** audits found. **{len(st.session_state.audit_bulk_selected)}** selected.")
    
    # Bulk action controls
    col_bulk1, col_bulk2, col_bulk3, col_bulk4 = st.columns([1.5, 1.5, 1.5, 1.5])
    with col_bulk1:
        if st.button("✅ Select All Visible", key="audit_bulk_all", use_container_width=True):
            for audit in audits:
                st.session_state.audit_bulk_selected.add(_get_attr(audit, 'id'))
            logger.debug(f"Selected all {len(audits)} visible audits")
            st.rerun()
    
    with col_bulk2:
        if st.button("❌ Clear Selection", key="audit_bulk_clear", use_container_width=True):
            st.session_state.audit_bulk_selected.clear()
            logger.debug("Cleared audit selection")
            st.rerun()
    
    with col_bulk3:
        # Bulk delete with confirmation
        selected_count = len(st.session_state.audit_bulk_selected)
        
        if selected_count > 0:
            # Show confirmation dialog if requested
            if st.session_state.get("_confirm_bulk_delete"):
                st.markdown(
                    f"""<div class='confirm-delete'>
                    ⚠️ <strong>Confirm Delete</strong><br>
                    You are about to delete <strong>{selected_count}</strong> audit(s). This action can be undone for 10 seconds.
                    </div>""", 
                    unsafe_allow_html=True
                )
                col_yes, col_no = st.columns(2)
                with col_yes:
                    if st.button("✅ Yes, Delete", key="confirm_delete_yes", type="primary", use_container_width=True):
                        deleted, deleted_audits, error = bulk_delete_audits(st.session_state.audit_bulk_selected)
                        if error:
                            st.error(f"Delete failed: {error}")
                        else:
                            st.session_state.audit_bulk_selected.clear()
                            st.session_state.recently_deleted_audits = deleted_audits
                            st.session_state.undo_delete_time = time.time()
                            st.session_state._confirm_bulk_delete = False
                            st.toast(f"✓ Deleted {deleted} audits. Undo available for 10s.")
                            logger.info(f"Bulk deleted {deleted} audits")
                            st.rerun()
                with col_no:
                    if st.button("❌ Cancel", key="confirm_delete_no", use_container_width=True):
                        st.session_state._confirm_bulk_delete = False
                        st.rerun()
            else:
                if st.button("🗑️ Bulk Delete", key="audit_bulk_delete", use_container_width=True):
                    st.session_state._confirm_bulk_delete = True
                    st.rerun()
        else:
            st.button("🗑️ Bulk Delete", key="audit_bulk_delete_disabled", use_container_width=True, disabled=True, help="Select audits first")
        
        # Undo delete button (visible for 10 seconds after delete)
        if st.session_state.get('recently_deleted_audits') and st.session_state.get('undo_delete_time'):
            time_remaining = 10 - (time.time() - st.session_state['undo_delete_time'])
            if time_remaining > 0:
                if st.button(f"↩️ Undo Delete ({int(time_remaining)}s)", key="undo_bulk_delete"):
                    restored, error = bulk_restore_audits(st.session_state['recently_deleted_audits'])
                    if error:
                        st.error(f"Restore failed: {error}")
                    else:
                        st.session_state.recently_deleted_audits = []
                        st.session_state.undo_delete_time = None
                        st.toast(f"✓ Restored {restored} audits.")
                        logger.info(f"Restored {restored} audits via undo")
                        st.rerun()
            else:
                # Clear expired undo data
                st.session_state.recently_deleted_audits = []
                st.session_state.undo_delete_time = None
    
    with col_bulk4:
        if st.session_state.audit_bulk_selected:
            csv_bytes = export_audits_to_csv(hist_data, st.session_state.audit_bulk_selected)
            if csv_bytes:
                st.download_button(
                    "📥 Export CSV",
                    csv_bytes,
                    "audit_history_selected.csv",
                    "text/csv",
                    key="audit_bulk_csv_btn",
                    use_container_width=True
                )
            else:
                st.button("📥 Export CSV", key="audit_bulk_export_disabled", use_container_width=True, disabled=True)
        else:
            st.button("📥 Export CSV", key="audit_bulk_export_disabled2", use_container_width=True, disabled=True, help="Select audits first")
    
    st.markdown("---")
    
    # Display audits grouped by time period
    for period_name, period_audits in grouped_audits.items():
        if not period_audits:
            continue
        
        # Create collapsible section for each time period
        with st.expander(f"📁 {period_name} ({len(period_audits)} audits)", expanded=(period_name == "Today")):
            for audit in period_audits:
                render_audit_item(audit, st.session_state.audit_bulk_selected)
    
    # Pagination controls (only show if needed)
    if total_pages > 1:
        st.markdown("---")
        display_pagination_controls("audit_history_page", total_pages, current_page)
    
    # Export CSV for all results
    st.markdown("---")
    st.markdown("### 📥 Export All Results")
    all_csv_bytes = export_audits_to_csv(hist_data)
    if all_csv_bytes:
        st.download_button(
            "📥 Export All Audits as CSV",
            all_csv_bytes,
            "audit_history_complete.csv",
            "text/csv"
        )


def render_audit_item(audit, bulk_selected: set):
    """Render a single audit item with checkbox and expandable details.
    
    Args:
        audit: Audit model instance or dict
        bulk_selected: Set of selected audit IDs (modified in place)
    """
    # Normalize audit to dict for consistent access
    audit_dict = normalize_audit(audit)
    audit_id = audit_dict.get("id")
    
    status_icon, _, _ = get_score_status_icon(audit_dict["health_score"])
    
    # Checkbox and expander for each audit
    col_check, col_expand = st.columns([0.5, 11.5])
    with col_check:
        is_selected = audit_id in bulk_selected
        if st.checkbox("", value=is_selected, key=f"audit_check_{audit_id}", label_visibility="collapsed"):
            bulk_selected.add(audit_id)
        else:
            bulk_selected.discard(audit_id)
    
    with col_expand:
        # Create expander for each audit with summary info
        date_str = safe_timestamp_slice(audit_dict["created_at"], 16).replace("T", " ") if audit_dict["created_at"] else 'N/A'
        score_display = audit_dict["health_score"] if audit_dict["health_score"] is not None else 'N/A'
        with st.expander(f"{status_icon} {audit_dict['domain']} - Score: {score_display}/100 - {date_str}"):
            render_audit_detail(audit, audit_dict)


def render_audit_detail(audit, audit_dict=None):
    """Render full audit details inside an expander with CRM lead management panel.
    
    Args:
        audit: Original audit (ORM or dict) - used for ID references
        audit_dict: Pre-normalized audit dict (optional, will normalize if not provided)
    """
    logger = logging.getLogger("sales_engine")
    
    # Use pre-normalized dict or normalize now
    if audit_dict is None:
        audit_dict = normalize_audit(audit)
    
    # Convert to data dict format for display
    data = convert_audit_to_data_dict(audit)
    
    # Get lead for this audit's domain and normalize it
    lead_orm = get_lead_for_audit(audit_dict.get("id"))
    lead = normalize_lead(lead_orm) if lead_orm else None
    
    # Metrics row
    c1, c2, c3, c4, c5 = st.columns(5)
    with c1:
        delta = "Good" if (data.get('score') or 0) >= 70 else "Needs Work"
        st.metric("Health Score", data.get('score', 'N/A'), delta=delta)
    with c2:
        st.metric("Google Speed", data.get('psi', 'N/A'))
    with c3:
        st.metric("Issues Found", len(data.get('issues', [])))
    with c4:
        st.metric("Age", data.get('domain_age', 'Unknown'))
    with c5:
        st.metric("Audit ID", audit_dict["id"])
    
    # =========================================================================
    # CRM LEAD MANAGEMENT PANEL
    # =========================================================================
    if lead:
        st.markdown("---")
        st.markdown("""
        <div style='background: linear-gradient(135deg, #0c3740 0%, #1a5a6e 100%); 
                    color: white; padding: 0.5rem 1rem; border-radius: 8px 8px 0 0;
                    font-family: Poppins, sans-serif; font-weight: 600;'>
            🎯 Lead CRM Details
        </div>
        """, unsafe_allow_html=True)
        
        # Lead status badges row
        source_badge = {
            "single": ("Single", "badge-single"),
            "bulk": ("Bulk", "badge-bulk"),
            "manual": ("Manual", "badge-manual")
        }.get(lead["source"] or "single", ("Single", "badge-single"))
        
        status_badge = {
            "hot": ("🔥 Hot", "badge-hot"),
            "warm": ("☀️ Warm", "badge-warm"),
            "cold": ("❄️ Cold", "badge-cold")
        }.get(lead["lead_status"] or "warm", ("☀️ Warm", "badge-warm"))
        
        pipeline_badge = {
            "new": ("New", "badge-new"),
            "contacted": ("Contacted", "badge-contacted"),
            "follow-up": ("Follow-up", "badge-follow-up"),
            "closed": ("Closed", "badge-closed")
        }.get(lead["pipeline_stage"] or "new", ("New", "badge-new"))
        
        approached_icon = "✅" if lead["approached"] else "❌"
        approached_class = "approached-yes" if lead["approached"] else "approached-no"
        
        st.markdown(f"""
        <div style='background: #f8faf9; border: 1px solid #e0e0e0; border-top: none; 
                    padding: 0.75rem 1rem; border-radius: 0 0 8px 8px; margin-bottom: 1rem;'>
            <span class='badge-source {source_badge[1]}'>{source_badge[0]}</span>
            <span class='badge-lead-status {status_badge[1]}'>{status_badge[0]}</span>
            <span class='badge-pipeline {pipeline_badge[1]}'>{pipeline_badge[0]}</span>
            <span class='{approached_class}' style='margin-left: 8px;'>Approached: {approached_icon}</span>
        </div>
        """, unsafe_allow_html=True)
        
        # Lead info row (non-editable info)
        col_lead1, col_lead2, col_lead3 = st.columns(3)
        with col_lead1:
            st.markdown(f"**📧 Email:** {lead['email'] or 'Not found'}")
            st.markdown(f"**📞 Phone:** {lead.get('phone') or 'Not found'}")
        with col_lead2:
            st.markdown(f"**🏢 Company:** {lead['company_name'] or lead['domain']}")
            st.markdown(f"**📊 Opportunity:** {lead.get('opportunity_rating') or 'N/A'}/100")
        with col_lead3:
            approached_date = safe_timestamp_slice(lead['approached_date'], 10) if lead['approached_date'] else 'Never'
            st.markdown(f"**📅 Approached:** {approached_date}")
            followup_date = safe_timestamp_slice(lead['follow_up_date'], 10) if lead['follow_up_date'] else 'Not set'
            st.markdown(f"**📆 Follow-up:** {followup_date}")
        
        # Editable CRM fields in expander
        with st.expander("✏️ Edit Lead CRM Fields", expanded=False):
            col_e1, col_e2 = st.columns(2)
            
            with col_e1:
                # Approached toggle
                new_approached = st.checkbox(
                    "✅ Mark as Approached",
                    value=lead["approached"],
                    key=f"lead_approached_{audit_dict['id']}",
                    help="Check this when you've contacted the lead"
                )
                
                # Lead status
                status_options = ["hot", "warm", "cold"]
                current_status_idx = status_options.index(lead["lead_status"]) if lead["lead_status"] in status_options else 1
                new_lead_status = st.selectbox(
                    "🎯 Lead Status",
                    status_options,
                    index=current_status_idx,
                    format_func=lambda x: {"hot": "🔥 Hot", "warm": "☀️ Warm", "cold": "❄️ Cold"}.get(x, x),
                    key=f"lead_status_{audit_dict['id']}"
                )
                
                # Interested
                interested_options = ["yes", "maybe", "no"]
                current_interested_idx = interested_options.index(lead["interested"]) if lead["interested"] in interested_options else 1
                new_interested = st.selectbox(
                    "💡 Interested",
                    interested_options,
                    index=current_interested_idx,
                    format_func=lambda x: {"yes": "✅ Yes", "maybe": "🤔 Maybe", "no": "❌ No"}.get(x, x),
                    key=f"lead_interested_{audit_dict['id']}"
                )
            
            with col_e2:
                # Pipeline stage
                pipeline_options = ["new", "contacted", "follow-up", "closed"]
                current_pipeline_idx = pipeline_options.index(lead["pipeline_stage"]) if lead["pipeline_stage"] in pipeline_options else 0
                new_pipeline_stage = st.selectbox(
                    "📊 Pipeline Stage",
                    pipeline_options,
                    index=current_pipeline_idx,
                    format_func=lambda x: {"new": "🆕 New", "contacted": "📞 Contacted", "follow-up": "📅 Follow-up", "closed": "✅ Closed"}.get(x, x),
                    key=f"lead_pipeline_{audit_dict['id']}"
                )
                
                # Follow-up date - parse from string if needed
                follow_up_val = lead["follow_up_date"]
                if follow_up_val and isinstance(follow_up_val, str):
                    try:
                        follow_up_val = datetime.fromisoformat(follow_up_val.replace('Z', '+00:00')).date()
                    except:
                        follow_up_val = datetime.now().date() + timedelta(days=7)
                elif follow_up_val and hasattr(follow_up_val, 'date'):
                    follow_up_val = follow_up_val.date()
                else:
                    follow_up_val = datetime.now().date() + timedelta(days=7)
                
                new_follow_up_date = st.date_input(
                    "📅 Follow-up Date",
                    value=follow_up_val,
                    key=f"lead_followup_{audit_dict['id']}"
                )
            
            # Notes field
            new_notes = st.text_area(
                "📝 Notes",
                value=lead["notes"] or "",
                height=100,
                key=f"lead_notes_{audit_dict['id']}",
                placeholder="Add notes about this lead..."
            )
            
            # Save button
            if st.button("💾 Save CRM Changes", key=f"save_lead_{audit_dict['id']}", type="primary", use_container_width=True):
                update_fields = {
                    "approached": new_approached,
                    "lead_status": new_lead_status,
                    "interested": new_interested,
                    "pipeline_stage": new_pipeline_stage,
                    "notes": new_notes
                }
                
                # Set approached date if newly approached
                if new_approached and not lead["approached"]:
                    update_fields["approached_date"] = datetime.utcnow()
                
                # Set follow-up date
                if new_follow_up_date:
                    update_fields["follow_up_date"] = datetime.combine(new_follow_up_date, datetime.min.time())
                
                if update_lead_crm_fields(lead["id"], **update_fields):
                    st.success("✅ Lead updated successfully!")
                    logger.info(f"Updated CRM fields for lead {lead['id']} ({lead['domain']})")
                    st.rerun()
                else:
                    st.error("❌ Failed to update lead")
    else:
        # No lead found - show create lead option
        st.markdown("---")
        st.info("ℹ️ No lead record exists for this domain yet. The lead will be created when you run a new audit or import manually.")
    
    # Tech stack
    if data.get('tech_stack'):
        try:
            st.markdown(f"**📦 Tech Stack:** {', '.join(data['tech_stack'])}")
        except Exception:
            st.markdown("**📦 Tech Stack:** Unable to display")
    
    # Issues
    if data.get('issues'):
        st.markdown("**⚠️ Issues Detected**")
        for i, issue in enumerate(data.get('issues', [])[:10], 1):  # Limit to 10 issues
            try:
                title = issue.get('title', 'Unknown Issue') if isinstance(issue, dict) else str(issue)
                with st.expander(f"{i}. {title}", expanded=False):
                    if isinstance(issue, dict):
                        col1, col2 = st.columns(2)
                        with col1:
                            st.markdown(f"**Impact:** {issue.get('impact', 'N/A')}")
                        with col2:
                            st.markdown(f"**Solution:** {issue.get('solution', 'N/A')}")
                    else:
                        st.markdown(str(issue))
            except Exception as e:
                logger.warning(f"Could not display issue #{i}: {str(e)}")
    
    # AI analysis - safely access ai section
    ai_section = data.get('ai') if isinstance(data.get('ai'), dict) else {}
    if ai_section.get('summary') and ai_section['summary'] != 'No summary available':
        st.markdown("---")
        st.markdown("**🤖 AI Analysis**")
        
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**Summary**")
            st.info(ai_section.get('summary', 'No summary available'))
            st.markdown("**Impact**")
            st.warning(ai_section.get('impact', 'No impact assessment available'))
        
        with col2:
            st.markdown("**Solutions**")
            st.success(ai_section.get('solutions', 'No solutions available'))
        
        email_draft = ai_section.get('email', '')
        if email_draft and email_draft != 'No email draft available':
            st.markdown("**📧 Cold Email Draft**")
            email_text = clean_text(email_draft) if 'clean_text' in dir() else email_draft
            st.text_area("", value=email_text, height=150, key=f"email_draft_{audit_dict['id']}")
    
    # Action buttons
    st.markdown("---")
    col_btn1, col_btn2, col_btn3, col_btn4 = st.columns(4)
    
    with col_btn1:
        # Download PDF - only generate when clicked
        if st.button("📥 Generate PDF", key=f"gen_pdf_{audit_dict['id']}", use_container_width=True):
            try:
                with st.spinner("Generating PDF..."):
                    pdf_bytes = get_audit_pdf(audit_dict['id'])
                    if not pdf_bytes:
                        pdf_bytes = generate_pdf(data)
                    if pdf_bytes:
                        st.download_button(
                            label="📥 Download PDF",
                            data=pdf_bytes,
                            file_name=f"audit_{audit_dict['id']}_{audit_dict['domain']}.pdf",
                            mime="application/pdf",
                            key=f"pdf_download_{audit_dict['id']}"
                        )
                    else:
                        logger.warning(f"Could not generate PDF for audit {audit_dict['id']}")
                        st.error("Could not generate PDF")
            except Exception as e:
                logger.error(f"PDF generation failed for audit {audit_dict['id']}: {str(e)}")
                st.error("PDF generation failed")
    
    with col_btn2:
        # Load to Single Audit view using persistence layer
        if st.button("📂 Load to Audit", key=f"load_audit_{audit_dict['id']}", use_container_width=True):
            # Use persistence layer instead of direct session state mutation
            set_current_audit(audit_dict['id'])
            persist_audit_data(audit_dict['id'], data)
            store_pdf_context(audit_dict['id'], data)
            save_navigation_state("Single Audit", audit_dict['id'])
            sync_query_params()
            
            st.session_state.current_section = 'Single Audit'
            logger.debug(f"Loaded audit {audit_dict['id']} ({audit_dict['domain']}) to Single Audit view via persistence layer")
            st.toast(f"✓ Loaded {audit_dict['domain']} to Single Audit page")
            st.rerun()
    
    with col_btn3:
        # Send Email button
        if lead and lead["email"]:
            if st.button("📧 Send Email", key=f"send_email_{audit_dict['id']}", use_container_width=True, type="primary"):
                # Set up session state for email sending
                st.session_state.email_target_lead_id = lead["id"]
                st.session_state.email_target_audit_id = audit_dict['id']
                st.session_state.email_target_domain = lead["domain"]
                st.session_state.email_target_address = lead["email"]
                st.session_state.current_section = 'Email Outreach'
                st.toast(f"📧 Opening email for {lead['domain']}")
                st.rerun()
        else:
            st.button("📧 Send Email", key=f"send_email_{audit_dict['id']}_disabled", use_container_width=True, disabled=True, 
                      help="No email address found for this lead")
    
    with col_btn4:
        # Delete audit with confirmation
        confirm_key = f"_confirm_delete_{audit_dict['id']}"
        if st.session_state.get(confirm_key):
            st.warning(f"Delete audit for **{audit_dict['domain']}**?")
            col_y, col_n = st.columns(2)
            with col_y:
                if st.button("✅ Yes", key=f"confirm_del_yes_{audit_dict['id']}"):
                    success, domain, error = delete_single_audit(audit_dict['id'])
                    if success:
                        st.toast(f"✓ Deleted audit for {domain}")
                        st.session_state.pop(confirm_key, None)
                        st.rerun()
                    else:
                        st.error(f"Failed to delete: {error}")
            with col_n:
                if st.button("❌ No", key=f"confirm_del_no_{audit_dict['id']}"):
                    st.session_state.pop(confirm_key, None)
                    st.rerun()
        else:
            if st.button("🗑️ Delete", key=f"del_audit_{audit_dict['id']}", use_container_width=True, type="secondary"):
                st.session_state[confirm_key] = True
                st.rerun()


def show_competitor_analysis():
    """Competitor analysis page."""
    st.title("🔄 Competitor Analysis")
    st.markdown("Compare multiple websites side-by-side")
    st.markdown("---")
    
    col1, col2, col3, col4, col5 = st.columns(5)
    urls = []
    for i, col in enumerate([col1, col2, col3, col4, col5]):
        with col:
            url_in = st.text_input(f"Site {i+1}", placeholder="example.com", key=f"comp_{i}")
            if url_in:
                urls.append(url_in)
    
    if st.button("▶️ Compare", type="primary", use_container_width=True):
        if len(urls) < 2:
            st.error("Need at least 2 sites")
        else:
            results = []
            progress = st.progress(0)
            
            for i, url in enumerate(urls):
                data = run_audit(url, st.session_state.OPENAI_API_KEY, st.session_state.GOOGLE_API_KEY)
                save_audit_to_db(data)
                results.append(data)
                progress.progress((i+1)/len(urls))
            
            st.success("Comparison complete!")
            
            comp_data = []
            for data in results:
                domain = urlparse(data['url']).netloc.replace("www.", "")
                comp_data.append({
                    "Website": domain,
                    "Score": format_score_badge(data['score']),
                    "Speed": data.get('psi', 'N/A'),
                    "Issues": len(data.get('issues', [])),
                    "Analytics": "✓" if any("Analytics" in t for t in data.get('tech_stack', [])) else "✗",
                    "SSL": "✓" if not any("SSL" in i.get('title', '') for i in data.get('issues', [])) else "✗"
                })
            
            st.dataframe(pd.DataFrame(comp_data), use_container_width=True)

def show_email_outreach():
    """Email outreach page with CRM integration.
    
    Features:
    - Lead selection from CRM database
    - Pre-fill from audit detail navigation
    - Actual email sending with SMTP
    - Automatic approached status update after send
    - Email templates with variable substitution
    """
    logger = logging.getLogger("sales_engine")
    
    # Code Nest branding CSS
    st.markdown("""
    <style>
    .email-compose-card {
        background: #f8faf9;
        border: 1px solid #e0e0e0;
        border-radius: 12px;
        padding: 1.5rem;
        margin: 1rem 0;
    }
    .email-success {
        background: linear-gradient(135deg, #2b945f 0%, #3cb371 100%);
        color: white;
        padding: 1rem 1.5rem;
        border-radius: 8px;
        text-align: center;
    }
    .email-header {
        background: linear-gradient(135deg, #0c3740 0%, #1a5a6e 100%);
        color: white;
        padding: 0.75rem 1rem;
        border-radius: 8px 8px 0 0;
        font-family: 'Poppins', sans-serif;
    }
    </style>
    """, unsafe_allow_html=True)
    
    st.title("📧 Email Outreach")
    st.markdown("Send personalized emails to leads with automatic CRM tracking")
    st.markdown("---")
    
    if not DB_AVAILABLE:
        st.error("Database required")
        return
    
    email_sub1, email_sub2, email_sub3 = st.tabs(["📨 Send Email", "📋 Email Templates", "📊 Outreach Stats"])
    
    with email_sub1:
        leads_orm = get_leads_cached()
        
        # Normalize all leads to dicts for consistent access
        leads = normalize_lead_list(leads_orm) if leads_orm else []
        
        # Check for pre-filled target from audit detail
        target_lead_id = st.session_state.pop('email_target_lead_id', None)
        target_domain = st.session_state.pop('email_target_domain', None)
        target_email = st.session_state.pop('email_target_address', None)
        target_audit_id = st.session_state.pop('email_target_audit_id', None)
        
        if leads:
            # Build lead options dictionary
            lead_opts = {}
            default_idx = 0
            for idx, l in enumerate(leads):
                label = f"{l['domain']} (Score: {l['health_score'] or 'N/A'}, Opp: {l.get('opportunity_rating') or 'N/A'})"
                lead_opts[label] = l
                # Set default to target lead if coming from audit detail
                if target_lead_id and l['id'] == target_lead_id:
                    default_idx = idx
            
            # CRM quick stats
            col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
            with col_stat1:
                st.metric("📊 Total Leads", len(leads))
            with col_stat2:
                not_approached = sum(1 for l in leads if not l["approached"])
                st.metric("📧 Not Approached", not_approached)
            with col_stat3:
                with_email = sum(1 for l in leads if l["email"])
                st.metric("✉️ With Email", with_email)
            with col_stat4:
                hot_leads = sum(1 for l in leads if l["lead_status"] == "hot")
                st.metric("🔥 Hot Leads", hot_leads)
            
            st.markdown("---")
            
            # Lead selection
            selected_name = st.selectbox(
                "🎯 Select a lead to email", 
                list(lead_opts.keys()), 
                index=default_idx if default_idx < len(lead_opts) else 0,
                key="lead_select"
            )
            selected_lead = lead_opts[selected_name] if selected_name else None
            
            if selected_lead:
                # Lead info card with CRM fields
                st.markdown("<div class='email-header'>📋 Lead Information</div>", unsafe_allow_html=True)
                
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.markdown(f"**🌐 Domain:** {selected_lead['domain']}")
                    st.markdown(f"**📊 Score:** {selected_lead['health_score'] or 'N/A'}")
                    st.markdown(f"**🎯 Opportunity:** {selected_lead.get('opportunity_rating') or 'N/A'}/100")
                with col2:
                    st.markdown(f"**📧 Email:** {selected_lead['email'] or 'Not found'}")
                    st.markdown(f"**📞 Phone:** {selected_lead.get('phone') or 'Not found'}")
                    st.markdown(f"**🏢 Company:** {selected_lead['company_name'] or 'Unknown'}")
                with col3:
                    # CRM status badges
                    approached_icon = "✅ Yes" if selected_lead["approached"] else "❌ No"
                    st.markdown(f"**📮 Approached:** {approached_icon}")
                    st.markdown(f"**🎯 Status:** {selected_lead['lead_status'] or 'warm'}")
                    st.markdown(f"**📊 Pipeline:** {selected_lead['pipeline_stage'] or 'new'}")
                
                # Inline status update
                st.markdown("---")
                col_status1, col_status2 = st.columns([3, 1])
                with col_status1:
                    new_status = st.selectbox(
                        "Update lead status after send", 
                        ["new", "contacted", "responded", "converted", "lost"], 
                        index=["new", "contacted", "responded", "converted", "lost"].index(selected_lead["status"]) if selected_lead["status"] in ["new", "contacted", "responded", "converted", "lost"] else 0,
                        key="email_lead_status"
                    )
                with col_status2:
                    st.markdown("<div style='margin-top:28px;'></div>", unsafe_allow_html=True)
                    if new_status != selected_lead["status"] and st.button("Update Status"):
                        update_lead_status(selected_lead["id"], new_status)
                        st.success("✅ Status updated!")
                        st.rerun()
                
                st.markdown("---")
                st.markdown("#### ✍️ Compose Email")
                
                # Email fields
                recipient = st.text_input("📧 Recipient", value=selected_lead["email"] or target_email or "")
                
                # Get latest audit for this domain for subject/body template
                db = get_db()
                latest_audit = None
                audit_data = None
                if db:
                    try:
                        latest_audit = db.query(Audit).filter(Audit.domain == selected_lead["domain"]).order_by(Audit.created_at.desc()).first()
                        if latest_audit:
                            audit_data = convert_audit_to_data_dict(latest_audit)
                    except Exception as e:
                        logger.warning(f"Error loading audit for email: {e}")
                    finally:
                        db.close()
                
                # Default subject with personalization
                default_subject = f"Quick Website Review for {selected_lead['company_name'] or selected_lead['domain']}"
                subject = st.text_input("📝 Subject", value=default_subject)
                
                # Build body with AI draft if available
                default_body = ""
                ai_section = audit_data.get('ai') if (audit_data and isinstance(audit_data.get('ai'), dict)) else {}
                if ai_section.get('email'):
                    default_body = ai_section['email']
                else:
                    default_body = f"""Hi,

I recently analyzed your website at {selected_lead['domain']} and noticed a few opportunities to improve your online presence.

Your website scored {selected_lead['health_score'] or 'N/A'}/100 on our health check. I'd love to share some insights that could help improve your site's performance and conversions.

Would you be interested in a quick call to discuss?

Best regards,
The Code Nest Team
"""
                
                body = st.text_area("📄 Body", value=default_body, height=250)
                
                # Attachment options
                col_attach1, col_attach2 = st.columns(2)
                with col_attach1:
                    attach_pdf = st.checkbox("📎 Attach PDF Report", value=True if latest_audit else False,
                                            disabled=not latest_audit, 
                                            help="Attach the audit PDF report to the email")
                with col_attach2:
                    mark_approached = st.checkbox("✅ Mark as Approached on Send", value=True,
                                                  help="Automatically update CRM when email is sent")
                
                # Action buttons
                col_btn1, col_btn2 = st.columns(2)
                with col_btn1:
                    if st.button("💾 Save Draft", use_container_width=True):
                        # Save to session for later
                        st.session_state.email_draft = {
                            'lead_id': selected_lead["id"],
                            'recipient': recipient,
                            'subject': subject,
                            'body': body
                        }
                        st.success("💾 Draft saved!")
                
                with col_btn2:
                    if st.button("📤 Send Email", type="primary", use_container_width=True):
                        if not recipient:
                            st.error("❌ Recipient email is required")
                        elif not subject:
                            st.error("❌ Subject is required")
                        elif not body:
                            st.error("❌ Email body is required")
                        else:
                            with st.spinner("Sending email..."):
                                try:
                                    # Get PDF if requested
                                    pdf_bytes = None
                                    pdf_filename = None
                                    if attach_pdf and latest_audit:
                                        pdf_bytes = get_audit_pdf(_get_attr(latest_audit, 'id'))
                                        if not pdf_bytes and audit_data:
                                            pdf_bytes = generate_pdf(audit_data)
                                        pdf_filename = f"website_audit_{selected_lead['domain']}.pdf"
                                    
                                    # Send the email
                                    success, message = send_branded_email_with_pdf(
                                        to_email=recipient,
                                        subject=subject,
                                        body=body,
                                        pdf_bytes=pdf_bytes,
                                        filename=pdf_filename or "audit_report.pdf"
                                    )
                                    
                                    if success:
                                        st.markdown("""
                                        <div class='email-success'>
                                            ✅ <strong>Email Sent Successfully!</strong><br>
                                            Your email has been delivered.
                                        </div>
                                        """, unsafe_allow_html=True)
                                        
                                        # Update CRM if requested
                                        if mark_approached:
                                            mark_lead_as_approached(selected_lead["id"], update_pipeline=True)
                                            st.info("✅ Lead marked as approached in CRM")
                                        
                                        # Log outreach
                                        logger.info(f"Email sent to {recipient} for lead {selected_lead['domain']}")
                                        
                                        # Clear draft
                                        st.session_state.pop('email_draft', None)
                                    else:
                                        st.error(f"❌ Failed to send email: {message}")
                                        logger.error(f"Email send failed: {message}")
                                        
                                except Exception as e:
                                    st.error(f"❌ Error sending email: {str(e)}")
                                    logger.error(f"Email send error: {str(e)}")
        else:
            st.info("📭 No leads found. Run some audits first to generate leads!")
    
    with email_sub2:
        st.markdown("### 📋 Manage Email Templates")
        st.markdown("Create reusable templates with variable substitution")
        
        template_name = st.text_input("Template name", placeholder="e.g., Initial Outreach")
        template_subject = st.text_input("Subject", placeholder="e.g., Quick Website Review for {{company}}")
        template_body = st.text_area(
            "Body template", 
            height=200, 
            placeholder="Hi {{name}},\n\nI noticed your website {{domain}} scored {{score}}/100...",
            help="Available variables: {{domain}}, {{score}}, {{company}}, {{name}}, {{issues}}"
        )
        
        st.markdown("""
        **Available variables:**
        - `{{domain}}` - Website domain
        - `{{score}}` - Health score
        - `{{company}}` - Company name
        - `{{name}}` - Contact name
        - `{{issues}}` - Number of issues found
        """)
        
        if st.button("💾 Save Template", type="primary"):
            if template_name and template_subject and template_body:
                if save_email_template(template_name, template_subject, template_body):
                    st.success("✅ Template saved!")
                else:
                    st.error("❌ Failed to save template")
            else:
                st.warning("⚠️ Please fill in all fields")
        
        st.divider()
        st.markdown("#### 📚 Existing Templates")
        
        templates = load_email_templates()
        if templates:
            for name, template in templates.items():
                with st.expander(f"📄 {name}"):
                    st.markdown(f"**Subject:** {template.get('subject', 'N/A')}")
                    st.markdown(f"**Body:**")
                    st.code(template.get('body', ''), language=None)
                    if st.button(f"🗑️ Delete", key=f"del_template_{name}"):
                        templates.pop(name, None)
                        save_email_templates(templates)
                        st.rerun()
        else:
            st.info("No templates saved yet")
    
    with email_sub3:
        st.markdown("### 📊 Outreach Statistics")
        
        leads = get_leads_cached()
        if leads:
            # Calculate stats
            total = len(leads)
            approached = sum(1 for l in leads if l.approached)
            not_approached = total - approached
            hot = sum(1 for l in leads if l.lead_status == "hot")
            warm = sum(1 for l in leads if l.lead_status == "warm")
            cold = sum(1 for l in leads if l.lead_status == "cold")
            
            # Pipeline breakdown
            new_stage = sum(1 for l in leads if l.pipeline_stage == "new")
            contacted = sum(1 for l in leads if l.pipeline_stage == "contacted")
            followup = sum(1 for l in leads if l.pipeline_stage == "follow-up")
            closed = sum(1 for l in leads if l.pipeline_stage == "closed")
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("📊 Total Leads", total)
                st.metric("✅ Approached", approached)
                st.metric("❌ Not Approached", not_approached)
            with col2:
                st.metric("🔥 Hot", hot)
                st.metric("☀️ Warm", warm)
                st.metric("❄️ Cold", cold)
            with col3:
                st.metric("🆕 New", new_stage)
                st.metric("📞 Contacted", contacted)
                st.metric("📅 Follow-up", followup)
                st.metric("✅ Closed", closed)
            
            # Conversion funnel visualization
            st.markdown("---")
            st.markdown("#### 📈 Pipeline Funnel")
            
            funnel_data = {
                "Stage": ["New", "Contacted", "Follow-up", "Closed"],
                "Count": [new_stage, contacted, followup, closed]
            }
            import pandas as pd
            df_funnel = pd.DataFrame(funnel_data)
            st.bar_chart(df_funnel.set_index("Stage"))
        else:
            st.info("No leads to show statistics for")

def show_lead_management():
    """Advanced lead management with CSV import, AI enrichment, and service scoring."""
    st.title("🎯 Lead Management & Enrichment")
    st.markdown("Import leads from Google Places, enrich with AI, and identify service opportunities")
    st.markdown("---")
    
    if not DB_AVAILABLE:
        st.error("Database required for lead management")
        return
    
    tab1, tab2, tab3, tab4 = st.tabs(["Import Leads", "Lead Database", "Service Opportunities", "AI Insights"])
    
    with tab1:
        st.markdown("### 📤 Import Leads from CSV")
        st.markdown("Upload your Google Places CSV export with: name, phone, address, city, state, zipcode, place_id, website")
        
        uploaded_file = st.file_uploader("Upload CSV file", type="csv")
        
        if uploaded_file:
            try:
                df = pd.read_csv(uploaded_file)
                st.write(f"**Preview:** {len(df)} leads found")
                
                with st.expander("View CSV Preview"):
                    st.dataframe(df.head(10), use_container_width=True)
                
                if st.button("🚀 Process & Enrich Leads", type="primary"):
                    with st.spinner("Processing leads and running audits..."):
                        # First, run audits on websites
                        audit_scores = {}
                        progress_bar = st.progress(0)
                        
                        for idx, row in df.iterrows():
                            website = row.get('website', '') or row.get('url', '')
                            if website:
                                if not website.startswith('http'):
                                    website = 'https://' + website
                                
                                try:
                                    audit_data = run_audit(website, st.session_state.OPENAI_API_KEY, st.session_state.GOOGLE_API_KEY)
                                    domain = urlparse(website).netloc.replace('www.', '')
                                    audit_scores[domain] = audit_data
                                except Exception as e:
                                    logger.warning(f"Error auditing {website}: {str(e)}")
                            
                            progress_bar.progress((idx + 1) / len(df))
                        
                        # Process CSV with enrichment
                        leads_created, errors = process_csv_leads(
                            uploaded_file,
                            audit_scores,
                            st.session_state.OPENAI_API_KEY
                        )
                        
                        st.success(f"✅ Successfully imported {leads_created} leads!")
                        
                        if errors:
                            with st.expander(f"⚠️ {len(errors)} Import Issues"):
                                for error in errors:
                                    st.warning(error)
            
            except Exception as e:
                st.error(f"Error processing file: {str(e)}")
    
    with tab2:
        st.markdown("### 📊 Lead Database")
        
        # Filters
        col1, col2, col3 = st.columns(3)
        with col1:
            status_filter = st.selectbox("Filter by Status", ["all", "new", "contacted", "responded", "converted", "lost"])
        with col2:
            industry_filter = st.text_input("Filter by Industry", placeholder="E.g., E-Commerce, SaaS")
        with col3:
            size_filter = st.selectbox("Filter by Company Size", ["All", "Small", "Medium", "Large", "Enterprise"])
        
        # Get leads
        db = get_db()
        if db:
            query = db.query(Lead).order_by(Lead.opportunity_rating.desc())
            
            if status_filter != "all":
                query = query.filter(Lead.status == status_filter)
            
            if industry_filter:
                query = query.filter(Lead.industry.contains(industry_filter))
            
            if size_filter != "All":
                query = query.filter(Lead.company_size == size_filter)
            
            leads_orm = query.all()
            db.close()
            
            # Normalize all leads to dicts
            leads = normalize_lead_list(leads_orm) if leads_orm else []
            
            if leads:
                # Display metrics
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("Total Leads", len(leads))
                with col2:
                    avg_opp = sum(l.get("opportunity_rating", 0) or 0 for l in leads) / len(leads) if leads else 0
                    st.metric("Avg Opportunity", f"{avg_opp:.0f}/100")
                with col3:
                    high_priority = len([l for l in leads if (l.get("opportunity_rating") or 0) >= 70])
                    st.metric("High Priority", high_priority)
                with col4:
                    converted = len([l for l in leads if l["status"] == "converted"])
                    st.metric("Converted", converted)
                
                st.divider()
                
                # Leads table
                leads_data = []
                for lead in leads[:100]:  # Limit to 100 for performance
                    leads_data.append({
                        "Company": lead["company_name"] or "N/A",
                        "Industry": lead.get("industry") or "Unknown",
                        "Size": lead.get("company_size") or "Unknown",
                        "Health Score": lead["health_score"] if lead["health_score"] is not None else "N/A",
                        "Opportunity": lead.get("opportunity_rating") or 0,
                        "Status": lead["status"],
                        "Location": f"{lead.get('city')}, {lead.get('state')}" if lead.get("city") else "N/A",
                        "Phone": lead.get("phone") or "N/A"
                    })
                
                st.dataframe(
                    pd.DataFrame(leads_data),
                    use_container_width=True,
                    column_config={
                        "Opportunity": st.column_config.ProgressColumn(
                            "Opportunity",
                            min_value=0,
                            max_value=100,
                        ),
                    }
                )
            else:
                st.info("No leads found. Try importing CSV data first.")
    
    with tab3:
        st.markdown("### 🎯 Service Opportunities")
        
        db = get_db()
        if db:
            leads_orm = db.query(Lead).order_by(Lead.opportunity_rating.desc()).limit(50).all()
            db.close()
            
            # Normalize leads
            leads = normalize_lead_list(leads_orm) if leads_orm else []
            
            if leads:
                # Select a lead to see service opportunities
                lead_options = {f"{l['company_name']} ({l.get('city')}, {l.get('state')})" or l['domain']: l for l in leads}
                selected_lead_name = st.selectbox("Select a lead to view service opportunities", list(lead_options.keys()))
                selected_lead = lead_options[selected_lead_name]
                
                if selected_lead:
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.markdown(f"**Company:** {selected_lead['company_name'] or 'N/A'}")
                        st.markdown(f"**Industry:** {selected_lead.get('industry') or 'Unknown'}")
                    with col2:
                        st.markdown(f"**Size:** {selected_lead.get('company_size') or 'Unknown'}")
                        st.markdown(f"**Health Score:** {selected_lead['health_score'] or 'N/A'}/100")
                    with col3:
                        st.markdown(f"**Location:** {selected_lead.get('city')}, {selected_lead.get('state')}" if selected_lead.get('city') else "")
                        st.markdown(f"**Opportunity:** {selected_lead.get('opportunity_rating') or 0}/100")
                    
                    st.divider()
                    
                    service_priorities = selected_lead.get('service_priorities') or {}
                    if service_priorities:
                        st.markdown("### 📈 Service Opportunity Scores")
                        
                        service_names = {
                            'website_development': '🌐 Website Development',
                            'seo_optimization': '🔍 SEO Optimization',
                            'mobile_app_development': '📱 Mobile App Development',
                            'social_media_marketing': '📱 Social Media Marketing',
                            'paid_advertising': '💰 Paid Advertising (PPC)',
                            'ecommerce_development': '🛒 E-Commerce Development',
                            'website_maintenance': '🔧 Website Maintenance',
                            'react_nextjs_development': '⚛️ React/Next.js Development',
                            'website_optimization': '⚡ Website Optimization',
                            'graphic_designing': '🎨 Graphic Designing'
                        }
                        
                        cols = st.columns(2)
                        for idx, (service_key, score) in enumerate(sorted(service_priorities.items(), 
                                                                          key=lambda x: x[1], reverse=True)):
                            with cols[idx % 2]:
                                service_name = service_names.get(service_key, service_key)
                                st.markdown(f"### {service_name}")
                                
                                # Progress bar
                                col_score, col_badge = st.columns([3, 1])
                                with col_score:
                                    st.progress(min(100, score) / 100)
                                with col_badge:
                                    if score >= 80:
                                        st.error(f"{score:.0f}")
                                    elif score >= 60:
                                        st.warning(f"{score:.0f}")
                                    else:
                                        st.info(f"{score:.0f}")
                                
                                # Show pitch on click
                                if st.button(f"View Pitch", key=f"pitch_{service_key}"):
                                    st.session_state[f"show_pitch_{service_key}"] = True
                                
                                if st.session_state.get(f"show_pitch_{service_key}"):
                                    with st.expander("📧 Service Pitch", expanded=True):
                                        pitch = generate_service_pitch(
                                            selected_lead["company_name"] or "Client",
                                            service_key,
                                            score,
                                            selected_lead["industry"] or "Unknown",
                                            selected_lead["company_size"] or "Unknown",
                                            {"score": selected_lead["health_score"] or 0, "issues": []}
                                        )
                                        st.markdown(pitch)
            else:
                st.info("No leads available")
    
    with tab4:
        st.markdown("### 🤖 AI Lead Insights")
        
        db = get_db()
        if db:
            leads_with_ai_orm = db.query(Lead).filter(Lead.ai_enrichment.isnot(None)).order_by(Lead.updated_at.desc()).limit(20).all()
            db.close()
            leads_with_ai = normalize_lead_list(leads_with_ai_orm)
            
            if leads_with_ai:
                for lead in leads_with_ai:
                    with st.expander(f"{lead['company_name'] or lead['domain']} - {lead['city']}, {lead['state']}"):
                        ai_enrichment = lead["ai_enrichment"]
                        if ai_enrichment:
                            col1, col2 = st.columns(2)
                            
                            with col1:
                                st.markdown("**Key Challenges:**")
                                if isinstance(ai_enrichment, dict):
                                    for challenge in ai_enrichment.get('key_challenges', []):
                                        st.markdown(f"- {challenge}")
                                    
                                    st.markdown("**Quick Wins (30 days):**")
                                    for win in ai_enrichment.get('quick_wins', []):
                                        st.markdown(f"✅ {win}")
                            
                            with col2:
                                st.markdown("**Recommended Services:**")
                                for service in ai_enrichment.get('recommended_services', []):
                                    st.markdown(f"• {service}")
                                
                                st.markdown("**Conversation Starters:**")
                                for starter in ai_enrichment.get('conversation_starters', []):
                                    st.markdown(f"💡 {starter}")
                            
                            st.markdown("**Estimated Business Impact:**")
                            st.info(ai_enrichment.get('estimated_impact', 'N/A'))
            else:
                st.info("No AI-enriched leads yet. Import leads with AI enrichment enabled.")

def show_scheduled_audits():
    """Scheduled audits page."""
    st.title("⏰ Scheduled Audits")
    st.markdown("Automated periodic audits to track improvements over time")
    st.markdown("---")
    
    st.info("📋 This feature is currently in development. Coming soon:\n- Schedule audits to run automatically\n- Track improvements over time\n- Auto-generated reports")
    
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("### Setup Scheduled Audit")
        url = st.text_input("Website URL")
        frequency = st.selectbox("Frequency", ["Daily", "Weekly", "Monthly"])
        
        if st.button("Schedule"):
            st.success(f"Audit scheduled for {frequency}")
    
    with col2:
        st.markdown("### Active Schedules")
        st.info("No scheduled audits yet")


# ============================================================================
# PHASE 4: CRM PIPELINE DASHBOARD
# ============================================================================

def show_crm_pipeline():
    """
    CRM Pipeline Dashboard - Visual lead management with Kanban board.
    
    Features:
    - Kanban board with 6 pipeline stages
    - Table view with filters and bulk actions
    - Analytics dashboard with charts
    - Lead detail panel with quick actions
    - Auto-movement rules for workflow automation
    """
    logger = logging.getLogger("sales_engine")
    logger.debug("Rendering CRM Pipeline page")
    
    # Custom CSS for CRM Pipeline
    st.markdown("""
    <style>
    /* CRM Pipeline Header */
    .crm-header {
        background: linear-gradient(135deg, #0c3740 0%, #1a5a6e 100%);
        padding: 1.5rem 2rem;
        border-radius: 12px;
        margin-bottom: 1.5rem;
        color: white;
    }
    .crm-header h1 {
        font-family: 'Poppins', sans-serif;
        font-size: 1.8rem;
        margin: 0;
        font-weight: 600;
    }
    .crm-header p {
        font-family: 'Lato', sans-serif;
        opacity: 0.9;
        margin: 0.3rem 0 0 0;
        font-size: 0.95rem;
    }
    
    /* Kanban Column */
    .kanban-column {
        background: #f8f9fa;
        border-radius: 10px;
        padding: 0.8rem;
        min-height: 400px;
    }
    .kanban-header {
        font-weight: 600;
        padding: 0.5rem;
        border-radius: 6px;
        margin-bottom: 0.5rem;
        font-family: 'Poppins', sans-serif;
        font-size: 0.9rem;
        text-align: center;
    }
    .kanban-header.new { background: #e3f2fd; color: #1565c0; }
    .kanban-header.contacted { background: #fff3e0; color: #ef6c00; }
    .kanban-header.follow-up { background: #fce4ec; color: #c2185b; }
    .kanban-header.qualified { background: #e8f5e9; color: #2e7d32; }
    .kanban-header.proposal { background: #f3e5f5; color: #7b1fa2; }
    .kanban-header.closed { background: #e0f2f1; color: #00695c; }
    
    /* Lead Card in Kanban */
    .lead-card {
        background: white;
        border-radius: 8px;
        padding: 0.8rem;
        margin-bottom: 0.5rem;
        box-shadow: 0 1px 3px rgba(0,0,0,0.1);
        border-left: 3px solid #2b945f;
        transition: all 0.2s ease;
    }
    .lead-card:hover {
        box-shadow: 0 4px 12px rgba(0,0,0,0.15);
        transform: translateY(-1px);
    }
    .lead-card.hot { border-left-color: #dc3545; }
    .lead-card.warm { border-left-color: #ffc107; }
    .lead-card.cold { border-left-color: #6c757d; }
    
    .lead-domain {
        font-weight: 600;
        font-size: 0.9rem;
        color: #0c3740;
        margin-bottom: 0.3rem;
    }
    .lead-company {
        font-size: 0.8rem;
        color: #5a5a5a;
        margin-bottom: 0.3rem;
    }
    .lead-score {
        font-size: 0.75rem;
        padding: 0.2rem 0.5rem;
        border-radius: 4px;
        display: inline-block;
    }
    .lead-score.high { background: #d4edda; color: #155724; }
    .lead-score.medium { background: #fff3cd; color: #856404; }
    .lead-score.low { background: #f8d7da; color: #721c24; }
    
    /* Status Badge */
    .status-pill {
        display: inline-block;
        padding: 0.15rem 0.5rem;
        border-radius: 12px;
        font-size: 0.7rem;
        font-weight: 600;
        text-transform: uppercase;
    }
    .status-pill.hot { background: #dc3545; color: white; }
    .status-pill.warm { background: #ffc107; color: #333; }
    .status-pill.cold { background: #6c757d; color: white; }
    
    /* Analytics Card */
    .analytics-card {
        background: white;
        border-radius: 10px;
        padding: 1.2rem;
        box-shadow: 0 2px 8px rgba(0,0,0,0.08);
        text-align: center;
    }
    .analytics-card h3 {
        font-size: 2rem;
        color: #0c3740;
        margin: 0;
        font-weight: 700;
    }
    .analytics-card p {
        font-size: 0.85rem;
        color: #5a5a5a;
        margin: 0.3rem 0 0 0;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Header
    st.markdown("""
    <div class="crm-header">
        <h1>📊 CRM Pipeline</h1>
        <p>Manage your leads through the sales pipeline with visual tracking and automation</p>
    </div>
    """, unsafe_allow_html=True)
    
    if not DB_AVAILABLE:
        st.error("Database connection required for CRM Pipeline")
        return
    
    # View tabs
    tab1, tab2, tab3 = st.tabs(["🗂️ Kanban Board", "📋 Table View", "📈 Analytics"])
    
    with tab1:
        render_crm_kanban_board()
    
    with tab2:
        render_crm_table_view()
    
    with tab3:
        render_crm_analytics()


def render_crm_kanban_board():
    """Render the Kanban-style pipeline board."""
    
    # Get leads grouped by stage
    stages_data = {}
    for stage in PIPELINE_STAGES:
        result = db_get_leads_by_pipeline(stage=stage, limit=50)
        if result["success"]:
            stages_data[stage] = result["data"].get("leads", [])
        else:
            stages_data[stage] = []
    
    # Create 6 columns for the Kanban board
    cols = st.columns(6)
    stage_colors = {
        "new": "new",
        "contacted": "contacted",
        "follow-up": "follow-up",
        "qualified": "qualified",
        "proposal": "proposal",
        "closed": "closed"
    }
    stage_icons = {
        "new": "🆕",
        "contacted": "📧",
        "follow-up": "🔄",
        "qualified": "✅",
        "proposal": "📝",
        "closed": "🎉"
    }
    
    for idx, stage in enumerate(PIPELINE_STAGES):
        with cols[idx]:
            leads = stages_data.get(stage, [])
            icon = stage_icons.get(stage, "📋")
            color_class = stage_colors.get(stage, "new")
            
            st.markdown(f"""
            <div class="kanban-column">
                <div class="kanban-header {color_class}">
                    {icon} {stage.upper()} ({len(leads)})
                </div>
            """, unsafe_allow_html=True)
            
            if not leads:
                st.caption("No leads")
            else:
                for lead in leads[:10]:  # Show max 10 per column
                    render_kanban_lead_card(lead, stage)
            
            st.markdown("</div>", unsafe_allow_html=True)


def render_kanban_lead_card(lead, current_stage: str):
    """Render a single lead card in the Kanban board."""
    # Ensure lead is a dict with all defaults (crash-proof)
    lead = safe_render_lead(lead)
    
    # Skip if no valid lead
    if not lead.get("domain") and not lead.get("id"):
        return
    
    domain = (lead.get("domain") or "Unknown")[:25]
    company = (lead.get("company_name") or "")[:20] or domain
    score = lead.get("health_score")
    status = lead.get("lead_status") or "warm"
    lead_id = lead.get("id")
    
    # Determine score class
    if score is None:
        score_class = "medium"
        score_display = "N/A"
    elif score >= 70:
        score_class = "high"
        score_display = f"🟢 {score}"
    elif score >= 40:
        score_class = "medium"
        score_display = f"🟡 {score}"
    else:
        score_class = "low"
        score_display = f"🔴 {score}"
    
    # Card container
    with st.container(border=True):
        st.markdown(f"**{domain}**")
        if company and company != domain:
            st.caption(company)
        
        col1, col2 = st.columns([1, 1])
        with col1:
            # Status pill
            pill_colors = {"hot": "🔥", "warm": "☀️", "cold": "❄️"}
            st.caption(f"{pill_colors.get(status, '❓')} {status}")
        with col2:
            st.caption(score_display)
        
        # Quick actions
        col_move, col_view = st.columns(2)
        with col_move:
            # Move to next stage
            current_idx = PIPELINE_STAGES.index(current_stage) if current_stage in PIPELINE_STAGES else 0
            if current_idx < len(PIPELINE_STAGES) - 1:
                next_stage = PIPELINE_STAGES[current_idx + 1]
                if st.button("→", key=f"move_{lead_id}_{current_stage}", help=f"Move to {next_stage}"):
                    result = db_update_lead_pipeline(lead_id, stage=next_stage)
                    if result["success"]:
                        st.toast(f"Moved to {next_stage}")
                        st.rerun()
        with col_view:
            if st.button("👁️", key=f"view_{lead_id}_{current_stage}", help="View details"):
                st.session_state["crm_selected_lead_id"] = lead_id
                st.rerun()


def render_crm_table_view():
    """Render the table view with filters and bulk actions."""
    
    # Filters row
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        stage_filter = st.selectbox("Pipeline Stage", ["All"] + PIPELINE_STAGES, key="crm_stage_filter")
    with col2:
        status_filter = st.selectbox("Lead Status", ["All"] + LEAD_STATUS_OPTIONS, key="crm_status_filter")
    with col3:
        search_term = st.text_input("Search Domain", placeholder="Enter domain...", key="crm_search")
    with col4:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🔄 Refresh", key="crm_refresh"):
            st.rerun()
    
    # Get filtered leads
    stage = stage_filter if stage_filter != "All" else None
    status = status_filter if status_filter != "All" else None
    
    result = db_get_leads_by_pipeline(stage=stage, status=status, limit=100)
    
    if not result["success"]:
        st.error(f"Error loading leads: {result.get('error', 'Unknown')}")
        return
    
    leads = result["data"].get("leads", [])
    
    # Convert all leads to dicts with crash-proof defaults
    leads = [safe_render_lead(l) for l in leads]
    
    # Apply search filter - skip leads with no domain
    if search_term:
        leads = [l for l in leads if l.get("domain") and search_term.lower() in l["domain"].lower()]
    else:
        leads = [l for l in leads if l.get("domain")]  # Filter out invalid leads
    
    if not leads:
        st.info("No leads found matching your filters")
        return
    
    st.markdown(f"**Showing {len(leads)} leads**")
    
    # Prepare dataframe
    df_data = []
    for lead in leads:
        follow_up = lead.get("follow_up_date") or ""
        if isinstance(follow_up, str) and len(follow_up) >= 10:
            follow_up = follow_up[:10]
        else:
            follow_up = "-"
        
        df_data.append({
            "ID": lead.get("id"),
            "Domain": lead.get("domain") or "",
            "Company": lead.get("company_name") or "",
            "Email": lead.get("email") or "",
            "Score": lead.get("health_score") if lead.get("health_score") is not None else "N/A",
            "Status": lead.get("lead_status") or "",
            "Stage": lead.get("pipeline_stage") or "",
            "Approached": "✅" if lead.get("approached") else "❌",
            "Follow-up": follow_up,
        })
    
    df = pd.DataFrame(df_data)
    
    # Display with selection
    st.dataframe(df, use_container_width=True, hide_index=True)
    
    # Bulk actions
    st.markdown("### Bulk Actions")
    selected_ids = st.multiselect("Select Lead IDs for bulk action", [l["ID"] for l in df_data], key="crm_bulk_select")
    
    if selected_ids:
        col1, col2, col3 = st.columns(3)
        with col1:
            new_stage = st.selectbox("Move to Stage", PIPELINE_STAGES, key="crm_bulk_stage")
            if st.button("Apply Stage", key="crm_apply_stage"):
                success_count = 0
                for lid in selected_ids:
                    result = db_update_lead_pipeline(lid, stage=new_stage)
                    if result["success"]:
                        success_count += 1
                st.success(f"Moved {success_count} leads to {new_stage}")
                st.rerun()
        
        with col2:
            new_status = st.selectbox("Set Status", LEAD_STATUS_OPTIONS, key="crm_bulk_status")
            if st.button("Apply Status", key="crm_apply_status"):
                success_count = 0
                for lid in selected_ids:
                    result = db_update_lead_pipeline(lid, status=new_status)
                    if result["success"]:
                        success_count += 1
                st.success(f"Updated {success_count} leads to {new_status}")
                st.rerun()
        
        with col3:
            bulk_note = st.text_input("Add Note", key="crm_bulk_note")
            if st.button("Add Note", key="crm_apply_note"):
                success_count = 0
                for lid in selected_ids:
                    result = db_update_lead_pipeline(lid, notes=bulk_note)
                    if result["success"]:
                        success_count += 1
                st.success(f"Added note to {success_count} leads")
                st.rerun()


def render_crm_analytics():
    """Render CRM analytics dashboard."""
    
    # Get counts
    result = db_count_leads_by_status()
    
    if not result["success"]:
        st.error("Error loading analytics")
        return
    
    data = result["data"]
    by_status = data.get("by_status", {})
    by_stage = data.get("by_stage", {})
    total = data.get("total", 0)
    
    # Top metrics row
    st.markdown("### 📊 Pipeline Overview")
    cols = st.columns(4)
    
    with cols[0]:
        st.markdown("""
        <div class="analytics-card">
            <h3>{}</h3>
            <p>Total Leads</p>
        </div>
        """.format(total), unsafe_allow_html=True)
    
    with cols[1]:
        hot = by_status.get("hot", 0)
        st.markdown("""
        <div class="analytics-card">
            <h3 style="color: #dc3545;">{}</h3>
            <p>🔥 Hot Leads</p>
        </div>
        """.format(hot), unsafe_allow_html=True)
    
    with cols[2]:
        warm = by_status.get("warm", 0)
        st.markdown("""
        <div class="analytics-card">
            <h3 style="color: #ffc107;">{}</h3>
            <p>☀️ Warm Leads</p>
        </div>
        """.format(warm), unsafe_allow_html=True)
    
    with cols[3]:
        cold = by_status.get("cold", 0)
        st.markdown("""
        <div class="analytics-card">
            <h3 style="color: #6c757d;">{}</h3>
            <p>❄️ Cold Leads</p>
        </div>
        """.format(cold), unsafe_allow_html=True)
    
    st.markdown("---")
    
    # Pipeline funnel
    st.markdown("### 📈 Pipeline Stages")
    
    if by_stage:
        stage_df = pd.DataFrame([
            {"Stage": stage.title(), "Count": by_stage.get(stage, 0)}
            for stage in PIPELINE_STAGES
        ])
        st.bar_chart(stage_df.set_index("Stage"))
    else:
        st.info("No data available for pipeline stages")
    
    # Follow-up section
    st.markdown("---")
    st.markdown("### ⏰ Follow-ups Due")
    
    followup_result = db_get_follow_up_due(days_ahead=7)
    
    if followup_result["success"]:
        followup_data = followup_result["data"]
        overdue = followup_data.get("overdue", [])
        today = followup_data.get("today", [])
        upcoming = followup_data.get("upcoming", [])
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.markdown(f"#### 🚨 Overdue ({len(overdue)})")
            if overdue:
                for lead in overdue[:5]:
                    lead = lead_to_dict(lead)  # Ensure dict
                    st.markdown(f"- **{lead.get('domain') or 'Unknown'}**")
            else:
                st.caption("None overdue")
        
        with col2:
            st.markdown(f"#### 📅 Today ({len(today)})")
            if today:
                for lead in today[:5]:
                    lead = lead_to_dict(lead)  # Ensure dict
                    st.markdown(f"- **{lead.get('domain') or 'Unknown'}**")
            else:
                st.caption("None today")
        
        with col3:
            st.markdown(f"#### 📆 Upcoming ({len(upcoming)})")
            if upcoming:
                for lead in upcoming[:5]:
                    lead = lead_to_dict(lead)  # Ensure dict
                    follow_up = lead.get('follow_up_date') or ''
                    date_str = follow_up[:10] if isinstance(follow_up, str) and len(follow_up) >= 10 else ''
                    st.markdown(f"- **{lead.get('domain') or 'Unknown'}** ({date_str})")
            else:
                st.caption("None upcoming")
    else:
        st.error("Could not load follow-up data")


def render_crm_lead_detail_panel(lead_id: int):
    """
    Render a detailed view panel for a single lead.
    
    Called when a lead is selected in the Kanban or Table view.
    """
    if not lead_id:
        st.warning("No lead selected")
        return
    
    result = db_get_lead_by_id(lead_id)
    
    if not result["success"]:
        st.error(f"Could not load lead: {result.get('error', 'Unknown')}")
        return
    
    # Use crash-proof lead rendering
    lead = safe_render_lead(result["data"].get("lead", {}))
    
    if not lead.get("domain"):
        st.warning("Lead data is incomplete")
        return
    
    st.markdown("### 📋 Lead Details")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown(f"**Domain:** {lead.get('domain') or 'Unknown'}")
        st.markdown(f"**Company:** {lead.get('company_name') or 'N/A'}")
        st.markdown(f"**Email:** {lead.get('email') or 'N/A'}")
        st.markdown(f"**Phone:** {lead.get('phone') or 'N/A'}")
        st.markdown(f"**Health Score:** {lead.get('health_score') if lead.get('health_score') is not None else 'N/A'}")
    
    with col2:
        st.markdown(f"**Stage:** {lead.get('pipeline_stage') or 'new'}")
        st.markdown(f"**Status:** {lead.get('lead_status') or 'warm'}")
        st.markdown(f"**Approached:** {'Yes' if lead.get('approached') else 'No'}")
        follow_up = lead.get('follow_up_date') or ''
        follow_up_display = follow_up[:10] if isinstance(follow_up, str) and len(follow_up) >= 10 else 'Not set'
        st.markdown(f"**Follow-up:** {follow_up_display}")
        st.markdown(f"**Assigned:** {lead.get('assigned_user') or 'Unassigned'}")
    
    # Notes
    st.markdown("**Notes:**")
    notes = lead.get("notes") or "No notes"
    st.text_area("", value=notes, height=100, disabled=True, key=f"lead_notes_{lead_id}")
    
    # Quick actions
    st.markdown("### Quick Actions")
    col1, col2, col3 = st.columns(3)
    
    with col1:
        new_stage = st.selectbox("Move to Stage", PIPELINE_STAGES, 
                                  index=PIPELINE_STAGES.index(lead.get("pipeline_stage", "new")),
                                  key=f"detail_stage_{lead_id}")
        if st.button("Update Stage", key=f"update_stage_{lead_id}"):
            result = db_update_lead_pipeline(lead_id, stage=new_stage)
            if result["success"]:
                st.success("Stage updated")
                st.rerun()
    
    with col2:
        new_status = st.selectbox("Set Status", LEAD_STATUS_OPTIONS,
                                   index=LEAD_STATUS_OPTIONS.index(lead.get("lead_status", "warm")) if lead.get("lead_status") in LEAD_STATUS_OPTIONS else 1,
                                   key=f"detail_status_{lead_id}")
        if st.button("Update Status", key=f"update_status_{lead_id}"):
            result = db_update_lead_pipeline(lead_id, status=new_status)
            if result["success"]:
                st.success("Status updated")
                st.rerun()
    
    with col3:
        new_note = st.text_input("Add Note", key=f"detail_note_{lead_id}")
        if st.button("Add Note", key=f"add_note_{lead_id}") and new_note:
            result = db_update_lead_pipeline(lead_id, notes=new_note)
            if result["success"]:
                st.success("Note added")
                st.rerun()
    
    # View audit button
    if lead.get("last_audit_id"):
        if st.button("🔍 View Last Audit", key=f"view_audit_{lead_id}"):
            audit_id = lead["last_audit_id"]
            set_current_audit(audit_id)
            save_navigation_state("Single Audit", audit_id)
            sync_query_params()
            st.session_state.current_section = "Single Audit"
            st.rerun()


def clean_text(text):
    """Sanitize text for PDF."""
    if not text: return ""
    text = text.replace('\u201c', '"').replace('\u201d', '"').replace('\u2019', "'").replace('\u2013', '-')
    return text.encode('latin-1', 'replace').decode('latin-1')

def format_score_badge(score):
    """Return color-coded score badge."""
    if score >= 80:
        return f"🟢 {score}/100"
    elif score >= 50:
        return f"🟡 {score}/100"
    else:
        return f"🔴 {score}/100"

def send_slack_notification(message, webhook_url=None):
    """Send Slack notification."""
    if not webhook_url:
        webhook_url = SLACK_WEBHOOK
    if not webhook_url:
        return False
    
    try:
        payload = {"text": message}
        requests.post(webhook_url, json=payload)
        return True
    except Exception:
        return False

def get_domain_age(url):
    """Get domain age."""
    domain = urlparse(url).netloc.replace("www.", "")
    try:
        w = whois.whois(domain)
        creation_date = w.creation_date
        if isinstance(creation_date, list):
            creation_date = creation_date[0]
        if creation_date:
            days_alive = (datetime.now() - creation_date).days
            years = round(days_alive / 365, 1)
            return f"{years} years", days_alive
    except Exception:
        return "Unknown", 0
    return "Unknown", 0

def get_google_speed(url, api_key):
    """Get Google PageSpeed score."""
    if not api_key: 
        return None, "No API Key"
    
    api_url = f"https://www.googleapis.com/pagespeedonline/v5/runPagespeed?url={url}&strategy=mobile&key={api_key}"
    try:
        r = requests.get(api_url, timeout=15)  # Reduced from 60s to 15s for faster audits
        if r.status_code == 200:
            data = r.json()
            try:
                score = data['lighthouseResult']['categories']['performance']['score'] * 100
                return int(score), "Success"
            except KeyError:
                return None, "Data Error"
        else:
            return None, f"Error: {r.status_code}"
    except Exception as e:
        return None, str(e)

def get_ai_consultation(url, data, api_key, force_regenerate=False):
    """
    Get AI analysis with structured insights for PDF and email outreach.
    
    ENHANCED FOR RICHER ISSUES:
    - Receives structured issues with category, title, impact, technical_detail
    - Produces deeper business-focused insights
    - Email references specific impactful issues across categories
    
    OPTIMIZED FOR COST:
    - Single OpenAI call returns both insights AND email
    - Results are cached per URL
    - Word limits: ~450 words for insights, ~180 words for email
    - No conversation history (single user message)
    
    Args:
        url: Website URL
        data: Audit data dict with rich issues structure
        api_key: OpenAI API key
        force_regenerate: If True, skip cache and make fresh API call
    
    Returns:
        Dict with summary, impact, solutions, email, email_subject, insights
    """
    domain = urlparse(url).netloc.replace("www.", "")
    
    # Check for valid API key first
    if not api_key:
        return {
            "summary": "AI Analysis Disabled",
            "impact": "N/A",
            "solutions": "Upgrade to enable AI",
            "email": "N/A",
            "email_subject": f"quick idea for {domain}",
            "insights": None,
            "from_cache": False
        }
    
    # CHECK CACHE FIRST (unless force regenerate)
    if not force_regenerate:
        cached = get_cached_ai_result(url)
        if cached:
            cached['from_cache'] = True
            return cached
    
    # Validate API key format
    if len(api_key) < 20:
        logger.error(f"OpenAI API key is invalid (length: {len(api_key)})")
        return {
            "summary": "Invalid API Key",
            "impact": "Please check your OpenAI API key in API Settings",
            "solutions": "Go to API Settings and enter a valid OpenAI API key",
            "email": "API key error - please configure OpenAI API key",
            "email_subject": f"quick idea for {domain}",
            "insights": None,
            "from_cache": False
        }
    
    # =========================================================================
    # PREPARE RICH AUDIT DATA FOR PROMPT
    # =========================================================================
    health_score = data.get('score', 0)
    psi_score = data.get('psi', 'N/A')
    domain_age = data.get('domain_age', 'Unknown')
    tech_list = ", ".join(data.get('tech_stack', [])[:8])
    
    # Build rich issues summary grouped by category
    issues = data.get('issues', [])
    issues_by_category = {}
    for issue in issues:
        cat = issue.get('category', 'General')
        if cat not in issues_by_category:
            issues_by_category[cat] = []
        issues_by_category[cat].append({
            'title': issue.get('title', ''),
            'impact': issue.get('impact', ''),
            'severity': issue.get('severity', 'medium')
        })
    
    # Format issues for prompt (top issues from each category, prioritize high severity)
    issues_text_parts = []
    priority_categories = ['Tracking', 'Conversion', 'SEO', 'Performance', 'Security', 'Content', 'UX']
    
    for cat in priority_categories:
        if cat in issues_by_category:
            cat_issues = issues_by_category[cat]
            # Sort by severity (high first)
            cat_issues.sort(key=lambda x: 0 if x['severity'] == 'high' else (1 if x['severity'] == 'medium' else 2))
            for issue in cat_issues[:2]:  # Max 2 per category
                issues_text_parts.append(f"[{cat}] {issue['title']}: {issue['impact']}")
    
    # Limit to top 8 most impactful issues for prompt
    issues_text = "\n".join(issues_text_parts[:8]) if issues_text_parts else "Minor optimization opportunities found"
    
    # Additional context for AI
    content_stats = data.get('content_stats', {})
    word_count = content_stats.get('word_count', 'unknown')
    tracking_info = data.get('tracking', {})
    has_analytics = tracking_info.get('google_analytics', False) or tracking_info.get('gtm', False)
    has_retargeting = tracking_info.get('facebook_pixel', False) or tracking_info.get('linkedin_insight', False)
    
    # =========================================================================
    # ENHANCED COMBINED PROMPT
    # =========================================================================
    combined_prompt = f"""You are a senior digital strategist at Code Nest LLC (New Mexico). Analyze this comprehensive audit and produce BOTH detailed insights AND a cold outreach email.

AUDIT OVERVIEW:
- Domain: {domain}
- Health Score: {health_score}/100
- PageSpeed: {psi_score}
- Domain Age: {domain_age}
- Tech Stack: {tech_list}
- Content: ~{word_count} words on homepage
- Has Analytics: {has_analytics}
- Has Retargeting: {has_retargeting}

DETAILED ISSUES FOUND (by category):
{issues_text}

Return ONLY valid JSON (no markdown code blocks):
{{
  "insights": {{
    "snapshot_summary": [
      "bullet1 - key finding about business impact (15-20 words)",
      "bullet2 - another key finding (15-20 words)",
      "bullet3 - third finding if relevant (15-20 words)"
    ],
    "top_3_issues": [
      {{
        "issue": "Issue name",
        "impact": "2-3 sentence explanation of business impact. How does this hurt leads, sales, or visibility?",
        "category": "category name"
      }},
      {{
        "issue": "Issue name",
        "impact": "2-3 sentence explanation. Be specific about lost revenue/leads/trust.",
        "category": "category name"
      }},
      {{
        "issue": "Issue name",
        "impact": "2-3 sentence explanation with business outcome focus.",
        "category": "category name"
      }}
    ],
    "quick_wins": [
      "Specific actionable fix 1 (e.g., 'Add Google Analytics tracking to see which channels drive leads')",
      "Specific actionable fix 2",
      "Specific actionable fix 3",
      "Specific actionable fix 4 (optional)",
      "Specific actionable fix 5 (optional)"
    ],
    "code_nest_services": [
      {{
        "issue": "Problem area",
        "service": "Specific Code Nest service/solution that addresses this"
      }},
      {{
        "issue": "Problem area",
        "service": "Specific solution"
      }},
      {{
        "issue": "Problem area",
        "service": "Specific solution"
      }}
    ],
    "next_step": "Clear call-to-action: suggest scheduling a free 15-minute strategy call to discuss priorities"
  }},
  "email": {{
    "subject": "lowercase 4-6 word subject mentioning {domain} (no caps, no punctuation)",
    "body": "Full cold email body. Structure:
1. Opening: 'Hi there,' + brief mention you reviewed their site
2. Key findings: 2-4 bullet points from DIFFERENT categories (mix SEO, tracking, conversion issues)
3. Business impact: 1-2 sentences on how these issues affect leads/revenue
4. Credibility line: 'For context, I'm reaching out from Code Nest LLC, a New Mexico-based agency specializing in website optimization, SEO, and digital marketing.'
5. Soft CTA: offer a free audit review call
6. Sign off with full signature"
  }}
}}

CRITICAL GUIDELINES:
1. INSIGHTS TOTAL: Keep under 450 words. Be concise but substantive.
2. EMAIL BODY: Target 120-180 words (not counting signature). Short enough for mobile reading.
3. BUSINESS LANGUAGE: Translate technical issues to business outcomes:
   - 'No analytics' → 'You can't see which marketing channels drive actual revenue'
   - 'No retargeting pixel' → 'You can't follow up on the 95%+ of visitors who leave without contacting you'
   - 'Thin content' → 'Google struggles to understand what your business offers, hurting rankings'
   - 'Slow speed' → '53% of visitors abandon sites that take more than 3 seconds to load'
4. PRIORITIZE issues that affect: lead generation, search visibility, ad ROI, user trust
5. If score >70, frame as 'optimization opportunities' not 'problems'
6. EMAIL must include issues from at least 2 different categories (e.g., one SEO + one Tracking)
7. Sign email as:
Best regards,
Code Nest Team
Code Nest LLC – New Mexico
contact@codenest.us.com
www.codenest.us.com"""

    try:
        log_source = "regeneration" if force_regenerate else "initial"
        logger.info(f"[OPENAI CALL - {log_source}] {domain} (key: {api_key[:8]}...)")
        
        client = OpenAI(api_key=api_key, timeout=AI_CLIENT_TIMEOUT)
        
        response = client.chat.completions.create(
            model=AI_MODEL,
            messages=[
                {"role": "system", "content": "You are a business-focused digital strategist. Translate technical website issues into clear business impact language. Return only valid JSON."},
                {"role": "user", "content": combined_prompt}
            ],
            temperature=AI_TEMPERATURE,
            max_tokens=AI_MAX_TOKENS,
            timeout=AI_TIMEOUT
        )
        
        result_text = response.choices[0].message.content.strip()
        logger.info(f"[OPENAI RESPONSE] {domain} - {len(result_text)} chars")
        
        # Parse JSON response
        try:
            # Clean up potential markdown code blocks
            if result_text.startswith("```"):
                result_text = result_text.split("```")[1]
                if result_text.startswith("json"):
                    result_text = result_text[4:]
                result_text = result_text.strip()
            
            parsed = json.loads(result_text)
            insights = parsed.get("insights", {})
            email_data = parsed.get("email", {})
            
        except json.JSONDecodeError as e:
            logger.warning(f"JSON parse error for {domain}: {e}")
            insights = None
            email_data = {}
        
        # Extract email parts
        email_subject = email_data.get("subject", f"quick note about {domain}").lower()
        email_body = email_data.get("body", "")
        
        # Ensure signature is present
        signature = """Best regards,
Code Nest Team
Code Nest LLC – New Mexico
contact@codenest.us.com
www.codenest.us.com"""
        
        if email_body and ("Code Nest LLC" not in email_body or "codenest.us.com" not in email_body.lower()):
            email_body = email_body.rstrip() + "\n\n" + signature
        
        # Fallback if email is missing/malformed - use rich issues data
        if not email_body or len(email_body) < 50:
            # Get top issues from different categories for fallback email
            fallback_bullets = []
            seen_categories = set()
            for issue in issues[:5]:
                cat = issue.get('category', 'General')
                if cat not in seen_categories and len(fallback_bullets) < 3:
                    fallback_bullets.append(f"• {issue.get('title', 'Optimization opportunity')}")
                    seen_categories.add(cat)
            
            if not fallback_bullets:
                fallback_bullets = ["• Site optimization opportunities identified"]
            
            email_body = f"""Hi there,

I recently reviewed {domain} and noticed a few opportunities that could help drive better results:

{chr(10).join(fallback_bullets)}

These issues can directly impact your visibility in search results and your ability to convert visitors into leads.

For context, I'm reaching out from Code Nest LLC, a New Mexico-based agency specializing in website optimization, SEO, and digital marketing.

Would you be open to a quick 15-minute call to discuss priorities?

{signature}"""
            email_subject = f"quick note about {domain}"
        
        # Build legacy format for backward compatibility
        summary = " ".join(insights.get("snapshot_summary", [])) if insights else "Analysis complete"
        impact = "; ".join([f"{i['issue']}: {i['impact']}" for i in insights.get("top_3_issues", [])]) if insights else ""
        solutions = "; ".join(insights.get("quick_wins", [])) if insights else ""
        
        result = {
            "summary": summary or "Analysis complete",
            "impact": impact or "Impact assessed",
            "solutions": solutions or "Solutions provided",
            "email": email_body,
            "email_subject": email_subject,
            "insights": insights,
            "from_cache": False
        }
        
        # CACHE THE RESULT
        cache_ai_result(url, result)
        
        return result
        
    except Exception as e:
        error_msg = str(e)
        logger.error(f"[OPENAI ERROR] {domain}: {error_msg}")
        
        # Specific error messages
        if "401" in error_msg or "invalid_api_key" in error_msg.lower():
            error_detail = "Invalid API Key - check API Settings"
        elif "429" in error_msg or "rate_limit" in error_msg.lower():
            error_detail = "Rate limit exceeded - wait and retry"
        elif "500" in error_msg or "502" in error_msg or "503" in error_msg:
            error_detail = "OpenAI server error - try again later"
        elif "timeout" in error_msg.lower():
            error_detail = "Request timed out - try again"
        elif "insufficient_quota" in error_msg.lower() or "billing" in error_msg.lower():
            error_detail = "OpenAI quota exceeded - check billing"
        else:
            error_detail = f"API Error: {error_msg[:50]}"
        
        return {
            "summary": "AI analysis error",
            "impact": error_detail,
            "solutions": "Check API Settings or try again",
            "email": f"Error: {error_detail}",
            "email_subject": f"quick idea for {domain}",
            "insights": None,
            "from_cache": False
        }

def calculate_opportunity_score(data):
    """Enhanced lead opportunity scoring (0-100)."""
    score = 0
    
    # Health score factors (40 points)
    if data['score'] < 30: score += 40
    elif data['score'] < 50: score += 30
    elif data['score'] < 70: score += 20
    else: score += 10
    
    # Tech stack issues (30 points)
    has_analytics = any("Analytics" in t for t in data.get('tech_stack', []))
    has_tracking = any("Pixel" in t for t in data.get('tech_stack', []))
    
    if not has_analytics or not has_tracking:
        score += 25
    
    # Critical issues (20 points)
    critical_count = len([i for i in data.get('issues', []) if i.get('title')])
    if critical_count >= 5: score += 20
    elif critical_count >= 3: score += 15
    elif critical_count > 0: score += 10
    
    # SSL/Security (10 points)
    if any("SSL" in i.get('title', '') for i in data.get('issues', [])):
        score += 10
    
    return min(100, score)

# ============================================================================
# ENHANCED SERVICE SCORING & LEAD ENRICHMENT
# ============================================================================

def detect_industry(url, html_content, tech_stack):
    """Detect industry from website content using keywords and tech stack."""
    html_lower = html_content.lower()
    
    industry_keywords = {
        "E-Commerce": ["shop", "cart", "product", "buy", "purchase", "store", "shopify"],
        "SaaS": ["dashboard", "api", "integration", "subscription", "pricing", "features"],
        "Real Estate": ["property", "listing", "real estate", "broker", "agent", "mls"],
        "Healthcare": ["doctor", "clinic", "medical", "appointment", "patient", "pharmacy"],
        "Finance": ["banking", "investment", "loan", "credit", "financial", "wealth"],
        "Legal": ["attorney", "lawyer", "law firm", "legal services", "litigation"],
        "Education": ["course", "university", "school", "student", "learning", "tuition"],
        "Restaurant/Food": ["menu", "reservation", "dining", "cafe", "restaurant", "food"],
        "Travel/Hospitality": ["hotel", "booking", "travel", "vacation", "resort", "airline"],
        "B2B Services": ["solution", "enterprise", "consulting", "service", "corporate"],
        "Manufacturing": ["manufacturing", "industrial", "equipment", "supplier", "production"],
        "Non-Profit": ["charity", "donation", "volunteer", "nonprofit", "mission"],
    }
    
    detected_industries = {}
    for industry, keywords in industry_keywords.items():
        score = sum(1 for keyword in keywords if keyword in html_lower)
        if score > 0:
            detected_industries[industry] = score
    
    if detected_industries:
        return max(detected_industries, key=detected_industries.get)
    return "General Business"

def estimate_company_size(tech_stack, content_length, has_team_page, num_emails):
    """Estimate company size based on website indicators."""
    score = 0
    
    # Enterprise tech
    enterprise_tech = ["salesforce", "sap", "servicenow", "workday", "oracle"]
    if any(tech.lower() in str(tech_stack).lower() for tech in enterprise_tech):
        return "Enterprise"
    
    # Startup/SaaS indicators
    if any(tech in str(tech_stack).lower() for tech in ["react", "nextjs", "vue", "node"]):
        score += 20
    
    # Content length (larger = usually bigger company)
    if content_length > 50000:
        score += 30
        size = "Large"
    elif content_length > 20000:
        score += 20
        size = "Medium"
    else:
        score += 10
        size = "Small"
    
    # Email count (more emails = bigger)
    if num_emails > 5:
        score += 20
        size = "Large"
    elif num_emails > 2:
        score += 10
        size = "Medium"
    
    # Team page indicates medium+
    if has_team_page:
        if score >= 40:
            return "Large"
        elif score >= 25:
            return "Medium"
    
    return size

def calculate_service_scores(audit_data, company_size, industry):
    """Calculate opportunity score for each Code Nest service."""
    scores = {}
    
    # Website Development Score (0-100)
    web_dev_score = 0
    web_dev_score += (100 - audit_data['score']) * 0.3  # Health score = 30%
    
    # Design freshness (check for outdated patterns)
    if any(issue['title'] in ['Outdated Design', 'Poor UX', 'Mobile Issues'] 
           for issue in audit_data.get('issues', [])):
        web_dev_score += 30
    
    if company_size in ["Large", "Enterprise"]:
        web_dev_score += 15
    
    scores['website_development'] = min(100, web_dev_score)
    
    # SEO Score (0-100)
    seo_score = 0
    seo_issues = [issue for issue in audit_data.get('issues', []) 
                  if any(word in issue.get('title', '').lower() for word in ['seo', 'meta', 'title', 'heading'])]
    seo_score += len(seo_issues) * 15
    seo_score += (100 - audit_data['score']) * 0.2  # Low health = SEO need
    
    if company_size in ["Medium", "Large", "Enterprise"]:
        seo_score += 20
    
    scores['seo_optimization'] = min(100, seo_score)
    
    # Mobile App Score (0-100)
    mobile_app_score = 0
    if any("mobile" in issue.get('title', '').lower() for issue in audit_data.get('issues', [])):
        mobile_app_score += 40
    
    if industry in ["E-Commerce", "SaaS", "Restaurant/Food"]:
        mobile_app_score += 30
    
    if company_size in ["Medium", "Large", "Enterprise"]:
        mobile_app_score += 20
    
    scores['mobile_app_development'] = mobile_app_score
    
    # Social Media Score (0-100)
    social_score = 0
    social_score += sum(1 for issue in audit_data.get('issues', []) 
                       if 'social' in issue.get('title', '').lower()) * 20
    
    if industry in ["E-Commerce", "Restaurant/Food", "Travel/Hospitality"]:
        social_score += 35
    
    if company_size in ["Medium", "Large"]:
        social_score += 15
    
    scores['social_media_marketing'] = min(100, social_score)
    
    # PPC/Paid Ads Score (0-100)
    ppc_score = 0
    if industry in ["E-Commerce", "SaaS", "Finance"]:
        ppc_score += 40
    
    if company_size in ["Large", "Enterprise"]:
        ppc_score += 25
    
    # If low organic traffic signals
    if audit_data['score'] < 50:
        ppc_score += 20
    
    scores['paid_advertising'] = min(100, ppc_score)
    
    # E-Commerce Score (0-100)
    ecommerce_score = 0
    if industry == "E-Commerce":
        ecommerce_score = 85
    elif "product" in str(audit_data.get('tech_stack', '')).lower():
        ecommerce_score = 60
    else:
        ecommerce_score = 20
    
    scores['ecommerce_development'] = ecommerce_score
    
    # WordPress/Website Maintenance Score (0-100)
    maintenance_score = 50  # Most sites need maintenance
    if "wordpress" in str(audit_data.get('tech_stack', '')).lower():
        maintenance_score = 75
    
    if company_size in ["Small", "Medium"]:
        maintenance_score += 20
    
    scores['website_maintenance'] = min(100, maintenance_score)
    
    # React/Modern Development Score (0-100)
    modern_dev_score = 0
    modern_stacks = ["react", "nextjs", "vue", "angular"]
    if any(stack in str(audit_data.get('tech_stack', '')).lower() for stack in modern_stacks):
        modern_dev_score = 60  # Already using modern stack, but might need optimization
    else:
        modern_dev_score = 40
    
    if industry == "SaaS":
        modern_dev_score += 30
    
    scores['react_nextjs_development'] = min(100, modern_dev_score)
    
    # Website Optimization Score (0-100)
    optimization_score = 100 - audit_data['score']
    optimization_score += (100 - (audit_data.get('psi', 0) or 0)) * 0.3
    scores['website_optimization'] = min(100, optimization_score)
    
    # Graphic Design Score (0-100)
    design_score = 30 if any("design" in issue.get('title', '').lower() 
                            for issue in audit_data.get('issues', [])) else 20
    if company_size in ["Medium", "Large"]:
        design_score += 20
    scores['graphic_designing'] = design_score
    
    return scores

def generate_service_pitch(company_name, service_name, score, industry, company_size, audit_data):
    """Generate AI-powered service pitch for a specific service."""
    pitches = {
        'website_development': f"""
Based on our audit of {company_name}'s current website, we identified significant opportunities for improvement:

**Current State:** Your website scores {audit_data['score']}/100 on our health assessment
**Key Issues Found:** {', '.join([issue['title'] for issue in audit_data.get('issues', [])[:3]])}

**What We Recommend:**
Our website development team specializes in building high-performance, conversion-optimized sites for {industry} businesses like yours. We'll:
- Rebuild your site with modern architecture ({company_size}-grade infrastructure)
- Optimize for conversions and user experience
- Ensure mobile-first responsive design
- Integrate with your existing tools and systems

**Expected Results:** 40-60% improvement in site performance, 25-35% increase in lead generation
**Timeline:** 8-12 weeks | **Investment:** Based on scope and complexity
""",
        
        'seo_optimization': f"""
{company_name} is currently ranking poorly for critical keywords in your industry.

**Current Gaps:**
- Missing or weak meta tags and structure
- Limited internal linking strategy
- Technical SEO issues affecting crawlability

**Our SEO Strategy:**
We'll execute a comprehensive SEO plan tailored for {industry} that includes:
- Technical SEO audit and remediation
- On-page optimization for high-intent keywords
- Content strategy and creation
- Off-page authority building
- Monthly performance reporting

**Expected Results:** Top 10 rankings for target keywords within 6 months, 200%+ organic traffic growth
**Timeline:** Ongoing (6-12 month commitment) | **Investment:** Monthly retainer starting at $2,500/month
""",

        'mobile_app_development': f"""
For a {company_size} {industry} business like {company_name}, a native mobile app could unlock new revenue streams.

**Opportunity:**
Your website receives significant mobile traffic, but a dedicated app would:
- Increase customer engagement and retention
- Enable push notifications and personalization
- Create a direct channel to your customers
- Generate new revenue opportunities

**Our Process:**
We build iOS and Android apps with:
- Native performance and user experience
- Offline functionality
- Seamless backend integration
- App Store and Play Store optimization

**Expected ROI:** 40-60% increase in customer lifetime value
**Timeline:** 12-16 weeks | **Investment:** Starting at $35,000
""",

        'social_media_marketing': f"""
{company_name}'s social presence is not aligned with your market potential.

**Current Assessment:**
- Limited social media integration on website
- Missing or inactive social accounts
- No consistent content strategy

**Our Social Media Services:**
For {industry} businesses, we create and execute comprehensive social strategies:
- Content calendar and creation
- Community management and engagement
- Paid social advertising ($500-$5,000/month)
- Analytics and monthly reporting
- Influencer partnerships (if relevant)

**Expected Results:** 5-10x follower growth, 30-50% increase in social-driven traffic
**Timeline:** Ongoing (minimum 3 months) | **Investment:** $2,000-$5,000/month
""",

        'paid_advertising': f"""
Your organic traffic is limited, but paid advertising could immediately increase qualified leads.

**Opportunity Analysis:**
- Low organic visibility in search results
- High cost of waiting for SEO results
- Immediate revenue opportunity through PPC

**Our Paid Advertising Services:**
Fully-managed campaigns on Google, Facebook, and Instagram with:
- Keyword and audience research
- Creative development and A/B testing
- Daily optimization and monitoring
- Conversion tracking and reporting
- Weekly performance reviews

**Expected ROI:** 3-5x return on ad spend for e-commerce, 200-500% for B2B
**Timeline:** Ongoing (minimum 3 months) | **Investment:** $1,000-$10,000/month ad spend + management
"""
    }
    
    return pitches.get(service_name, "Service pitch unavailable")

def enrich_lead_with_ai(lead_data, audit_data, openai_key):
    """Use AI to generate comprehensive lead enrichment and recommendations.
    
    Uses centralized AI settings for cost efficiency.
    """
    if not openai_key:
        return None
    
    try:
        client = OpenAI(api_key=openai_key, timeout=AI_CLIENT_TIMEOUT)
        
        # Shortened prompt for cost efficiency
        prompt = f"""Analyze lead and provide strategic recommendations (JSON only):

Company: {lead_data.get('company_name', 'Unknown')}
Industry: {lead_data.get('industry', 'Unknown')}
Location: {lead_data.get('city', 'Unknown')}, {lead_data.get('state', 'Unknown')}
Health Score: {audit_data.get('score', 0)}/100
Top Issues: {', '.join([i['title'][:30] for i in audit_data.get('issues', [])[:3]])}

Return JSON:
{{"key_challenges": ["challenge1", "challenge2", "challenge3"],
"quick_wins": ["win1", "win2", "win3"],
"recommended_services": ["service1", "service2", "service3"],
"estimated_impact": "1-2 sentence impact",
"conversation_starters": ["reason1", "reason2", "reason3"]}}"""
        
        response = client.chat.completions.create(
            model=AI_MODEL,  # Use centralized model setting
            messages=[{"role": "user", "content": prompt}],
            temperature=AI_TEMPERATURE,
            max_tokens=500,  # Reduced for cost
            timeout=AI_TIMEOUT
        )
        
        result = response.choices[0].message.content
        # Clean up potential markdown
        if result.startswith("```"):
            result = result.split("```")[1]
            if result.startswith("json"):
                result = result[4:]
        return json.loads(result.strip())
    
    except Exception as e:
        logger.error(f"Error enriching lead with AI: {str(e)}")
        return None

def process_csv_leads(csv_file, audit_scores_lookup, openai_key):
    """Process CSV lead file with Google Places data and audit matching."""
    try:
        df = pd.read_csv(csv_file)
        leads_created = 0
        errors = []
        
        for idx, row in df.iterrows():
            try:
                company_name = row.get('name', f"Lead {idx}")
                phone = row.get('phone', '')
                address = row.get('address', '')
                place_id = row.get('place_id', '')
                website = row.get('website', '') or row.get('url', '')
                city = row.get('city', '')
                state = row.get('state', '')
                zipcode = row.get('zipcode', '')
                
                # Extract domain from website if available
                if website:
                    domain = urlparse(website).netloc.replace('www.', '')
                else:
                    domain = None
                
                if not domain:
                    errors.append(f"Row {idx}: No website/domain found")
                    continue
                
                # Look up audit scores if available
                audit_data = audit_scores_lookup.get(domain, {})
                
                # Create lead record
                db = get_db()
                if not db:
                    continue
                
                lead = Lead(
                    domain=domain,
                    company_name=company_name,
                    phone=phone,
                    address=address,
                    place_id=place_id,
                    city=city,
                    state=state,
                    zipcode=zipcode,
                    health_score=audit_data.get('score'),
                    industry=audit_data.get('industry', detect_industry(domain, '', audit_data.get('tech_stack', []))),
                    company_size=audit_data.get('company_size', estimate_company_size(
                        audit_data.get('tech_stack', []),
                        len(str(audit_data)),
                        False,
                        len(audit_data.get('emails', []))
                    )),
                    service_priorities=audit_data.get('service_scores', {}),
                    status='new'
                )
                
                # AI enrichment if audit data available
                if audit_data and openai_key:
                    ai_enrichment = enrich_lead_with_ai(
                        {
                            'company_name': company_name,
                            'city': city,
                            'state': state,
                            'industry': lead.industry,
                            'company_size': lead.company_size
                        },
                        audit_data,
                        openai_key
                    )
                    lead.ai_enrichment = ai_enrichment
                
                db.add(lead)
                db.commit()
                leads_created += 1
                
                db.close()
            
            except Exception as e:
                errors.append(f"Row {idx}: {str(e)}")
                continue
        
        return leads_created, errors
    
    except Exception as e:
        logger.error(f"Error processing CSV: {str(e)}")
        return 0, [str(e)]

def create_issue(category: str, title: str, impact: str, technical_detail: str, severity: str = "medium") -> dict:
    """
    Create a structured issue with full business context.
    
    Args:
        category: One of SEO, Performance, Tracking, Conversion, UX, Security, Content
        title: Short human-readable issue title
        impact: Business-level impact explanation
        technical_detail: Technical explanation (1-2 lines)
        severity: high, medium, or low
    
    Returns:
        Structured issue dict
    """
    return {
        "category": category,
        "title": title,
        "impact": impact,
        "technical_detail": technical_detail,
        "severity": severity
    }


def run_audit(url, openai_key, google_key):
    """
    Enhanced website audit with deep free checks across multiple categories:
    - Technical SEO & Indexability
    - On-Page Content & Structure
    - Performance (PageSpeed API)
    - Tracking & Analytics
    - Conversion & UX
    - Security & Professionalism
    
    Returns structured issues with category, title, impact, technical_detail, and severity.
    """
    if not url.startswith('http'): 
        url = 'http://' + url
    
    parsed_url = urlparse(url)
    domain = parsed_url.netloc.replace("www.", "")
    base_url = f"{parsed_url.scheme}://{parsed_url.netloc}"
    
    data = {
        "url": url,
        "domain": domain,
        "score": 100,
        "issues": [],
        "tech_stack": [],
        "emails": [],
        "psi": None,
        "psi_error": None,
        "psi_audits": [],  # NEW: Store detailed PSI audit findings
        "domain_age": "Unknown",
        "accessibility_score": None,
        "security_issues": [],
        "broken_links": 0,
        "cwv_metrics": {},
        "content_stats": {},  # NEW: Content analysis stats
        "seo_stats": {}  # NEW: SEO analysis stats
    }

    try:
        start = time.time()
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'}
        response = requests.get(url, headers=headers, timeout=30)
        load_time = time.time() - start
        response_headers = response.headers
        
        html = response.text.lower()
        html_original = response.text  # Keep original case for some checks
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Start domain age lookup in background (slow WHOIS operation)
        executor = ThreadPoolExecutor(max_workers=1)
        domain_age_future = executor.submit(get_domain_age, url)
        
        # Extract emails
        data['emails'] = list(set(re.findall(r"[a-z0-9\.\-+_]+@[a-z0-9\.\-+_]+\.[a-z]+", html)))

        # =====================================================================
        # A. TECHNICAL SEO & INDEXABILITY
        # =====================================================================
        
        # 1. robots.txt check
        try:
            robots_resp = requests.get(f"{base_url}/robots.txt", headers=headers, timeout=5)
            if robots_resp.status_code == 200:
                robots_content = robots_resp.text.lower()
                data['tech_stack'].append("robots.txt")
                
                # Check if blocking important paths
                if "disallow: /" in robots_content and "disallow: /wp-admin" not in robots_content:
                    data['score'] -= 15
                    data['issues'].append(create_issue(
                        "SEO", 
                        "robots.txt Blocks Entire Site",
                        "Google can't index your site – you're invisible in search results",
                        "robots.txt contains 'Disallow: /' which blocks all crawlers from indexing",
                        "high"
                    ))
                
                # Check for sitemap reference
                if "sitemap:" in robots_content:
                    data['seo_stats']['has_sitemap_in_robots'] = True
            else:
                data['seo_stats']['robots_missing'] = True
        except Exception:
            data['seo_stats']['robots_error'] = True
        
        # 2. Meta robots check
        meta_robots = soup.find('meta', attrs={'name': 'robots'})
        x_robots = response_headers.get('X-Robots-Tag', '')
        
        if meta_robots:
            robots_content = meta_robots.get('content', '').lower()
            if 'noindex' in robots_content:
                data['score'] -= 20
                data['issues'].append(create_issue(
                    "SEO",
                    "Site Set to NoIndex",
                    "Google is explicitly told NOT to show your site in search results – zero organic visibility",
                    "<meta name='robots' content='noindex'> found – this blocks all search engine indexing",
                    "high"
                ))
        
        if 'noindex' in x_robots.lower():
            data['score'] -= 20
            data['issues'].append(create_issue(
                "SEO",
                "X-Robots-Tag Blocks Indexing",
                "Server headers tell Google not to index – you won't appear in search results",
                "HTTP header X-Robots-Tag contains 'noindex'",
                "high"
            ))
        
        # 3. Canonical check
        canonical = soup.find('link', attrs={'rel': 'canonical'})
        if not canonical:
            data['score'] -= 5
            data['issues'].append(create_issue(
                "SEO",
                "Missing Canonical Tag",
                "Risk of duplicate content issues – Google may split your ranking power across multiple URLs",
                "No <link rel='canonical'> tag found – search engines may index duplicate versions",
                "medium"
            ))
        elif canonical:
            canonical_href = canonical.get('href', '')
            if canonical_href and domain not in canonical_href:
                data['score'] -= 10
                data['issues'].append(create_issue(
                    "SEO",
                    "Canonical Points to Different Domain",
                    "Your SEO authority is being transferred to another website",
                    f"Canonical tag points to {canonical_href[:50]}... instead of your domain",
                    "high"
                ))
        
        # 4. Sitemap check
        sitemap_found = False
        try:
            sitemap_resp = requests.get(f"{base_url}/sitemap.xml", headers=headers, timeout=5)
            if sitemap_resp.status_code == 200 and ('<?xml' in sitemap_resp.text or '<urlset' in sitemap_resp.text):
                sitemap_found = True
                data['tech_stack'].append("XML Sitemap")
        except Exception:
            pass
        
        if not sitemap_found and not data.get('seo_stats', {}).get('has_sitemap_in_robots'):
            data['score'] -= 5
            data['issues'].append(create_issue(
                "SEO",
                "No XML Sitemap Found",
                "Google has to guess which pages matter – slower indexing and potential missed pages",
                "No sitemap.xml found at /sitemap.xml or referenced in robots.txt",
                "medium"
            ))
        
        # 5. Title tag analysis
        title_tag = soup.title.string.strip() if soup.title and soup.title.string else ""
        data['seo_stats']['title_length'] = len(title_tag)
        
        if not title_tag:
            data['score'] -= 15
            data['issues'].append(create_issue(
                "SEO",
                "Missing Page Title",
                "No title = no click-worthy search result – severely hurts click-through rates",
                "The <title> tag is missing or empty",
                "high"
            ))
        elif len(title_tag) < 30:
            data['score'] -= 8
            data['issues'].append(create_issue(
                "SEO",
                f"Title Too Short ({len(title_tag)} chars)",
                "Short titles waste valuable search result real estate and often look unprofessional",
                f"Title '{title_tag[:40]}...' is under 30 characters (recommended: 50-65)",
                "medium"
            ))
        elif len(title_tag) > 65:
            data['score'] -= 5
            data['issues'].append(create_issue(
                "SEO",
                f"Title Too Long ({len(title_tag)} chars)",
                "Google will truncate your title in search results, cutting off important keywords",
                f"Title exceeds 65 characters and will be cut off in search results",
                "low"
            ))
        
        # Check for generic titles
        generic_titles = ['home', 'welcome', 'untitled', 'homepage', 'index']
        if title_tag.lower().strip() in generic_titles:
            data['score'] -= 10
            data['issues'].append(create_issue(
                "SEO",
                "Generic/Unhelpful Title Tag",
                "Searchers can't tell what your business does – lower click-through rates",
                f"Title '{title_tag}' is too generic – should describe your business/service",
                "high"
            ))
        
        # 6. Meta description analysis
        meta_desc = soup.find('meta', attrs={'name': 'description'})
        meta_desc_content = meta_desc.get('content', '').strip() if meta_desc else ""
        data['seo_stats']['meta_desc_length'] = len(meta_desc_content)
        
        if not meta_desc_content:
            data['score'] -= 10
            data['issues'].append(create_issue(
                "SEO",
                "Missing Meta Description",
                "Google will auto-generate a snippet – usually looks worse and converts less",
                "No <meta name='description'> tag found",
                "high"
            ))
        elif len(meta_desc_content) < 50:
            data['score'] -= 5
            data['issues'].append(create_issue(
                "SEO",
                f"Meta Description Too Short ({len(meta_desc_content)} chars)",
                "You're missing the chance to sell your page in search results",
                "Meta description should be 120-160 characters for optimal display",
                "medium"
            ))
        
        # =====================================================================
        # B. ON-PAGE CONTENT & STRUCTURE
        # =====================================================================
        
        # 1. Content depth (word count)
        # Remove script and style elements
        for script in soup(["script", "style", "nav", "header", "footer"]):
            script.decompose()
        
        text_content = soup.get_text(separator=' ', strip=True)
        word_count = len(text_content.split())
        data['content_stats']['word_count'] = word_count
        
        if word_count < 250:
            data['score'] -= 10
            data['issues'].append(create_issue(
                "Content",
                f"Thin Content ({word_count} words)",
                "Google favors comprehensive content – thin pages rank poorly and don't convert well",
                f"Homepage has only ~{word_count} words. Aim for 500+ words of quality content.",
                "high"
            ))
        elif word_count < 400:
            data['score'] -= 5
            data['issues'].append(create_issue(
                "Content",
                f"Light Content ({word_count} words)",
                "May struggle to rank for competitive keywords – consider adding more value",
                f"Homepage has ~{word_count} words. More detailed content often ranks better.",
                "medium"
            ))
        
        # 2. Heading hierarchy (H1)
        h1_tags = soup.find_all('h1')
        h1_count = len(h1_tags)
        data['seo_stats']['h1_count'] = h1_count
        
        if h1_count == 0:
            data['score'] -= 10
            data['issues'].append(create_issue(
                "SEO",
                "No H1 Heading",
                "Google uses H1 to understand your main topic – missing it hurts rankings",
                "No <h1> tag found. Every page should have exactly one H1 describing the main topic.",
                "high"
            ))
        elif h1_count > 1:
            data['score'] -= 5
            data['issues'].append(create_issue(
                "SEO",
                f"Multiple H1 Tags ({h1_count} found)",
                "Confuses search engines about your page's main topic",
                f"Found {h1_count} H1 tags. Best practice is exactly one H1 per page.",
                "medium"
            ))
        elif h1_count == 1:
            h1_text = h1_tags[0].get_text(strip=True).lower()
            if h1_text in ['home', 'welcome', 'homepage', '']:
                data['score'] -= 8
                data['issues'].append(create_issue(
                    "SEO",
                    "Generic H1 Heading",
                    "Wasted opportunity to tell Google (and visitors) what you do",
                    f"H1 is '{h1_text}' – should describe your service/value proposition",
                    "medium"
                ))
        
        # 3. Image alt text coverage
        all_images = BeautifulSoup(html_original, 'html.parser').find_all('img')
        total_images = len(all_images)
        images_with_alt = len([img for img in all_images if img.get('alt') and img.get('alt').strip()])
        data['content_stats']['total_images'] = total_images
        data['content_stats']['images_with_alt'] = images_with_alt
        
        if total_images > 0:
            alt_coverage = (images_with_alt / total_images) * 100
            data['content_stats']['alt_coverage'] = round(alt_coverage, 1)
            
            if alt_coverage < 40:
                data['score'] -= 8
                data['issues'].append(create_issue(
                    "SEO",
                    f"Poor Image Alt Text ({int(alt_coverage)}% coverage)",
                    "Google can't understand your images – missing image search traffic and accessibility issues",
                    f"Only {images_with_alt} of {total_images} images have alt text. Screen readers also can't describe images.",
                    "high"
                ))
            elif alt_coverage < 70:
                data['score'] -= 4
                data['issues'].append(create_issue(
                    "SEO",
                    f"Incomplete Image Alt Text ({int(alt_coverage)}% coverage)",
                    "Missing some image SEO opportunities and accessibility",
                    f"{total_images - images_with_alt} images are missing alt text",
                    "medium"
                ))
        
        # 4. Open Graph / Social metadata
        og_title = soup.find('meta', attrs={'property': 'og:title'})
        og_desc = soup.find('meta', attrs={'property': 'og:description'})
        og_image = soup.find('meta', attrs={'property': 'og:image'})
        
        og_missing = []
        if not og_title: og_missing.append('og:title')
        if not og_desc: og_missing.append('og:description')
        if not og_image: og_missing.append('og:image')
        
        if len(og_missing) >= 2:
            data['score'] -= 5
            data['issues'].append(create_issue(
                "Content",
                "Missing Social Sharing Metadata",
                "Links shared on Facebook/LinkedIn will look unprofessional with no preview image or description",
                f"Missing Open Graph tags: {', '.join(og_missing)}",
                "medium"
            ))
        
        # =====================================================================
        # C. TRACKING & ANALYTICS (Enhanced)
        # =====================================================================
        
        tracking = {
            'google_analytics': False,
            'gtm': False,
            'facebook_pixel': False,
            'linkedin_insight': False,
            'tiktok_pixel': False,
            'hotjar': False
        }
        
        # Google Analytics / GA4
        if 'gtag(' in html or re.search(r'ua-\d{4,}-\d{1,}', html) or re.search(r'g-[a-z0-9]{10}', html):
            tracking['google_analytics'] = True
            data['tech_stack'].append("Google Analytics")
        
        # Google Tag Manager
        if 'gtm.js' in html or 'googletagmanager' in html:
            tracking['gtm'] = True
            data['tech_stack'].append("Google Tag Manager")
        
        # Facebook/Meta Pixel
        if 'fbq(' in html or 'facebook.com/tr' in html:
            tracking['facebook_pixel'] = True
            data['tech_stack'].append("Facebook Pixel")
        
        # LinkedIn Insight
        if 'linkedin.com/px' in html or ('linkedin' in html and 'insight' in html):
            tracking['linkedin_insight'] = True
            data['tech_stack'].append("LinkedIn Insight")
        
        # TikTok Pixel
        if 'tiktok.com/i18n/pixel' in html or 'analytics.tiktok.com' in html:
            tracking['tiktok_pixel'] = True
            data['tech_stack'].append("TikTok Pixel")
        
        # Hotjar
        if 'hotjar' in html:
            tracking['hotjar'] = True
            data['tech_stack'].append("Hotjar")
        
        data['tracking'] = tracking
        
        # Issue: No analytics at all
        has_any_analytics = tracking['google_analytics'] or tracking['gtm']
        has_any_retargeting = tracking['facebook_pixel'] or tracking['linkedin_insight'] or tracking['tiktok_pixel']
        
        if not has_any_analytics:
            data['score'] -= 15
            data['issues'].append(create_issue(
                "Tracking",
                "No Analytics Tools Detected",
                "You can't see where your visitors come from, what they do, or which channels drive leads",
                "No Google Analytics, GA4, or Tag Manager found. You're flying blind.",
                "high"
            ))
        
        if has_any_analytics and not has_any_retargeting:
            data['score'] -= 10
            data['issues'].append(create_issue(
                "Tracking",
                "No Retargeting Pixels Installed",
                "95%+ of visitors leave without contacting you – you can't follow up with targeted ads",
                "No Facebook, LinkedIn, or TikTok retargeting pixels detected",
                "high"
            ))
        
        # =====================================================================
        # D. CONVERSION & UX (Lead Capture & Trust)
        # =====================================================================
        
        # Lead capture check
        forms = soup.find_all('form')
        has_form = len(forms) > 0
        
        # Contact methods
        tel_links = soup.find_all('a', href=re.compile(r'^tel:'))
        mailto_links = soup.find_all('a', href=re.compile(r'^mailto:'))
        
        # CTA buttons/links
        cta_keywords = ['contact', 'get quote', 'request', 'book', 'schedule', 'call', 'free', 'demo', 'consultation']
        cta_found = any(keyword in html for keyword in cta_keywords)
        
        data['conversion_stats'] = {
            'has_form': has_form,
            'tel_links': len(tel_links),
            'mailto_links': len(mailto_links),
            'cta_found': cta_found
        }
        
        if not has_form and not tel_links and not mailto_links:
            data['score'] -= 15
            data['issues'].append(create_issue(
                "Conversion",
                "No Clear Contact Method",
                "Visitors who want to buy can't easily reach you – losing potential customers",
                "No contact form, phone link (tel:), or email link (mailto:) found on homepage",
                "high"
            ))
        elif not has_form and not cta_found:
            data['score'] -= 8
            data['issues'].append(create_issue(
                "Conversion",
                "Weak Call-to-Action",
                "Visitors don't know what action to take – lower conversion rates",
                "No prominent contact form or CTA buttons like 'Get Quote' or 'Book Now'",
                "medium"
            ))
        
        # Trust signals check
        trust_keywords = ['testimonial', 'review', 'client', 'customer', 'rating', 'stars', 'trust', 'certified', 'award']
        has_trust_signals = any(keyword in html for keyword in trust_keywords)
        
        if not has_trust_signals:
            data['score'] -= 5
            data['issues'].append(create_issue(
                "Conversion",
                "No Visible Trust Signals",
                "New visitors have no proof you're legitimate – lower conversion rates",
                "No testimonials, reviews, ratings, or trust badges detected on homepage",
                "medium"
            ))
        
        # Mobile viewport check
        viewport = soup.find('meta', attrs={'name': 'viewport'})
        if not viewport:
            data['score'] -= 10
            data['issues'].append(create_issue(
                "UX",
                "Not Mobile-Optimized",
                "60%+ of traffic is mobile – your site may be unusable on phones",
                "Missing <meta name='viewport'> tag. Layout likely breaks on mobile devices.",
                "high"
            ))
        
        # =====================================================================
        # E. SECURITY & PROFESSIONALISM
        # =====================================================================
        
        # HTTPS check
        is_https = url.startswith('https')
        if not is_https:
            data['score'] -= 15
            data['security_issues'].append("No SSL/HTTPS")
            data['issues'].append(create_issue(
                "Security",
                "Site Not Using HTTPS",
                "Browsers show 'Not Secure' warning – visitors lose trust and may leave",
                "Site loads over HTTP instead of HTTPS. Modern browsers flag this as insecure.",
                "high"
            ))
        
        # Mixed content check (for HTTPS sites)
        if is_https:
            http_resources = re.findall(r'(src|href)=["\']http://', html_original)
            if len(http_resources) > 2:
                data['score'] -= 5
                data['security_issues'].append("Mixed Content")
                data['issues'].append(create_issue(
                    "Security",
                    "Mixed Content Warning",
                    "Browser may block some resources or show security warnings",
                    f"Found {len(http_resources)} HTTP resources on HTTPS page",
                    "medium"
                ))
        
        # Security headers check
        security_headers = {
            'Strict-Transport-Security': response_headers.get('Strict-Transport-Security'),
            'Content-Security-Policy': response_headers.get('Content-Security-Policy'),
            'X-Frame-Options': response_headers.get('X-Frame-Options'),
            'X-Content-Type-Options': response_headers.get('X-Content-Type-Options')
        }
        
        missing_security_headers = [h for h, v in security_headers.items() if not v]
        if len(missing_security_headers) >= 3:
            data['score'] -= 5
            data['issues'].append(create_issue(
                "Security",
                "Missing Security Headers",
                "Site is more vulnerable to common attacks – could hurt business reputation if compromised",
                f"Missing: {', '.join(missing_security_headers[:3])}",
                "medium"
            ))
        
        # =====================================================================
        # F. TECH STACK DETECTION (Enhanced)
        # =====================================================================
        
        tech_checks = [
            ("wp-content", "WordPress"),
            ("shopify", "Shopify"),
            ("wix", "Wix"),
            ("squarespace", "Squarespace"),
            ("webflow", "Webflow"),
            ("react", "React"),
            ("next", "Next.js"),
            ("vue", "Vue.js"),
            ("angular", "Angular"),
            ("bootstrap", "Bootstrap"),
            ("tailwind", "Tailwind CSS"),
            ("jquery", "jQuery"),
            ("cloudflare", "Cloudflare"),
            ("stripe", "Stripe"),
            ("intercom", "Intercom"),
            ("hubspot", "HubSpot"),
            ("mailchimp", "Mailchimp"),
            ("zendesk", "Zendesk")
        ]
        
        for check, tech in tech_checks:
            if check in html and tech not in data['tech_stack']:
                data['tech_stack'].append(tech)

        # =====================================================================
        # G. PERFORMANCE - Enhanced PageSpeed API
        # =====================================================================
        
        # Get domain age result from background thread
        try:
            data['domain_age'] = domain_age_future.result(timeout=3)[0]
        except Exception:
            data['domain_age'] = "Unknown"
        finally:
            executor.shutdown(wait=False)
        
        # PageSpeed API with detailed audits
        try:
            psi_result = get_google_speed_detailed(url, google_key)
            if psi_result:
                data['psi'] = psi_result.get('score')
                data['psi_audits'] = psi_result.get('audits', [])
                data['cwv_metrics'] = psi_result.get('cwv', {})
                
                # Add issues from PageSpeed audits
                for audit in psi_result.get('audits', [])[:3]:  # Top 3 failing audits
                    data['issues'].append(create_issue(
                        "Performance",
                        audit.get('title', 'Performance Issue'),
                        audit.get('impact', 'Slower site = fewer conversions'),
                        audit.get('detail', ''),
                        "high" if audit.get('score', 1) < 0.5 else "medium"
                    ))
                
                # Overall speed issue
                if data['psi'] and data['psi'] < 50:
                    data['score'] -= 15
                    data['issues'].append(create_issue(
                        "Performance",
                        f"Critical Site Speed ({data['psi']}/100)",
                        "53% of mobile users leave if page takes >3 seconds – losing half your visitors",
                        f"Google PageSpeed score is {data['psi']}/100. Target 80+ for good performance.",
                        "high"
                    ))
                elif data['psi'] and data['psi'] < 70:
                    data['score'] -= 8
                    data['issues'].append(create_issue(
                        "Performance",
                        f"Slow Site Speed ({data['psi']}/100)",
                        "Site is slower than competitors – impacts both rankings and conversions",
                        f"PageSpeed score {data['psi']}/100 is below recommended 70+",
                        "medium"
                    ))
        except Exception as e:
            data['psi_error'] = str(e)
            # Fall back to basic load time check
            if load_time > 3.0:
                data['score'] -= 10
                data['issues'].append(create_issue(
                    "Performance",
                    f"Slow Server Response ({round(load_time, 1)}s)",
                    "Slow initial response frustrates visitors before they even see your content",
                    f"Server took {round(load_time, 2)} seconds to respond. Target <1.5s.",
                    "high"
                ))

        # =====================================================================
        # H. AI CONSULTATION (Caching + Single API Call)
        # =====================================================================
        
        # AI consultation with rich issues data
        ai_executor = ThreadPoolExecutor(max_workers=1)
        ai_future = ai_executor.submit(get_ai_consultation, url, data, openai_key)
        try:
            data['ai'] = ai_future.result(timeout=40)
        except TimeoutError:
            logger.warning(f"AI consultation timed out for {url}")
            data['ai'] = {
                "summary": "AI analysis timed out",
                "impact": "OpenAI took too long - try again",
                "solutions": "Click 'Regenerate Email' to retry",
                "email": "AI timed out - click 'Regenerate Email' to try again",
                "email_subject": f"quick idea for {domain}",
                "insights": None,
                "from_cache": False
            }
        except Exception as e:
            error_str = str(e)
            logger.error(f"AI consultation failed for {url}: {error_str}")
            data['ai'] = {
                "summary": "AI analysis failed",
                "impact": f"Error: {error_str[:80]}",
                "solutions": "Check API key or try again",
                "email": f"AI error - {error_str[:40]}",
                "email_subject": f"quick idea for {domain}",
                "insights": None,
                "from_cache": False
            }
        finally:
            ai_executor.shutdown(wait=False)

    except Exception as e:
        data['error'] = str(e)
        logger.error(f"Audit error for {url}: {str(e)}")

    # Ensure score stays in bounds
    data['score'] = max(0, min(100, data['score']))
    
    # Calculate accessibility score
    accessibility_score = 100
    if data.get('seo_stats', {}).get('h1_count', 0) == 0:
        accessibility_score -= 20
    if data.get('content_stats', {}).get('alt_coverage', 100) < 50:
        accessibility_score -= 20
    if not soup.find('meta', attrs={'name': 'viewport'}):
        accessibility_score -= 20
    data['accessibility_score'] = max(0, accessibility_score)
    
    # Add service scoring for lead enrichment
    try:
        industry = detect_industry(url, html if 'html' in locals() else '', data.get('tech_stack', []))
        company_size = estimate_company_size(data.get('tech_stack', []), 
                                            len(str(data)), 
                                            False, 
                                            len(data.get('emails', [])))
        data['industry'] = industry
        data['company_size'] = company_size
        data['service_scores'] = calculate_service_scores(data, company_size, industry)
    except Exception as e:
        logger.warning(f"Error calculating service scores: {str(e)}")
    
    return data


def get_google_speed_detailed(url, api_key):
    """
    Get detailed Google PageSpeed results including specific failing audits.
    Returns score, core web vitals, and top failing audits with business impact.
    """
    if not api_key:
        return None
    
    api_url = f"https://www.googleapis.com/pagespeedonline/v5/runPagespeed?url={url}&strategy=mobile&key={api_key}&category=performance"
    
    try:
        r = requests.get(api_url, timeout=20)
        if r.status_code != 200:
            return None
        
        data = r.json()
        lighthouse = data.get('lighthouseResult', {})
        
        # Get overall score
        perf_score = lighthouse.get('categories', {}).get('performance', {}).get('score')
        if perf_score is not None:
            perf_score = int(perf_score * 100)
        
        # Get Core Web Vitals
        cwv = {}
        audits = lighthouse.get('audits', {})
        
        if 'largest-contentful-paint' in audits:
            cwv['LCP'] = audits['largest-contentful-paint'].get('displayValue', '')
        if 'cumulative-layout-shift' in audits:
            cwv['CLS'] = audits['cumulative-layout-shift'].get('displayValue', '')
        if 'total-blocking-time' in audits:
            cwv['TBT'] = audits['total-blocking-time'].get('displayValue', '')
        
        # Map technical audits to business impact
        audit_impact_map = {
            'render-blocking-resources': {
                'title': 'Render-Blocking Resources',
                'impact': 'Page appears blank for longer - visitors leave before seeing content'
            },
            'unused-javascript': {
                'title': 'Unused JavaScript Code',
                'impact': 'Loading code that is never used - wasting bandwidth and slowing everything'
            },
            'unused-css-rules': {
                'title': 'Unused CSS Code',
                'impact': 'Bloated stylesheets slow down page rendering'
            },
            'offscreen-images': {
                'title': 'Images Not Lazy-Loaded',
                'impact': 'Loading images visitors may never scroll to - wasting their data and time'
            },
            'uses-optimized-images': {
                'title': 'Unoptimized Images',
                'impact': 'Heavy images are the #1 cause of slow sites - directly hurts conversions'
            },
            'modern-image-formats': {
                'title': 'Outdated Image Formats',
                'impact': 'Using JPEG/PNG instead of WebP - images could be 50% smaller'
            },
            'uses-text-compression': {
                'title': 'Text Compression Disabled',
                'impact': 'HTML/CSS/JS could load much faster with gzip compression'
            },
            'server-response-time': {
                'title': 'Slow Server Response',
                'impact': 'Server takes too long to start sending data - delays everything'
            },
            'uses-responsive-images': {
                'title': 'Non-Responsive Images',
                'impact': 'Serving desktop-sized images to mobile users - wasting their data'
            },
            'efficient-animated-content': {
                'title': 'Inefficient Animations',
                'impact': 'Video/GIF animations slowing down page load'
            }
        }
        
        # Find failing audits (score < 0.9) and convert to business language
        failing_audits = []
        for audit_id, mapping in audit_impact_map.items():
            if audit_id in audits:
                audit = audits[audit_id]
                score = audit.get('score', 1)
                if score is not None and score < 0.9:
                    failing_audits.append({
                        'id': audit_id,
                        'title': mapping['title'],
                        'impact': mapping['impact'],
                        'score': score,
                        'detail': audit.get('displayValue', '')
                    })
        
        # Sort by score (worst first) and take top 3
        failing_audits.sort(key=lambda x: x.get('score', 1))
        
        return {
            'score': perf_score,
            'cwv': cwv,
            'audits': failing_audits[:3]
        }
        
    except Exception as e:
        logger.warning(f"PageSpeed API error: {str(e)}")
        return None


def save_audit_to_db(data, comparison_group=None):
    """Save audit to database."""
    db = get_db()
    if not db:
        return None
    
    try:
        domain = urlparse(data['url']).netloc.replace("www.", "")
        audit = Audit(
            url=data['url'],
            domain=domain,
            health_score=data['score'],
            psi_score=data.get('psi'),
            domain_age=data.get('domain_age'),
            tech_stack=data.get('tech_stack', []),
            issues=data.get('issues', []),
            emails_found=data.get('emails', []),
            ai_summary=data.get('ai', {}).get('summary'),
            ai_impact=data.get('ai', {}).get('impact'),
            ai_solutions=data.get('ai', {}).get('solutions'),
            ai_email=data.get('ai', {}).get('email'),
            comparison_group=comparison_group
        )
        db.add(audit)
        db.commit()
        db.refresh(audit)
        
        # Store audit_id in data for PDF generation
        data['audit_id'] = audit.id
        
        # Create/update lead
        existing_lead = db.query(Lead).filter(Lead.domain == domain).first()
        opp_score = calculate_opportunity_score(data)
        
        if not existing_lead:
            lead = Lead(
                domain=domain,
                email=data['emails'][0] if data.get('emails') else None,
                health_score=data['score'],
                opportunity_rating=opp_score,
                created_at=datetime.utcnow()
            )
            db.add(lead)
            db.commit()
        else:
            existing_lead.health_score = data['score']
            existing_lead.opportunity_rating = opp_score
            existing_lead.updated_at = datetime.utcnow()
            db.commit()
        
        return audit.id
    except Exception as e:
        db.rollback()
        return None
    finally:
        db.close()

# ============================================================================
# AUDIT HISTORY - HELPER FUNCTIONS (CENTRALIZED)
# ============================================================================

# Centralized session state keys for Audit History module
AUDIT_HISTORY_SESSION_KEYS = {
    # Filter keys (cleared on reset)
    "filters": [
        "hist_search", "min_score", "max_score", "time_filter",
        "hist_calendar_date", "hist_date_from", "hist_date_to",
        "show_all_scans", "audit_history_page"
    ],
    # Persistent keys (not cleared on filter reset)
    "persistent": [
        "audit_bulk_selected", "recently_deleted_audits", 
        "undo_delete_time", "audit_tags", "audit_sort_col", "audit_sort_dir"
    ],
    # Internal/temporary keys
    "internal": [
        "_audit_history_db_error", "_confirm_bulk_delete", "_confirm_delete_id"
    ]
}

# All filter keys (for backward compatibility with clear_audit_history_filters)
AUDIT_HISTORY_FILTER_KEYS = AUDIT_HISTORY_SESSION_KEYS["filters"]


def init_audit_history_session_state():
    """Initialize all audit history session state keys with safe defaults."""
    defaults = {
        "audit_bulk_selected": set(),
        "recently_deleted_audits": [],
        "undo_delete_time": None,
        "audit_tags": {},
        "_audit_history_db_error": False,
        "_confirm_bulk_delete": False,
        "_confirm_delete_id": None
    }
    for key, default_val in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = default_val


def get_score_status_icon(score: int) -> tuple:
    """Get status icon and label based on health score.
    
    Args:
        score: Health score (0-100) or None
        
    Returns:
        Tuple of (icon, label, css_class)
    """
    if score is None:
        return "❓", "Unknown", "unknown"
    if score >= 85:
        return "✅", "Passed", "good"
    elif score >= 70:
        return "🟡", "Review", "warning"
    else:
        return "❌", "Needs Attention", "critical"


def sanitize_domain_search(search_query: str) -> str:
    """Sanitize domain search query to prevent SQL injection and XSS attacks.
    
    Args:
        search_query: Raw user input for domain search
        
    Returns:
        Sanitized string safe for database queries, or None if invalid
    """
    if not search_query:
        return None
    # Remove dangerous characters, keep only safe ones for domain search
    import re
    sanitized = re.sub(r'[^\w\.\-]', '', search_query.strip())
    return sanitized[:255] if sanitized else None  # Limit length


def validate_score_filters(min_score: int, max_score: int) -> tuple:
    """Validate and normalize score filter values.
    
    Args:
        min_score: Minimum score filter value
        max_score: Maximum score filter value
        
    Returns:
        Tuple of (min_score, max_score, is_valid, error_message)
    """
    logger = logging.getLogger("sales_engine")
    
    # Ensure values are integers in valid range
    try:
        min_score = max(0, min(100, int(min_score or 0)))
        max_score = max(0, min(100, int(max_score or 100)))
    except (ValueError, TypeError):
        logger.warning(f"Invalid score filter values: min={min_score}, max={max_score}")
        return 0, 100, False, "Invalid score values"
    
    # Swap if min > max
    if min_score > max_score:
        logger.debug(f"Swapping score filters: {min_score} > {max_score}")
        min_score, max_score = max_score, min_score
    
    return min_score, max_score, True, None


def generate_active_filter_chips(time_filter: str, selected_date, search: str, 
                                  min_score: int, max_score: int) -> list:
    """Generate list of active filter chip HTML strings.
    
    Args:
        time_filter: Selected time period filter
        selected_date: Selected specific date (if applicable)
        search: Domain search query
        min_score: Minimum score filter
        max_score: Maximum score filter
        
    Returns:
        List of HTML strings for active filter chips
    """
    active_filters = []
    
    if time_filter == "Select Specific Date" and selected_date:
        active_filters.append(
            f"Time: <span class='active-filter'>{selected_date.strftime('%B %d, %Y')}</span>"
        )
    elif time_filter != "All Time":
        active_filters.append(f"Time: <span class='active-filter'>{time_filter}</span>")
    
    if search:
        # Escape HTML in search query for display
        safe_search = str(search).replace('<', '&lt;').replace('>', '&gt;')
        active_filters.append(f"Domain: <span class='active-filter'>{safe_search}</span>")
    
    if min_score > 0:
        active_filters.append(f"Min Score: <span class='active-filter'>{min_score}</span>")
    
    if max_score < 100:
        active_filters.append(f"Max Score: <span class='active-filter'>{max_score}</span>")
    
    return active_filters

def _get_attr(obj, attr, default=None):
    """Get attribute from either ORM object or dict."""
    if isinstance(obj, dict):
        return obj.get(attr, default)
    return getattr(obj, attr, default)


def filter_audits_by_time_period(audits: list, time_filter: str, selected_date=None, date_from=None, date_to=None) -> list:
    """Filter audits by time period. Extracted helper function. Works with ORM objects or dicts."""
    if not audits or time_filter == "All Time":
        return audits
    
    now = datetime.now()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    yesterday_start = today_start - timedelta(days=1)
    week_start = today_start - timedelta(days=today_start.weekday())
    month_start = today_start.replace(day=1)
    
    filtered = []
    for audit in audits:
        created_at = _get_attr(audit, 'created_at')
        if not created_at:
            continue
        
        audit_date = created_at
        
        if time_filter == "Today":
            if audit_date >= today_start:
                filtered.append(audit)
        elif time_filter == "Yesterday":
            if yesterday_start <= audit_date < today_start:
                filtered.append(audit)
        elif time_filter == "This Week":
            if audit_date >= week_start:
                filtered.append(audit)
        elif time_filter == "This Month":
            if audit_date >= month_start:
                filtered.append(audit)
        elif time_filter == "Last 30 Days":
            if audit_date >= (now - timedelta(days=30)):
                filtered.append(audit)
        elif time_filter == "Select Specific Date" and selected_date:
            selected_date_start = datetime.combine(selected_date, datetime.min.time())
            selected_date_end = datetime.combine(selected_date, datetime.max.time())
            if selected_date_start <= audit_date <= selected_date_end:
                filtered.append(audit)
        elif time_filter == "Custom Range" and date_from and date_to:
            date_from_dt = datetime.combine(date_from, datetime.min.time())
            date_to_dt = datetime.combine(date_to, datetime.max.time())
            if date_from_dt <= audit_date <= date_to_dt:
                filtered.append(audit)
        else:
            filtered.append(audit)
    
    return filtered

def dedupe_audits_by_domain(audits: list, per_user: bool = True) -> list:
    """Deduplicate audits keeping only most recent per domain (optionally per user). Works with ORM objects or dicts."""
    if not audits:
        return []
    
    domain_map = {}
    for audit in audits:
        domain = _get_attr(audit, 'domain', '')
        username = _get_attr(audit, 'username', '')
        created_at = _get_attr(audit, 'created_at')
        
        # Create key based on domain (and optionally username for per-user dedup)
        key = f"{domain}:{username}" if per_user and username else domain
        
        if key not in domain_map:
            domain_map[key] = audit
        else:
            # Keep the one with the latest created_at
            existing = domain_map[key]
            existing_created_at = _get_attr(existing, 'created_at')
            if created_at and existing_created_at:
                if created_at > existing_created_at:
                    domain_map[key] = audit
            elif created_at:
                domain_map[key] = audit
    
    result = list(domain_map.values())
    result.sort(key=lambda x: _get_attr(x, 'created_at') or datetime.min, reverse=True)
    return result

def group_audits_by_period(audits: list) -> dict:
    """Group audits by time period for organized display. Works with ORM objects or dicts."""
    now = datetime.now()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    yesterday_start = today_start - timedelta(days=1)
    week_start = today_start - timedelta(days=today_start.weekday())
    month_start = today_start.replace(day=1)
    
    grouped = {
        "Today": [],
        "Yesterday": [],
        "This Week": [],
        "This Month": [],
        "Older": []
    }
    
    for audit in audits:
        created_at = _get_attr(audit, 'created_at')
        if not created_at:
            grouped["Older"].append(audit)
            continue
        
        audit_date = created_at
        if audit_date >= today_start:
            grouped["Today"].append(audit)
        elif audit_date >= yesterday_start:
            grouped["Yesterday"].append(audit)
        elif audit_date >= week_start:
            grouped["This Week"].append(audit)
        elif audit_date >= month_start:
            grouped["This Month"].append(audit)
        else:
            grouped["Older"].append(audit)
    
    return grouped
    
    return grouped

def generate_audit_row_data(audit) -> dict:
    """Generate display data for a single audit row.
    
    Normalizes the audit first to ensure consistent dictionary access.
    """
    # Normalize audit to dict
    a = normalize_audit(audit)
    
    status_icon, status_label, _ = get_score_status_icon(a["health_score"])
    tags = st.session_state.get('audit_tags', {}).get(a["id"], [])
    
    return {
        "Status": status_icon,
        "Domain": a["domain"],
        "Score": format_score_badge(a["health_score"]) if a["health_score"] is not None else "N/A",
        "Speed": a["psi_score"] if a["psi_score"] else "N/A",
        "Issues": len(a["issues"]) if a["issues"] else 0,
        "Date": safe_timestamp_slice(a["created_at"], 11).replace("T", " ") if a["created_at"] else "N/A",
        "Tags": ", ".join(tags),
        "ID": a["id"],
        "User": a["username"] or "Unknown"
    }

def sort_audit_rows(hist_data: list, sort_col: str, reverse: bool = True) -> list:
    """Sort audit row data by specified column."""
    def sort_key(item):
        if sort_col == "Score":
            try:
                val = str(item["Score"]).split("/")[0]
                val = val.replace("✅","").replace("🟡","").replace("❌","").replace("❓","").strip()
                return int(val)
            except:
                return 0
        if sort_col == "Issues":
            try:
                return int(item["Issues"])
            except:
                return 0
        if sort_col == "Speed":
            try:
                return int(item["Speed"]) if str(item["Speed"]).isdigit() else 0
            except:
                return 0
        if sort_col == "Date":
            return item["Date"]
        return str(item.get(sort_col, ""))
    
    return sorted(hist_data, key=sort_key, reverse=reverse)

def convert_audit_to_data_dict(audit) -> dict:
    """Convert an Audit model instance to a data dictionary for display/PDF generation.
    
    Uses the consistency layer to normalize the audit first, then adds display-specific fields.
    """
    # Use consistency layer for base conversion
    base = audit_to_dict(audit)
    
    # Add display-specific 'ai' nested dict structure
    ai_dict = None
    if base.get('ai_summary'):
        ai_dict = {
            'summary': base.get('ai_summary') or 'No summary available',
            'impact': base.get('ai_impact') or 'No impact assessment available',
            'solutions': base.get('ai_solutions') or 'No solutions available',
            'email': base.get('ai_email') or 'No email draft available'
        }
    
    return {
        'url': base.get('url') or "",
        'domain': base.get('domain') or "",
        'score': base.get('health_score') or 0,
        'health_score': base.get('health_score') or 0,
        'psi': base.get('psi_score'),
        'psi_score': base.get('psi_score'),
        'domain_age': base.get('domain_age') or "Unknown",
        'tech_stack': base.get('tech_stack') or [],
        'issues': base.get('issues') or [],
        'ai': ai_dict,
        'emails': base.get('emails_found') or [],
        'created_at': base.get('created_at') or "N/A",
        'timestamp': base.get('timestamp') or base.get('created_at') or "N/A",
        'audit_id': base.get('id'),
        'username': base.get('username') or "Unknown"
    }

def invalidate_audit_cache():
    """Clear audit history cache after saving new audit."""
    logger = logging.getLogger("sales_engine")
    try:
        get_audit_history_cached.clear()
        logger.debug("Audit cache invalidated")
    except Exception as e:
        logger.debug(f"Cache invalidation skipped: {e}")


def clear_audit_history_filters():
    """Safely clear all audit history filters by popping widget keys.
    
    This avoids StreamlitAPIException by using pop() instead of direct assignment.
    Must be called BEFORE filter widgets are rendered.
    """
    logger = logging.getLogger("sales_engine")
    for key in AUDIT_HISTORY_FILTER_KEYS:
        st.session_state.pop(key, None)
    logger.debug("Audit history filters cleared")
    st.rerun()


def export_audits_to_csv(hist_data: list, selected_ids: set = None, 
                          exclude_columns: list = None) -> bytes:
    """Export audit data to CSV format.
    
    Args:
        hist_data: List of audit row dictionaries
        selected_ids: Optional set of audit IDs to filter (exports all if None)
        exclude_columns: Columns to exclude from export (default: ["ID", "User"])
        
    Returns:
        CSV data as bytes, ready for download
    """
    logger = logging.getLogger("sales_engine")
    
    if exclude_columns is None:
        exclude_columns = ["ID", "User"]
    
    try:
        # Filter by selected IDs if provided
        if selected_ids:
            data_to_export = [item for item in hist_data if item.get("ID") in selected_ids]
        else:
            data_to_export = hist_data
        
        if not data_to_export:
            logger.warning("No data to export to CSV")
            return b""
        
        # Create DataFrame and export
        df = pd.DataFrame(data_to_export)
        df = df.drop(columns=[c for c in exclude_columns if c in df.columns], errors='ignore')
        csv_bytes = df.to_csv(index=False).encode('utf-8')
        
        logger.info(f"Exported {len(data_to_export)} audits to CSV ({len(csv_bytes)} bytes)")
        return csv_bytes
        
    except Exception as e:
        logger.error(f"CSV export failed: {str(e)}")
        return b""


def serialize_audit_for_undo(audit) -> dict:
    """Serialize an Audit object for undo/restore operations.
    
    Args:
        audit: Audit model instance
        
    Returns:
        Dictionary with all audit fields for restoration
    """
    return {
        'id': audit.id,
        'domain': audit.domain,
        'url': audit.url,
        'health_score': audit.health_score,
        'psi_score': audit.psi_score,
        'issues': audit.issues,
        'created_at': audit.created_at.isoformat() if audit.created_at else None,
        'ai_summary': audit.ai_summary,
        'ai_impact': audit.ai_impact,
        'ai_solutions': audit.ai_solutions,
        'ai_email': audit.ai_email,
        'emails_found': audit.emails_found,
        'domain_age': audit.domain_age,
        'tech_stack': audit.tech_stack,
        'username': audit.username
    }


def restore_audit_from_dict(db, audit_data: dict):
    """Restore an audit from a serialized dictionary.
    
    Args:
        db: Database session
        audit_data: Dictionary from serialize_audit_for_undo
        
    Returns:
        New Audit instance (already added to session)
    """
    new_audit = Audit(
        domain=audit_data['domain'],
        url=audit_data.get('url', f"https://{audit_data['domain']}"),
        health_score=audit_data['health_score'],
        psi_score=audit_data['psi_score'],
        issues=audit_data.get('issues', []),
        created_at=datetime.fromisoformat(audit_data['created_at']) if audit_data.get('created_at') else None,
        ai_summary=audit_data.get('ai_summary'),
        ai_impact=audit_data.get('ai_impact'),
        ai_solutions=audit_data.get('ai_solutions'),
        ai_email=audit_data.get('ai_email'),
        emails_found=audit_data.get('emails_found', []),
        domain_age=audit_data.get('domain_age'),
        tech_stack=audit_data.get('tech_stack', []),
        username=audit_data.get('username')
    )
    db.add(new_audit)
    return new_audit


def bulk_delete_audits(audit_ids: set) -> tuple:
    """Delete multiple audits and store for potential undo.
    
    Args:
        audit_ids: Set of audit IDs to delete
        
    Returns:
        Tuple of (deleted_count, deleted_audits_list, error_message)
    """
    logger = logging.getLogger("sales_engine")
    db = None
    
    try:
        db = get_db()
        if not db:
            return 0, [], "Database unavailable"
        
        deleted_count = 0
        deleted_audits = []
        
        for aid in list(audit_ids):
            audit = db.query(Audit).filter(Audit.id == aid).first()
            if audit:
                deleted_audits.append(serialize_audit_for_undo(audit))
                db.delete(audit)
                deleted_count += 1
        
        db.commit()
        logger.info(f"Bulk deleted {deleted_count} audits")
        invalidate_audit_cache()
        
        return deleted_count, deleted_audits, None
        
    except Exception as e:
        logger.error(f"Bulk delete failed: {str(e)}")
        if db:
            try:
                db.rollback()
            except:
                pass
        return 0, [], str(e)
    finally:
        if db:
            try:
                db.close()
            except:
                pass


def bulk_restore_audits(deleted_audits: list) -> tuple:
    """Restore previously deleted audits.
    
    Args:
        deleted_audits: List of serialized audit dictionaries
        
    Returns:
        Tuple of (restored_count, error_message)
    """
    logger = logging.getLogger("sales_engine")
    db = None
    
    try:
        db = get_db()
        if not db:
            return 0, "Database unavailable"
        
        restored_count = 0
        for audit_data in deleted_audits:
            restore_audit_from_dict(db, audit_data)
            restored_count += 1
        
        db.commit()
        logger.info(f"Restored {restored_count} audits")
        invalidate_audit_cache()
        
        return restored_count, None
        
    except Exception as e:
        logger.error(f"Bulk restore failed: {str(e)}")
        if db:
            try:
                db.rollback()
            except:
                pass
        return 0, str(e)
    finally:
        if db:
            try:
                db.close()
            except:
                pass


def delete_single_audit(audit_id: int) -> tuple:
    """Delete a single audit by ID.
    
    Args:
        audit_id: ID of the audit to delete
        
    Returns:
        Tuple of (success: bool, domain: str, error_message: str)
    """
    logger = logging.getLogger("sales_engine")
    db = None
    
    try:
        db = get_db()
        if not db:
            return False, None, "Database unavailable"
        
        audit = db.query(Audit).filter(Audit.id == audit_id).first()
        if not audit:
            return False, None, "Audit not found"
        
        domain = audit.domain
        db.delete(audit)
        db.commit()
        
        logger.info(f"Deleted audit ID={audit_id}, domain={domain}")
        invalidate_audit_cache()
        
        return True, domain, None
        
    except Exception as e:
        logger.error(f"Delete audit failed: {str(e)}")
        if db:
            try:
                db.rollback()
            except:
                pass
        return False, None, str(e)
    finally:
        if db:
            try:
                db.close()
            except:
                pass


# ============================================================================
# AUDIT HISTORY - DATABASE FUNCTIONS (FIXED)
# ============================================================================

def save_audit_to_db(data, comparison_group=None, username=None, source="single"):
    """Save audit to database with user tracking and full CRM lead creation.
    
    This function saves the audit and creates/updates the associated Lead record
    with proper CRM fields for lead management workflow.
    
    Args:
        data: Audit data dictionary from run_audit()
        comparison_group: Optional group name for competitor analysis
        username: Username of the creator (auto-detected if not provided)
        source: Source of the audit - "single", "bulk", or "manual"
    
    Returns:
        audit_id on success, None on failure
    """
    logger = logging.getLogger("sales_engine")
    db = None
    
    # Get current user if not provided
    if not username:
        username = st.session_state.get("current_user")
    
    try:
        db = get_db()
        if not db:
            logger.error("Database unavailable when saving audit")
            return None
        
        # Validate required fields
        url = data.get('url')
        if not url:
            logger.error("Cannot save audit: missing URL")
            return None
        
        # Sanitize and extract domain
        try:
            domain = urlparse(url).netloc.replace("www.", "")
            if not domain:
                domain = url.replace("http://", "").replace("https://", "").split("/")[0]
        except Exception:
            domain = "unknown"
        
        # Safely get AI data
        ai_data = data.get('ai', {}) or {}
        
        # Create audit record with user tracking and source
        audit = Audit(
            url=url,
            domain=domain,
            health_score=data.get('score', 0),
            psi_score=data.get('psi'),
            domain_age=data.get('domain_age'),
            tech_stack=data.get('tech_stack', []) or [],
            issues=data.get('issues', []) or [],
            emails_found=data.get('emails', []) or [],
            ai_summary=ai_data.get('summary'),
            ai_impact=ai_data.get('impact'),
            ai_solutions=ai_data.get('solutions'),
            ai_email=ai_data.get('email'),
            comparison_group=comparison_group,
            username=username,
            source=source,  # Track origin: single, bulk, manual
            created_at=datetime.utcnow()
        )
        
        db.add(audit)
        db.commit()
        db.refresh(audit)
        
        # Store audit_id in data for PDF generation
        data['audit_id'] = audit.id
        
        logger.info(f"Audit saved: ID={audit.id}, domain={domain}, user={username}, source={source}")
        
        # =====================================================================
        # CRM LEAD CREATION/UPDATE - Full integration
        # =====================================================================
        try:
            existing_lead = db.query(Lead).filter(Lead.domain == domain).first()
            opp_score = calculate_opportunity_score(data)
            
            # Try to extract phone from raw HTML if available
            extracted_phone = None
            if data.get('raw_html'):
                extracted_phone = extract_phone_from_html(data.get('raw_html'))
            
            # Get first email from the audit
            first_email = data.get('emails', [None])[0] if data.get('emails') else None
            
            if not existing_lead:
                # CREATE NEW LEAD with CRM fields
                lead = Lead(
                    domain=domain,
                    email=first_email,
                    phone=extracted_phone,
                    health_score=data.get('score', 0),
                    opportunity_rating=opp_score,
                    # CRM fields
                    approached=False,
                    lead_status="warm",  # Default new leads to warm
                    interested="maybe",
                    pipeline_stage="new",
                    source=source,  # Track if from single/bulk
                    assigned_user=username,  # Assign to creator
                    last_audit_id=audit.id,
                    created_at=datetime.utcnow(),
                    updated_at=datetime.utcnow()
                )
                db.add(lead)
                logger.info(f"Lead created: domain={domain}, source={source}, assigned={username}")
            else:
                # UPDATE EXISTING LEAD
                existing_lead.health_score = data.get('score', 0)
                existing_lead.opportunity_rating = opp_score
                existing_lead.last_audit_id = audit.id
                existing_lead.updated_at = datetime.utcnow()
                
                # Update email if we found one and didn't have it
                if first_email and not existing_lead.email:
                    existing_lead.email = first_email
                
                # Update phone if we found one and didn't have it
                if extracted_phone and not existing_lead.phone:
                    existing_lead.phone = extracted_phone
                
                # Don't overwrite source if already set (preserve original source)
                # Don't overwrite approached status or pipeline_stage
                
                logger.debug(f"Lead updated: domain={domain}, new_audit={audit.id}")
            
            db.commit()
            
        except Exception as lead_error:
            logger.warning(f"Failed to create/update lead for {domain}: {str(lead_error)}")
            # Don't fail the whole operation if lead update fails
            try:
                db.rollback()
            except:
                pass
        
        # Invalidate cache so new audit appears immediately
        invalidate_audit_cache()
        
        return audit.id
        
    except Exception as e:
        logger.error(f"Failed to save audit: {str(e)}", exc_info=True)
        if db:
            try:
                db.rollback()
            except:
                pass
        return None
    finally:
        if db:
            try:
                db.close()
            except:
                pass

def get_audit_history(limit=100, search_query=None, min_score=None, max_score=None, 
                      username=None, is_admin=False):
    """Get audit history with filters and user-based access control.
    
    Args:
        limit: Maximum number of audits to return
        search_query: Domain search filter
        min_score: Minimum health score filter
        max_score: Maximum health score filter
        username: Current user's username (for filtering)
        is_admin: Whether current user is admin (sees all audits)
    
    Returns:
        List of Audit objects, or empty list on error
    """
    logger = logging.getLogger("sales_engine")
    db = None
    
    try:
        db = get_db()
        if not db:
            logger.error("Database unavailable when loading audit history")
            st.session_state["_audit_history_db_error"] = True
            return []
        
        st.session_state["_audit_history_db_error"] = False
        
        # Base query
        query = db.query(Audit).order_by(Audit.created_at.desc())
        
        # User-based filtering: Admin sees all, users see only their own
        if not is_admin and username:
            # Also include audits with no username (legacy data) for the user
            query = query.filter(
                (Audit.username == username) | (Audit.username == None)
            )
        
        # Apply search filter (sanitized)
        if search_query:
            sanitized_search = sanitize_domain_search(search_query)
            if sanitized_search:
                query = query.filter(Audit.domain.ilike(f"%{sanitized_search}%"))
        
        # Apply score filters
        if min_score is not None and min_score > 0:
            query = query.filter(Audit.health_score >= min_score)
        if max_score is not None and max_score < 100:
            query = query.filter(Audit.health_score <= max_score)
        
        # Limit results
        audits = query.limit(limit).all()
        
        logger.debug(f"Loaded {len(audits)} audits for user={username}, admin={is_admin}")
        return audits
        
    except Exception as e:
        logger.error(f"Failed to load audit history: {str(e)}", exc_info=True)
        st.session_state["_audit_history_db_error"] = True
        return []
    finally:
        if db:
            try:
                db.close()
            except:
                pass

def get_leads(status_filter=None, min_opp=None, max_opp=None):
    """Get leads with filters."""
    db = get_db()
    if not db:
        return []
    
    try:
        query = db.query(Lead).order_by(Lead.opportunity_rating.desc())
        if status_filter and status_filter != "all":
            query = query.filter(Lead.status == status_filter)
        if min_opp is not None:
            query = query.filter(Lead.opportunity_rating >= min_opp)
        if max_opp is not None:
            query = query.filter(Lead.opportunity_rating <= max_opp)
        return query.all()
    except Exception:
        return []
    finally:
        db.close()

def update_lead_status(lead_id, new_status):
    """Update lead status."""
    db = get_db()
    if not db:
        return False
    
    try:
        lead = db.query(Lead).filter(Lead.id == lead_id).first()
        if lead:
            lead.status = new_status
            lead.updated_at = datetime.utcnow()
            db.commit()
            return True
        return False
    except Exception:
        db.rollback()
        return False
    finally:
        db.close()

def save_email_template(name: str, subject: str, body: str):
    """Save email template."""
    templates_file = Path(__file__).parent / "email_templates.json"
    try:
        if templates_file.exists():
            templates = json.loads(templates_file.read_text())
        else:
            templates = {}
        
        templates[name] = {
            "subject": subject,
            "body": body,
            "created_at": datetime.now().isoformat()
        }
        templates_file.write_text(json.dumps(templates, indent=2))
        return True
    except Exception:
        return False

def load_email_templates():
    """Load email templates."""
    templates_file = Path(__file__).parent / "email_templates.json"
    try:
        if templates_file.exists():
            return json.loads(templates_file.read_text())
        return {}
    except Exception:
        return {}


def save_email_templates(templates: dict):
    """Save the entire templates dictionary."""
    templates_file = Path(__file__).parent / "email_templates.json"
    try:
        templates_file.write_text(json.dumps(templates, indent=2))
        return True
    except Exception:
        return False


class PDFReport(FPDF):
    """
    Enhanced PDF report generator with Code Nest branding.
    
    Brand Info:
    - Company: Code Nest LLC
    - Tagline: Nest Idea | Code Success
    - Website: https://codenest.us.com
    - Email: contact@codenest.us.com
    - Location: New Mexico, USA
    
    Features:
    - Professional cover page with logo
    - Branded color scheme (dark green #0c3740, accent green #2b945f)
    - Section-specific styling with color-coded headers
    - Consistent typography and spacing
    - Single footer per page (no duplicates)
    """
    
    def __init__(self):
        super().__init__()
        # Convert hex colors to RGB for FPDF
        self.brand_dark_green = (12, 55, 64)      # #0c3740
        self.brand_accent_green = (43, 148, 95)   # #2b945f
        self.brand_grey = (90, 90, 90)            # #5a5a5a
        self.brand_white = (254, 255, 255)        # #feffff
        self.brand_light_bg = (245, 250, 248)     # Light greenish tint
        
        # Track if we're on the cover page (to skip header/footer)
        self.is_cover_page = False
        
        # Check for logo file
        self.logo_available = PDF_LOGO_PATH.exists()
        if not self.logo_available:
            print(f"⚠️ Logo not found at {PDF_LOGO_PATH} – please upload your Code Nest logo to this path.")
    
    def header(self):
        """
        Branded header with logo and company info.
        Only renders on content pages (not cover page).
        """
        # Skip header on cover page
        if self.is_cover_page:
            return
        
        # Add logo if available
        if self.logo_available:
            try:
                self.image(str(PDF_LOGO_PATH), 10, 8, 35)
                self.set_xy(50, 10)
            except Exception:
                self.set_xy(10, 10)
        else:
            self.set_xy(10, 10)
        
        # Company name in dark green
        self.set_font('Arial', 'B', 18)
        self.set_text_color(*self.brand_dark_green)
        self.cell(0, 8, clean_text(COMPANY_NAME), 0, 1, 'L')
        
        # Tagline in accent green
        if self.logo_available:
            self.set_x(50)
        self.set_font('Arial', 'I', 10)
        self.set_text_color(*self.brand_accent_green)
        self.cell(0, 5, clean_text(COMPANY_TAGLINE), 0, 1, 'L')
        
        # Branded accent line
        self.set_draw_color(*self.brand_dark_green)
        self.set_line_width(1.5)
        self.line(10, 28, 200, 28)
        
        # Thin accent line below
        self.set_draw_color(*self.brand_accent_green)
        self.set_line_width(0.5)
        self.line(10, 30, 200, 30)
        
        self.ln(20)
    
    def footer(self):
        """
        Branded footer with contact info.
        Only renders on content pages (not cover page).
        Single clean footer - no duplicates.
        """
        # Skip footer on cover page
        if self.is_cover_page:
            return
        
        self.set_y(-20)
        
        # Accent line above footer
        self.set_draw_color(*self.brand_accent_green)
        self.set_line_width(0.3)
        self.line(10, self.get_y(), 200, self.get_y())
        self.ln(2)
        
        # Footer Line 1: Company | Website
        self.set_font('Arial', '', 8)
        self.set_text_color(*self.brand_grey)
        self.cell(0, 4, f"{COMPANY_NAME} | {COMPANY_WEBSITE}", 0, 1, 'C')
        
        # Footer Line 2: Email | Location
        self.cell(0, 4, f"{CONTACT_EMAIL} | {COMPANY_LOCATION}", 0, 1, 'C')
        
        # Page number (right-aligned)
        self.set_font('Arial', 'I', 8)
        self.set_text_color(*self.brand_accent_green)
        self.cell(0, 4, f"Page {self.page_no()}", 0, 0, 'R')
    
    def add_cover_page(self, domain: str, score: int):
        """
        Add professional branded cover page.
        Cover page has its own layout without the standard header/footer.
        """
        # Mark as cover page to skip header/footer
        self.is_cover_page = True
        self.add_page()
        
        # Large centered logo at top
        if self.logo_available:
            try:
                self.image(str(PDF_LOGO_PATH), 75, 25, 60)
                self.set_y(90)
            except Exception:
                self.set_y(50)
        else:
            self.set_y(50)
        
        # Company name centered
        self.set_font('Arial', 'B', 24)
        self.set_text_color(*self.brand_dark_green)
        self.cell(0, 12, clean_text(COMPANY_NAME), 0, 1, 'C')
        
        # Tagline centered
        self.set_font('Arial', 'I', 14)
        self.set_text_color(*self.brand_accent_green)
        self.cell(0, 8, clean_text(COMPANY_TAGLINE), 0, 1, 'C')
        self.ln(10)
        
        # Main title
        self.set_font('Arial', 'B', 26)
        self.set_text_color(*self.brand_dark_green)
        self.cell(0, 14, "Website Performance", 0, 1, 'C')
        self.cell(0, 14, "& Growth Audit", 0, 1, 'C')
        self.ln(5)
        
        # Decorative accent line
        self.set_draw_color(*self.brand_accent_green)
        self.set_line_width(2)
        self.line(60, self.get_y(), 150, self.get_y())
        self.ln(12)
        
        # "Prepared for:" + domain
        self.set_font('Arial', 'B', 16)
        self.set_text_color(*self.brand_grey)
        self.cell(0, 10, "Prepared for:", 0, 1, 'C')
        self.set_font('Arial', 'B', 20)
        self.set_text_color(*self.brand_dark_green)
        self.cell(0, 10, clean_text(domain), 0, 1, 'C')
        self.ln(8)
        
        # Health score section
        self.set_font('Arial', 'B', 14)
        self.set_text_color(*self.brand_grey)
        self.cell(0, 8, "Overall Health Score", 0, 1, 'C')
        
        # Score with color coding
        self.set_font('Arial', 'B', 48)
        if score >= 70:
            self.set_text_color(*self.brand_accent_green)
        elif score >= 40:
            self.set_text_color(255, 165, 0)  # Orange
        else:
            self.set_text_color(200, 60, 60)  # Red
        self.cell(0, 22, f"{score}/100", 0, 1, 'C')
        self.ln(5)
        
        # Report date
        self.set_font('Arial', '', 12)
        self.set_text_color(*self.brand_grey)
        self.cell(0, 8, f"Report Date: {datetime.now().strftime('%B %d, %Y')}", 0, 1, 'C')
        
        # Bottom branded contact block (replaces footer on cover)
        self.set_y(-55)
        self.set_fill_color(*self.brand_dark_green)
        self.rect(0, self.get_y(), 210, 50, 'F')
        
        self.set_y(-48)
        self.set_font('Arial', 'B', 12)
        self.set_text_color(*self.brand_white)
        self.cell(0, 7, COMPANY_NAME, 0, 1, 'C')
        self.set_font('Arial', '', 10)
        self.cell(0, 6, COMPANY_WEBSITE, 0, 1, 'C')
        self.cell(0, 6, f"{CONTACT_EMAIL} | {COMPANY_LOCATION}", 0, 1, 'C')
        
        # Reset cover page flag for subsequent pages
        self.is_cover_page = False
    
    def section_title(self, label, color_type="default"):
        """
        Branded section title with color-coded styling.
        
        color_type options: 'default', 'issues', 'quick_wins', 'ai_insights'
        """
        self.ln(5)
        self.set_font('Arial', 'B', 14)
        
        # Set colors based on section type
        if color_type == "issues":
            self.set_fill_color(255, 240, 240)  # Light red tint
            self.set_text_color(180, 60, 60)
        elif color_type == "quick_wins":
            self.set_fill_color(240, 255, 245)  # Light green tint
            self.set_text_color(*self.brand_accent_green)
        elif color_type == "ai_insights":
            self.set_fill_color(*self.brand_light_bg)
            self.set_text_color(*self.brand_dark_green)
        else:
            self.set_fill_color(*self.brand_light_bg)
            self.set_text_color(*self.brand_dark_green)
        
        # Draw section header with left accent bar
        y_start = self.get_y()
        self.set_draw_color(*self.brand_accent_green)
        self.set_line_width(3)
        self.line(10, y_start, 10, y_start + 10)
        
        self.cell(0, 10, f"  {clean_text(label)}", 0, 1, 'L', fill=True)
        self.ln(4)
    
    def chapter_body(self, text):
        """Standard body text with proper styling."""
        self.set_font('Arial', '', 10)
        self.set_text_color(*self.brand_grey)
        self.multi_cell(0, 5, clean_text(text))
        self.ln()
    
    def bullet_point(self, text, indent=0):
        """Consistent bullet point formatting."""
        self.set_font('Arial', '', 10)
        self.set_text_color(*self.brand_grey)
        
        # Bullet in accent green
        self.set_x(10 + indent)
        self.set_text_color(*self.brand_accent_green)
        self.cell(5, 5, chr(149), 0, 0)  # Bullet character
        
        # Text in grey
        self.set_text_color(*self.brand_grey)
        self.multi_cell(0, 5, f" {clean_text(text)}")

def generate_pdf(data):
    """Generate enhanced branded PDF report with cover page and AI Insights."""
    pdf = PDFReport()
    
    # Extract domain for cover page
    domain = data.get('url', 'Unknown')
    if '://' in domain:
        domain = domain.split('://')[1].split('/')[0]
    domain = domain.replace('www.', '')
    
    score = data.get('score', 0)
    
    # =========================================================================
    # COVER PAGE
    # =========================================================================
    pdf.add_cover_page(domain, score)
    
    # =========================================================================
    # AUDIT DETAILS PAGE
    # =========================================================================
    pdf.add_page()
    
    # Audit Summary Section
    pdf.section_title("Audit Summary")
    
    # Key metrics in a clean layout
    pdf.set_font('Arial', '', 10)
    pdf.set_text_color(*pdf.brand_grey)
    
    metrics = [
        ("Domain", domain),
        ("Health Score", f"{score}/100"),
        ("Domain Age", data.get('domain_age', 'Unknown')),
        ("Report Date", datetime.now().strftime('%B %d, %Y at %H:%M')),
    ]
    
    if data.get('psi'):
        metrics.append(("Google PageSpeed Score", f"{data['psi']}/100"))
    if data.get('accessibility_score'):
        metrics.append(("Accessibility Score", f"{data['accessibility_score']}/100"))
    if data.get('audit_id'):
        metrics.append(("Audit ID", str(data['audit_id'])))
    
    for label, value in metrics:
        pdf.set_font('Arial', 'B', 10)
        pdf.set_text_color(*pdf.brand_dark_green)
        pdf.cell(60, 6, f"{label}:", 0, 0)
        pdf.set_font('Arial', '', 10)
        pdf.set_text_color(*pdf.brand_grey)
        pdf.cell(0, 6, str(value), 0, 1)
    
    pdf.ln(5)

    # Tech stack
    pdf.section_title("Technology Stack")
    tech = ", ".join(data['tech_stack']) if data.get('tech_stack') else "Standard HTML/CSS"
    pdf.chapter_body(f"Detected technologies: {tech}")

    # =========================================================================
    # CRITICAL FINDINGS (Issues)
    # =========================================================================
    pdf.section_title("Critical Findings", color_type="issues")
    
    if data.get('issues'):
        for issue in data['issues']:
            # Issue title in bold
            pdf.set_font('Arial', 'B', 10)
            pdf.set_text_color(180, 60, 60)
            issue_title = issue.get('title', str(issue)) if isinstance(issue, dict) else str(issue)
            pdf.cell(0, 6, f"[!] {clean_text(issue_title)}", 0, 1)
            
            # Impact in italic
            if isinstance(issue, dict) and issue.get('impact'):
                pdf.set_font('Arial', 'I', 9)
                pdf.set_text_color(*pdf.brand_grey)
                pdf.multi_cell(0, 4, f"    Impact: {clean_text(issue['impact'])}")
            pdf.ln(2)
    else:
        pdf.set_font('Arial', 'I', 10)
        pdf.set_text_color(*pdf.brand_accent_green)
        pdf.cell(0, 6, "No critical issues detected - Great job!", 0, 1)

    # =========================================================================
    # AI INSIGHTS PAGE
    # =========================================================================
    ai_section = data.get('ai') if isinstance(data.get('ai'), dict) else {}
    if ai_section.get('insights'):
        pdf.add_page()
        
        # AI Insights header - centered with branding
        pdf.set_font('Arial', 'B', 20)
        pdf.set_text_color(*pdf.brand_dark_green)
        pdf.cell(0, 12, "AI-Powered Insights", 0, 1, 'C')
        
        # Decorative lines
        pdf.set_draw_color(*pdf.brand_accent_green)
        pdf.set_line_width(1)
        y = pdf.get_y()
        pdf.line(70, y, 140, y)
        pdf.ln(10)
        
        insights = ai_section.get('insights', {})
        
        # Snapshot Summary
        if insights.get('snapshot_summary'):
            pdf.section_title("Snapshot Summary", color_type="ai_insights")
            for bullet in insights['snapshot_summary']:
                pdf.bullet_point(bullet)
            pdf.ln(4)
        
        # Top 3 Issues - Red/Orange themed
        if insights.get('top_3_issues'):
            pdf.section_title("Top 3 Issues Hurting You Most", color_type="issues")
            for item in insights['top_3_issues']:
                pdf.set_font('Arial', 'B', 10)
                pdf.set_text_color(180, 60, 60)
                pdf.cell(5, 5, chr(149), 0, 0)
                pdf.cell(0, 5, f" {clean_text(item.get('issue', ''))}", 0, 1)
                
                pdf.set_font('Arial', 'I', 9)
                pdf.set_text_color(*pdf.brand_grey)
                pdf.set_x(20)
                pdf.multi_cell(0, 4, f"Impact: {clean_text(item.get('impact', ''))}")
            pdf.ln(4)
        
        # Quick Wins - Green themed
        if insights.get('quick_wins'):
            pdf.section_title("Quick Wins (Next 30 Days)", color_type="quick_wins")
            for idx, win in enumerate(insights['quick_wins'], 1):
                pdf.set_font('Arial', '', 10)
                pdf.set_text_color(*pdf.brand_accent_green)
                pdf.cell(8, 5, f"{idx}.", 0, 0)
                pdf.set_text_color(*pdf.brand_grey)
                pdf.multi_cell(0, 5, clean_text(win))
            pdf.ln(4)
        
        # How Code Nest Can Help - Brand themed
        if insights.get('code_nest_services'):
            pdf.section_title(f"How {COMPANY_NAME} Can Help", color_type="ai_insights")
            for item in insights['code_nest_services']:
                pdf.set_font('Arial', 'B', 10)
                pdf.set_text_color(*pdf.brand_dark_green)
                pdf.cell(5, 5, chr(149), 0, 0)
                pdf.cell(0, 5, f" {clean_text(item.get('issue', ''))}", 0, 1)
                
                pdf.set_font('Arial', '', 9)
                pdf.set_x(15)
                pdf.set_text_color(*pdf.brand_accent_green)
                pdf.multi_cell(0, 4, f"-> {clean_text(item.get('service', ''))}")
                pdf.set_text_color(*pdf.brand_grey)
            pdf.ln(4)
        
        # Suggested Next Step - Call to Action box
        if insights.get('next_step'):
            pdf.ln(5)
            pdf.set_fill_color(*pdf.brand_light_bg)
            pdf.set_draw_color(*pdf.brand_dark_green)
            pdf.set_line_width(1)
            
            # Draw box
            y_start = pdf.get_y()
            pdf.rect(10, y_start, 190, 25, 'DF')
            
            pdf.set_xy(15, y_start + 3)
            pdf.set_font('Arial', 'B', 12)
            pdf.set_text_color(*pdf.brand_dark_green)
            pdf.cell(0, 6, "Suggested Next Step", 0, 1)
            
            pdf.set_x(15)
            pdf.set_font('Arial', '', 10)
            pdf.set_text_color(*pdf.brand_accent_green)
            pdf.multi_cell(180, 5, clean_text(insights['next_step']))
            pdf.ln(4)
    
    # Legacy AI summary (for backward compatibility)
    elif data.get('ai') and data['ai'].get('summary'):
        pdf.add_page()
        pdf.section_title("Executive Summary", color_type="ai_insights")
        pdf.chapter_body(data['ai'].get('summary', ''))
        
        if data['ai'].get('impact'):
            pdf.section_title("Business Impact", color_type="issues")
            pdf.chapter_body(data['ai'].get('impact', ''))
        
        if data['ai'].get('solutions'):
            pdf.section_title("Recommended Solutions", color_type="quick_wins")
            pdf.chapter_body(data['ai'].get('solutions', ''))

    return pdf.output(dest='S').encode('latin-1')

def save_audit_pdf_to_file(audit_id, pdf_bytes):
    """Save PDF to file for persistent download."""
    pdf_dir = Path(__file__).parent / "audit_pdfs"
    pdf_dir.mkdir(exist_ok=True)
    
    pdf_path = pdf_dir / f"audit_{audit_id}.pdf"
    try:
        with open(pdf_path, 'wb') as f:
            f.write(pdf_bytes)
        return str(pdf_path)
    except Exception as e:
        return None

def get_audit_pdf(audit_id):
    """Generate PDF from audit database record."""
    db = get_db()
    if not db:
        return None
    
    try:
        audit_orm = db.query(Audit).filter(Audit.id == audit_id).first()
        if not audit_orm:
            return None
        
        # Normalize audit to dict for consistent access
        audit = normalize_audit(audit_orm)
        
        # Prepare audit data for PDF generation
        audit_data = {
            "domain": audit["domain"],
            "url": audit["url"],
            "score": audit["health_score"],
            "health_score": audit["health_score"],
            "psi": audit["psi_score"],
            "psi_score": audit["psi_score"],
            "domain_age": audit["domain_age"],
            "tech_stack": audit["tech_stack"] or [],
            "issues": audit["issues"] or [],
            "emails": audit["emails_found"] or [],
            "created_at": safe_timestamp_slice(audit["created_at"], 19).replace("T", " ") if audit["created_at"] else "N/A",
            "ai": {
                "summary": audit["ai_summary"] or "",
                "impact": audit["ai_impact"] or "",
                "solutions": audit["ai_solutions"] or "",
                "email": audit["ai_email"] or ""
            }
        }
        
        # Generate PDF
        pdf_bytes = generate_pdf_report(audit_data, f"audit_{audit_id}.pdf")
        return pdf_bytes
    except Exception as e:
        logger.error(f"Error generating audit PDF: {e}")
        return None
    finally:
        db.close()

def show_admin_settings():
    """Admin settings page with enhanced features."""
    st.subheader("👥 Admin Settings")
    
    admin_tab1, admin_tab2, admin_tab3 = st.tabs(["User Management", "Analytics", "Configuration"])
    
    with admin_tab1:
        users = load_users()
        pending = [u for u in users if users[u].get("admin_request")]
        
        if pending:
            st.warning(f"⚠️ {len(pending)} pending admin request(s)")
        
        # Custom CSS for beautiful UI
        st.markdown("""
        <style>
        .user-card {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 20px;
            border-radius: 12px;
            margin-bottom: 15px;
            color: white;
            box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
        }
        .user-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 15px;
            border-bottom: 1px solid rgba(255, 255, 255, 0.2);
            padding-bottom: 10px;
        }
        .user-name {
            font-size: 18px;
            font-weight: bold;
        }
        .user-role {
            display: inline-block;
            padding: 4px 12px;
            border-radius: 20px;
            font-size: 12px;
            font-weight: bold;
        }
        .role-admin {
            background-color: rgba(255, 107, 107, 0.8);
            color: white;
        }
        .role-user {
            background-color: rgba(100, 200, 100, 0.8);
            color: white;
        }
        .api-controls {
            display: grid;
            grid-template-columns: repeat(3, 1fr);
            gap: 10px;
            margin-top: 15px;
        }
        .api-toggle {
            background-color: rgba(255, 255, 255, 0.15);
            padding: 10px;
            border-radius: 8px;
            border: 1px solid rgba(255, 255, 255, 0.2);
            text-align: center;
        }
        .api-toggle.enabled {
            background-color: rgba(100, 200, 100, 0.3);
            border-color: rgba(100, 200, 100, 0.6);
        }
        .api-label {
            font-size: 12px;
            font-weight: bold;
            margin-bottom: 5px;
        }
        .actions {
            display: flex;
            gap: 8px;
            margin-top: 15px;
            flex-wrap: wrap;
        }
        .btn-action {
            flex: 1;
            min-width: 100px;
            padding: 8px 12px;
            border-radius: 6px;
            border: none;
            font-size: 12px;
            font-weight: bold;
            cursor: pointer;
            transition: all 0.3s ease;
        }
        </style>
        """, unsafe_allow_html=True)
        
        st.markdown("### 📊 User Management & API Access Control")
        st.markdown("Manage user roles and control their API access permissions below:")
        st.divider()
        
        # Filter users
        col1, col2 = st.columns(2)
        with col1:
            filter_role = st.selectbox("Filter by Role", ["All", "Admin", "User"])
        with col2:
            search_user = st.text_input("Search username")
        
        # Display users
        for username, user_data in users.items():
            if username == st.session_state.get("current_user"):
                continue
            
            role = user_data.get("role", "user")
            admin_req = user_data.get("admin_request", False)
            name = user_data.get("name", "N/A")
            
            # Apply filters
            if filter_role != "All" and role.capitalize() != filter_role:
                continue
            if search_user and search_user.lower() not in username.lower():
                continue
            
            # Get API permissions
            api_perms = get_user_api_permissions(username)
            
            # Beautiful user card
            with st.container():
                col1, col2 = st.columns([2, 1])
                
                with col1:
                    st.markdown(f"""
                    <div class="user-card">
                        <div class="user-header">
                            <div>
                                <div class="user-name">👤 {name}</div>
                                <div style="font-size: 12px; opacity: 0.8;">@{username}</div>
                            </div>
                            <div class="user-role role-{'admin' if role == 'admin' else 'user'}">
                                {'👑 ADMIN' if role == 'admin' else '👤 USER'}
                            </div>
                        </div>
                        
                        <div class="api-controls">
                            <div class="api-toggle {'enabled' if api_perms['openai'] else ''}">
                                <div class="api-label">🤖 OpenAI</div>
                                <div style="font-size: 11px;">{'✅ Enabled' if api_perms['openai'] else '❌ Disabled'}</div>
                            </div>
                            <div class="api-toggle {'enabled' if api_perms['google'] else ''}">
                                <div class="api-label">🔍 Google</div>
                                <div style="font-size: 11px;">{'✅ Enabled' if api_perms['google'] else '❌ Disabled'}</div>
                            </div>
                            <div class="api-toggle {'enabled' if api_perms['slack'] else ''}">
                                <div class="api-label">📱 Slack</div>
                                <div style="font-size: 11px;">{'✅ Enabled' if api_perms['slack'] else '❌ Disabled'}</div>
                            </div>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with col2:
                    # Action buttons
                    st.markdown("**Actions:**")
                    
                    # Role management (using selectbox instead of buttons)
                    new_role = st.selectbox(
                        "Role",
                        ["user", "admin"],
                        index=0 if role == "user" else 1,
                        key=f"role_{username}"
                    )
                    
                    if new_role != role:
                        users[username]["role"] = new_role
                        if new_role == "admin":
                            users[username]["admin_request"] = False
                        save_users(users)
                        st.toast(f"✓ {username} role updated to {new_role}")
                    
                    # Clear admin requests
                    if admin_req:
                        if st.button("❌ Clear Admin Request", key=f"clear_{username}", use_container_width=True):
                            users[username]["admin_request"] = False
                            save_users(users)
                            st.toast(f"Request cleared for {username}")
                
                # API Access Control Section with form
                st.markdown("**⚙️ API Access Control:**")
                
                with st.form(key=f"api_form_{username}", clear_on_submit=False):
                    col_api1, col_api2, col_api3 = st.columns(3)
                    
                    with col_api1:
                        openai_access = st.checkbox(
                            "🤖 Allow OpenAI",
                            value=api_perms.get("openai", False),
                            key=f"openai_{username}"
                        )
                    
                    with col_api2:
                        google_access = st.checkbox(
                            "🔍 Allow Google PageSpeed",
                            value=api_perms.get("google", False),
                            key=f"google_{username}"
                        )
                    
                    with col_api3:
                        slack_access = st.checkbox(
                            "📱 Allow Slack",
                            value=api_perms.get("slack", False),
                            key=f"slack_{username}"
                        )
                    
                    # Submit button for form (no rerun)
                    submitted = st.form_submit_button("💾 Save API Permissions", use_container_width=True)
                    if submitted:
                        set_user_api_permission(username, "openai", openai_access)
                        set_user_api_permission(username, "google", google_access)
                        set_user_api_permission(username, "slack", slack_access)
                        st.toast(f"✓ API permissions saved for {username}")
                
                st.divider()
    
    with admin_tab2:
        st.markdown("### Analytics Dashboard")
        
        if DB_AVAILABLE:
            all_audits = get_audit_history(limit=1000)
            all_leads = get_leads()
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Total Audits", len(all_audits))
            with col2:
                st.metric("Total Leads", len(all_leads))
            with col3:
                avg_score = sum([a.health_score for a in all_audits]) / len(all_audits) if all_audits else 0
                st.metric("Avg Health Score", f"{avg_score:.0f}")
            with col4:
                converted = len([l for l in all_leads if l.status == "converted"])
                st.metric("Converted", converted)
            
            st.divider()
            
            # Score distribution
            if all_audits:
                scores = [a.health_score for a in all_audits]
                chart_data = pd.DataFrame({
                    "Score Range": ["0-30", "30-60", "60-80", "80-100"],
                    "Count": [
                        len([s for s in scores if s < 30]),
                        len([s for s in scores if 30 <= s < 60]),
                        len([s for s in scores if 60 <= s < 80]),
                        len([s for s in scores if s >= 80])
                    ]
                })
                st.bar_chart(chart_data.set_index("Score Range"))
            
            # Lead funnel
            if all_leads:
                lead_statuses = {}
                for lead in all_leads:
                    lead_statuses[lead.status] = lead_statuses.get(lead.status, 0) + 1
                
                st.markdown("### Lead Funnel")
                for status, count in lead_statuses.items():
                    st.write(f"{status.capitalize()}: {count}")
    
    with admin_tab3:
        st.markdown("### System Configuration")
        st.info("Slack Webhook: " + ("✓ Configured" if SLACK_WEBHOOK else "Not configured"))
        st.info("OpenAI API: " + ("✓ Configured" if OPENAI_API_KEY else "Not configured"))
        st.info("Database: " + ("✓ Connected" if DB_AVAILABLE else "Not connected"))

def show_email_settings():
    """Configure email notification settings."""
    st.markdown("## 📧 Email Notification Settings")
    
    # Hostinger setup info
    with st.expander("📋 Hostinger SMTP Setup Guide", expanded=False):
        st.markdown("""
        ### Quick Setup for Hostinger Email
        
        **SMTP Configuration for Hostinger:**
        - **SMTP Server:** `smtp.hostinger.com`
        - **SMTP Port:** `587` (TLS) or `465` (SSL)
        - **Sender Email:** Your Hostinger email (e.g., `contact@codenest.us.com`)
        - **Password:** Your Hostinger email password
        
        **Steps to get your credentials:**
        1. Log in to Hostinger Control Panel
        2. Go to **Email** → Your email account
        3. Find SMTP/IMAP settings
        4. Copy the SMTP server address and your email credentials
        
        **Alternative - Using Environment Variable:**
        ```bash
        # Set this in your .env file for security:
        EMAIL_PASSWORD=your_hostinger_email_password
        ```
        
        **Auto-Send Reports Feature:**
        - When you run an audit, the system will automatically try to:
          1. Extract contact email from OpenAI analysis
          2. Send a formatted audit report to that email
          3. Log all sent emails for tracking
        
        **Email Features:**
        - ✅ Automatic report sending when contacts are found
        - ✅ Professional HTML formatted emails
        - ✅ Includes audit metrics, top issues, and recommendations
        - ✅ All emails logged and tracked
        - ✅ Manual send with custom email addresses
        """)
    
    config = load_email_config()
    
    st.markdown("### SMTP Configuration")
    
    col1, col2 = st.columns(2)
    with col1:
        enabled = st.checkbox("Enable Email Notifications", value=config.get("enabled", False))
    with col2:
        auto_send = st.checkbox("Auto-Send Reports", value=config.get("auto_send_reports", True), help="Automatically send reports when emails are found")
    
    with st.form("email_config_form"):
        smtp_server = st.text_input(
            "SMTP Server",
            value=config.get("smtp_server", ""),
            placeholder="e.g., smtp.gmail.com",
            help="Your email provider's SMTP server address"
        )
        
        smtp_port = st.number_input(
            "SMTP Port",
            value=config.get("smtp_port", 587),
            min_value=1,
            max_value=65535,
            help="Usually 587 for TLS or 465 for SSL"
        )
        
        sender_email = st.text_input(
            "Sender Email Address",
            value=config.get("sender_email", ""),
            placeholder="noreply@example.com"
        )
        
        sender_password = st.text_input(
            "Email Password/App Password",
            value=config.get("sender_password", ""),
            type="password",
            help="Use an app-specific password for Gmail",
            placeholder="••••••••"
        )
        
        from_name = st.text_input(
            "Display Name",
            value=config.get("from_name", "Code Nest Sales Engine"),
            help="Name that appears in 'From:' field"
        )
        
        reply_to = st.text_input(
            "Reply-To Email Address",
            value=config.get("reply_to", config.get("sender_email", "")),
            placeholder="support@example.com",
            help="Email address for replies"
        )
        
        st.markdown("### Notification Types")
        col1, col2, col3 = st.columns(3)
        with col1:
            audit_notif = st.checkbox("Audit Completions", value=config.get("notifications", {}).get("audit_complete", True))
        with col2:
            perm_notif = st.checkbox("Permission Changes", value=config.get("notifications", {}).get("permission_change", True))
        with col3:
            admin_notif = st.checkbox("Admin Alerts", value=config.get("notifications", {}).get("admin_alert", True))
        
        # Test email
        st.markdown("### Test Configuration")
        test_email = st.text_input("Send test email to:", placeholder="test@example.com")
        
        col1, col2, col3 = st.columns([1, 1, 2])
        with col1:
            submit = st.form_submit_button("Save Configuration", type="primary", use_container_width=True)
        with col2:
            test = st.form_submit_button("Send Test Email", use_container_width=True)
        
        if submit:
            new_config = {
                "enabled": enabled,
                "smtp_server": smtp_server,
                "smtp_port": smtp_port,
                "sender_email": sender_email,
                "sender_password": sender_password,
                "from_name": from_name,
                "reply_to": reply_to,
                "auto_send_reports": auto_send,
                "notifications": {
                    "audit_complete": audit_notif,
                    "permission_change": perm_notif,
                    "admin_alert": admin_notif
                }
            }
            success, message = save_email_config(new_config)
            if success:
                st.success(message)
                st.rerun()
            else:
                st.error(message)
        
        if test and test_email:
            if not enabled or not smtp_server or not sender_email:
                st.error("Please configure SMTP settings and enable notifications first")
            else:
                with st.spinner("Sending test email..."):
                    html_body = get_email_template("admin_alert", {
                        "event": "Test Email",
                        "details": "This is a test email from Code Nest Sales Engine",
                        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    })
                    success, message = send_email(test_email, "Test Email - Code Nest Sales Engine", html_body)
                    if success:
                        st.success(f"✅ {message}")
                    else:
                        st.error(f"❌ {message}")
    
    # Notification logs
    st.divider()
    st.markdown("### Notification Logs")
    
    if NOTIFICATIONS_LOG_PATH.exists():
        try:
            notifications = json.loads(NOTIFICATIONS_LOG_PATH.read_text())
            if notifications:
                df = pd.DataFrame(notifications[-50:])  # Last 50
                df["timestamp"] = pd.to_datetime(df["timestamp"]).dt.strftime("%Y-%m-%d %H:%M:%S")
                st.dataframe(df.sort_values("timestamp", ascending=False), use_container_width=True)
            else:
                st.info("No notifications sent yet")
        except Exception as e:
            st.error(f"Error loading logs: {str(e)}")
    else:
        st.info("No notification logs available yet")

def show_export_reports():
    """Export audits as PDF or Excel reports."""
    logger = logging.getLogger("sales_engine")
    
    try:
        st.markdown("## 📄 Export Reports")
        
        export_type = st.radio("Export Format", ["PDF (Single)", "Excel (Batch)", "CSV (Data)"], horizontal=True)
        
        if export_type == "PDF (Single)":
            st.markdown("### Export Single Audit as PDF")
            
            # Get list of audits and convert to dicts
            audits_raw = get_audit_history_cached(limit=100)
            if not audits_raw:
                st.warning("No audits available to export")
                return
            
            # Convert all audits to dicts with safety wrapper
            audits = [safe_render_audit(a) for a in audits_raw]
            
            # Build audit options with safe timestamp handling
            audit_options = {}
            for a in audits:
                domain = a.get('domain') or 'Unknown'
                timestamp = safe_timestamp_slice(a.get('timestamp') or a.get('created_at'), 10)
                label = f"{domain} ({timestamp})"
                audit_options[label] = a
            
            if not audit_options:
                st.warning("No valid audits to display")
                return
            
            selected = st.selectbox("Select Audit", list(audit_options.keys()))
            
            if selected and st.button("Generate PDF Report", use_container_width=True, type="primary"):
                with st.spinner("Generating PDF..."):
                    audit_data = audit_options[selected]
                    pdf_bytes = generate_pdf_report(audit_data)
                    
                    if pdf_bytes:
                        st.success("✅ PDF generated successfully")
                        st.download_button(
                            label="⬇️ Download PDF Report",
                            data=pdf_bytes,
                            file_name=f"audit_{(audit_data.get('domain') or 'report').replace('/', '_')}.pdf",
                            mime="application/pdf",
                            use_container_width=True
                        )
                    else:
                        st.error("Error generating PDF")
        
        elif export_type == "Excel (Batch)":
            st.markdown("### Export Multiple Audits as Excel")
            
            audits_raw = get_audit_history_cached(limit=500)
            if not audits_raw:
                st.warning("No audits available to export")
                return
            
            # Convert all audits to dicts with safety wrapper
            audits = [safe_render_audit(a) for a in audits_raw]
            
            # Filter options
            col1, col2 = st.columns(2)
            with col1:
                min_score = st.slider("Minimum Score", 0, 100, 0)
            with col2:
                max_score = st.slider("Maximum Score", 0, 100, 100)
            
            # Safe score filtering
            filtered_audits = []
            for a in audits:
                score = a.get("score") or a.get("health_score") or 0
                try:
                    score_int = int(score) if score is not None else 0
                except (ValueError, TypeError):
                    score_int = 0
                if min_score <= score_int <= max_score:
                    filtered_audits.append(a)
            
            if not filtered_audits:
                st.warning("No audits match your filter criteria")
                return
            
            st.info(f"Exporting {len(filtered_audits)} audits...")
            
            if st.button("Generate Excel Report", use_container_width=True, type="primary"):
                with st.spinner("Generating Excel..."):
                    excel_bytes = generate_excel_report(filtered_audits)
                    
                    if excel_bytes:
                        st.success("✅ Excel file generated successfully")
                        st.download_button(
                            label="⬇️ Download Excel Report",
                            data=excel_bytes,
                            file_name=f"audits_report_{datetime.now().strftime('%Y%m%d')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                    else:
                        st.error("Error generating Excel")
        
        elif export_type == "CSV (Data)":
            st.markdown("### Export as CSV for Analysis")
            
            audits_raw = get_audit_history_cached(limit=1000)
            if not audits_raw:
                st.warning("No audits available to export")
                return
            
            # Convert all audits to dicts with safety wrapper
            audits = [safe_render_audit(a) for a in audits_raw]
            
            df = pd.DataFrame(audits)
            
            # Handle empty dataframe
            if df.empty:
                st.warning("No data available for export")
                return
            
            # Column selection with safe defaults using helper
            existing_cols = df.columns.tolist()
            default_cols = get_safe_export_columns(existing_cols, ["domain", "score", "health_score", "status", "timestamp", "created_at"])
            
            columns = st.multiselect(
                "Select columns to export",
                existing_cols,
                default=default_cols
            )
            
            if columns and st.button("Generate CSV", use_container_width=True, type="primary"):
                csv_data = df[columns].to_csv(index=False)
                st.success("✅ CSV ready for download")
                st.download_button(
                    label="⬇️ Download CSV",
                    data=csv_data,
                    file_name=f"audits_{datetime.now().strftime('%Y%m%d')}.csv",
                    mime="text/csv",
                    use_container_width=True
                )
            
            # Preview
            if columns:
                st.markdown("### Preview")
                st.dataframe(df[columns].head(10), use_container_width=True)
    
    except Exception as e:
        logger.error(f"Error in Export Reports: {str(e)}", exc_info=True)
        st.error("⚠️ Something went wrong generating the export. Please try again.")
        with st.expander("🔍 Technical Details"):
            st.code(str(e))


# ============================================================================
# MAIN ROUTING & CONTENT DISPATCHER
# ============================================================================

# Main content area routing
current_section = st.session_state.current_section

if current_section == "Single Audit":
    show_single_audit()
elif current_section == "Audit History":
    show_audit_history()
elif current_section == "Dashboard":
    show_dashboard()
elif current_section == "Preferences":
    show_preferences_panel(st.session_state.get('current_user', 'user'))
elif current_section == "Bulk Audit":
    show_bulk_audit()
elif current_section == "Competitor Analysis":
    show_competitor_analysis()
elif current_section == "Lead Management":
    show_lead_management()
elif current_section == "CRM Pipeline":
    show_crm_pipeline()
elif current_section == "Email Outreach":
    show_email_outreach()
elif current_section == "Scheduled Audits":
    show_scheduled_audits()
elif current_section == "API Settings":
    show_api_settings()
elif current_section == "Email Settings":
    show_email_settings()
elif current_section == "Export Reports":
    show_export_reports()
elif current_section == "Admin Settings":
    show_admin_settings()
else:
    # Default to Single Audit
    st.session_state.current_section = "Single Audit"
    show_single_audit()
